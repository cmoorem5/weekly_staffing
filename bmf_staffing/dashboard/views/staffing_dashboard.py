"""Staffing dashboard page and CSV/XLSX exports."""

import csv
import io
import json
from collections import defaultdict
from datetime import date, timedelta
from typing import cast

from django.contrib import messages
from django.http import Http404, HttpResponse
from django.shortcuts import redirect, render
from openpyxl import Workbook
from sqlalchemy import func
from staffing_tool.db import session_scope
from staffing_tool.fiscal_year import (
    fy_end_date,
    fy_label_year,
    fy_week1_sunday_containing,
    pay_periods_for_fy,
)
from staffing_tool.time_buckets import bucket_label, buckets_for_range
from staffing_tool.leave_grid import EXCEPTION_COL_BREAKDOWN_KEYS
from staffing_tool.metrics import compute_period_rollups, compute_week_metrics
from staffing_tool.models import (
    BaseConfig,
    WeeklyBaseCoverage,
    WeeklyLeaveDetail,
    WeeklyManagerShift,
    WeeklyStaffing,
)

from .dashboard_filters import (
    fy_choice_rows,
    last_closed_pay_period_end_for_fy,
    parse_date_param,
    parse_fy_week1_from_request,
    serialize_filters_query,
    serialize_filters_query_from_parts,
)
from .helpers import DB_PATH, FY_AND_PAY_PERIOD_POLICY_NOTE, _ensure_db, _utc_now_iso


def _parse_multi_param(request, key: str) -> list[str]:
    """
    Read a multi-select query param.
    Supports repeated params (?k=a&k=b) and comma-separated (?k=a,b).
    """
    vals = []
    if hasattr(request.GET, "getlist"):
        vals.extend([v for v in request.GET.getlist(key) if v is not None])
    raw = (request.GET.get(key) or "").strip()
    if raw:
        vals.extend(raw.split(","))
    out: list[str] = []
    for v in vals:
        s = str(v).strip()
        if not s:
            continue
        out.append(s)
    # de-dupe preserving order
    seen = set()
    deduped: list[str] = []
    for v in out:
        if v in seen:
            continue
        seen.add(v)
        deduped.append(v)
    return deduped


def _exception_type_key_map() -> dict[str, list[str]]:
    """
    UI exception types -> WeeklyLeaveDetail.leave_type keys.
    Includes an ALL pseudo-option.
    """
    # Keep names aligned with the exception grid columns where possible.
    base = {k: list(v) for k, v in EXCEPTION_COL_BREAKDOWN_KEYS.items()}
    base["ALL"] = sorted({t for keys in base.values() for t in keys})
    return base


def _normalize_exception_types(selected: list[str]) -> list[str]:
    mapping = _exception_type_key_map()
    normalized: list[str] = []
    for s in selected:
        key = s.strip().upper()
        if key == "SL":
            key = "SICK"
        if key == "SICK/SL":
            key = "SICK"
        if key in mapping:
            normalized.append(key)
    # de-dupe
    seen = set()
    out: list[str] = []
    for v in normalized:
        if v in seen:
            continue
        seen.add(v)
        out.append(v)
    return out


def _exc_types_label(exc_types: list[str]) -> str:
    if not exc_types:
        return ""
    if "ALL" in exc_types:
        return "All exceptions"
    if len(exc_types) == 1:
        return exc_types[0]
    return ", ".join(exc_types)


# Exception chart/table series order (matches leave grid columns + Other for unmapped).
EXC_BREAKDOWN_SERIES_ORDER = ["LT", "LOA", "SICK", "AT", "JURY", "BREV", "Other"]


def _exc_breakdown_groups() -> dict[str, list[str]]:
    """
    Exception breakdown series for the staffing dashboard chart/export.

    Keys are display series names; values are WeeklyLeaveDetail.leave_type keys that
    roll up into that series. Any leave_type not covered here is treated as "Other".
    """
    return {
        "LT": EXCEPTION_COL_BREAKDOWN_KEYS["LT"],
        "LOA": EXCEPTION_COL_BREAKDOWN_KEYS["LOA"],
        "SICK": EXCEPTION_COL_BREAKDOWN_KEYS["SICK"],
        "AT": EXCEPTION_COL_BREAKDOWN_KEYS["AT"],
        "JURY": EXCEPTION_COL_BREAKDOWN_KEYS["JURY"],
        "BREV": EXCEPTION_COL_BREAKDOWN_KEYS["BREV"],
    }


def _exc_group_for_leave_type(leave_type: str) -> str:
    lt = (leave_type or "").strip().upper()
    groups = _exc_breakdown_groups()
    for group, keys in groups.items():
        if lt in {k.upper() for k in keys}:
            return group
    return "Other"


def _build_staffing_dashboard_context(request) -> dict[str, object]:
    """Build context for staffing dashboard and export endpoints (single source)."""
    _ensure_db()
    if not DB_PATH:
        raise Http404("Database is not configured (STAFFING_DB_PATH).")

    today = date.today()
    fy_start = parse_fy_week1_from_request(request, today)
    fy_end = fy_end_date(fy_start)
    fy_label = fy_label_year(fy_start)
    fy_choices = fy_choice_rows(
        fy_label_year(fy_week1_sunday_containing(today))
    )

    granularity = (request.GET.get("granularity") or "pay_period").strip().lower()
    if granularity not in {"quarter", "month", "pay_period"}:
        granularity = "pay_period"

    # FY-to-date ends at last closed pay period within the selected FY when that FY is current.
    last_closed_in_fy = last_closed_pay_period_end_for_fy(today, fy_start)
    is_current_fy = fy_start == fy_week1_sunday_containing(today)
    default_end = last_closed_in_fy if is_current_fy else fy_end
    default_start = fy_start

    date_start = parse_date_param(request.GET.get("date_start", ""), default_start)
    date_end = parse_date_param(request.GET.get("date_end", ""), default_end)
    date_start = max(date_start, fy_start)
    date_end = min(date_end, fy_end)
    if date_start > date_end:
        date_start, date_end = default_start, default_end

    buckets = buckets_for_range(granularity, date_start, date_end)
    labels: list[str] = []
    staffing_rate_series: list[float] = []
    ot_dependency_series: list[float] = []
    shift_exception_series: list[float] = []
    system_rw_series: list[float] = []
    system_gr_series: list[float] = []
    table_rows: list[dict[str, object]] = []

    # Manager line shifts (counts) + exceptions (counts)
    manager_line_shifts_total_series: list[int] = []
    manager_line_shifts_breakdown_series: dict[str, list[int]] = {}
    manager_line_shifts_breakdown_order: list[str] = []
    manager_line_shifts_table: list[dict[str, object]] = []
    exc_total_series: list[int] = []
    exc_breakdown_series: dict[str, list[int]] = {
        k: [] for k in EXC_BREAKDOWN_SERIES_ORDER
    }
    exc_table_rows: list[dict[str, object]] = []

    # Data quality panel: expected week_start Sundays that fall in the selected date range.
    expected_week_starts: list[str] = []
    first_sun = date_start + timedelta(days=(6 - date_start.weekday()) % 7)
    cur = first_sun
    while cur <= date_end:
        expected_week_starts.append(cur.isoformat())
        cur += timedelta(days=7)

    # Preset ranges for dashboard buttons (no DB access required).
    periods = pay_periods_for_fy(fy_start)
    end_anchor = default_end
    closed = [p for p in periods if p.end <= end_anchor]
    last6 = closed[-6:] if len(closed) >= 6 else closed
    last6_start = last6[0].start if last6 else fy_start
    last6_end = last6[-1].end if last6 else end_anchor
    last12w_start = max(fy_start, end_anchor - timedelta(days=83))
    preset_links = {
        "fy_ytd": {
            "label": "FY YTD (last closed PP)",
            "qs": serialize_filters_query_from_parts(
                {
                    "fy": str(fy_label),
                    "granularity": "pay_period",
                    "date_start": fy_start.isoformat(),
                    "date_end": end_anchor.isoformat(),
                }
            ),
        },
        "last_6_pp": {
            "label": "Last 6 pay periods",
            "qs": serialize_filters_query_from_parts(
                {
                    "fy": str(fy_label),
                    "granularity": "pay_period",
                    "date_start": last6_start.isoformat(),
                    "date_end": last6_end.isoformat(),
                }
            ),
        },
        "last_12_weeks": {
            "label": "Last 12 weeks",
            "qs": serialize_filters_query_from_parts(
                {
                    "fy": str(fy_label),
                    "granularity": "month",
                    "date_start": last12w_start.isoformat(),
                    "date_end": end_anchor.isoformat(),
                }
            ),
        },
        "full_fy": {
            "label": "Full FY",
            "qs": serialize_filters_query_from_parts(
                {
                    "fy": str(fy_label),
                    "granularity": "quarter",
                    "date_start": fy_start.isoformat(),
                    "date_end": fy_end.isoformat(),
                }
            ),
        },
    }

    with session_scope(DB_PATH) as session:
        weeks = (
            session.query(WeeklyStaffing)
            .filter(WeeklyStaffing.week_start >= date_start.isoformat())
            .filter(WeeklyStaffing.week_start <= date_end.isoformat())
            .order_by(WeeklyStaffing.week_start.asc())
            .all()
        )
        latest_week_start = weeks[-1].week_start if weeks else None
        latest_updated_at = getattr(weeks[-1], "updated_at", None) if weeks else None

        # Best-effort "last schedule import applied": use the most recently updated row that
        # looks like it came from the schedule import flow.
        imported = (
            session.query(WeeklyStaffing)
            .filter(
                (WeeklyStaffing.entered_by == "import")
                | (WeeklyStaffing.notes.ilike("imported from schedule%"))
            )
            .order_by(WeeklyStaffing.updated_at.desc(), WeeklyStaffing.week_start.desc())
            .first()
        )
        last_import_week_start = getattr(imported, "week_start", None) if imported else None
        last_import_updated_at = getattr(imported, "updated_at", None) if imported else None

        # Data quality: missing week_start rows and missing schedule-import markers.
        staff_by_week: dict[str, tuple[str | None, str | None]] = {}
        if expected_week_starts:
            staff_rows = (
                session.query(
                    WeeklyStaffing.week_start,
                    WeeklyStaffing.entered_by,
                    WeeklyStaffing.notes,
                )
                .filter(WeeklyStaffing.week_start.in_(expected_week_starts))
                .all()
            )
            staff_by_week = {
                str(ws): (cast(str | None, entered_by), cast(str | None, notes))
                for ws, entered_by, notes in staff_rows
            }
            mgr_weeks = (
                session.query(WeeklyManagerShift.week_start)
                .filter(WeeklyManagerShift.week_start.in_(expected_week_starts))
                .distinct()
                .all()
            )
            mgr_week_set = {str(ws) for (ws,) in mgr_weeks}
        else:
            mgr_week_set = set()

        data_quality_rows: list[dict[str, str]] = []
        for ws in expected_week_starts:
            if ws not in staff_by_week:
                data_quality_rows.append(
                    {"week_start": ws, "issue": "Weekly staffing row missing"}
                )
                continue
            entered_by, notes = staff_by_week.get(ws, (None, None))
            notes_s = (notes or "").strip().lower()
            imported_marker = (
                ws in mgr_week_set
                or (entered_by or "").strip().lower() == "import"
                or notes_s.startswith("imported from schedule")
            )
            if not imported_marker:
                data_quality_rows.append(
                    {
                        "week_start": ws,
                        "issue": "No schedule import marker (manual week or incomplete import)",
                    }
                )
        if not weeks:
            return {
                "fy_label": fy_label,
                "fy_choices": fy_choices,
                "fy_start": fy_start.isoformat(),
                "fy_end": fy_end.isoformat(),
                "granularity": granularity,
                "date_start": date_start.isoformat(),
                "date_end": date_end.isoformat(),
                "data_through": default_end.isoformat(),
                "is_current_fy": is_current_fy,
                "latest_week_start": latest_week_start,
                "latest_updated_at": latest_updated_at,
                "last_import_week_start": last_import_week_start,
                "last_import_updated_at": last_import_updated_at,
                "labels_json": "[]",
                "staffing_rate_series_json": "[]",
                "ot_dependency_series_json": "[]",
                "shift_exception_series_json": "[]",
                "system_rw_series_json": "[]",
                "system_gr_series_json": "[]",
                "table_rows": [],
                "filters_qs": serialize_filters_query(
                    fy_label, granularity, date_start, date_end
                ),
                "manager_line_shifts_total_series_json": "[]",
                "manager_line_shifts_breakdown_order": [],
                "manager_line_shifts_breakdown_series_json": "{}",
                "manager_line_shifts_table": [],
                "exc_breakdown_order": EXC_BREAKDOWN_SERIES_ORDER,
                "exc_total_series_json": "[]",
                "exc_breakdown_series_json": "{}",
                "exc_table_rows": [],
                "no_data": True,
                "data_quality_rows": data_quality_rows,
                "expected_week_starts_count": len(expected_week_starts),
                "preset_links": preset_links,
            }

        week_starts = [w.week_start for w in weeks]
        cov_rows = (
            session.query(WeeklyBaseCoverage)
            .filter(WeeklyBaseCoverage.week_start.in_(week_starts))
            .all()
        )
        coverages_by_week: dict[str, list[WeeklyBaseCoverage]] = defaultdict(list)
        for c in cov_rows:
            coverages_by_week[c.week_start].append(c)

        bases = list(session.query(BaseConfig).all())

        # Precompute weekly metrics objects.
        weekly_metrics: list[tuple[date, object]] = []
        for w in weeks:
            m = compute_week_metrics(w, coverages_by_week[w.week_start], bases)
            weekly_metrics.append((date.fromisoformat(w.week_start), m))

        # Exceptions: weekly totals from WeeklyLeaveDetail by leave_type, rolled up into groups.
        exc_by_week_by_group: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
        exc_total_by_week: dict[str, int] = defaultdict(int)
        exc_q = (
            session.query(
                WeeklyLeaveDetail.week_start,
                WeeklyLeaveDetail.leave_type,
                func.sum(WeeklyLeaveDetail.count),
            )
            .filter(WeeklyLeaveDetail.week_start >= date_start.isoformat())
            .filter(WeeklyLeaveDetail.week_start <= date_end.isoformat())
            .group_by(WeeklyLeaveDetail.week_start, WeeklyLeaveDetail.leave_type)
            .all()
        )
        for ws, lt, total in exc_q:
            n = int(total or 0)
            ws_s = str(ws)
            group = _exc_group_for_leave_type(str(lt or ""))
            exc_by_week_by_group[ws_s][group] += n
            exc_total_by_week[ws_s] += n

        # Manager line shifts: count per bucket from WeeklyManagerShift.shift_date.
        # Default exec-friendly breakdown is by base_name.
        mgr_total_by_shift_date: dict[str, int] = defaultdict(int)
        mgr_by_shift_date_by_base: dict[str, dict[str, int]] = defaultdict(
            lambda: defaultdict(int)
        )
        mgr_rows = (
            session.query(
                WeeklyManagerShift.shift_date,
                WeeklyManagerShift.base_name,
                func.count(WeeklyManagerShift.id),
            )
            .filter(WeeklyManagerShift.shift_date >= date_start.isoformat())
            .filter(WeeklyManagerShift.shift_date <= date_end.isoformat())
            .filter(
                (WeeklyManagerShift.event_type == "line_shift")
                | (WeeklyManagerShift.event_type.is_(None))
                | (WeeklyManagerShift.event_type == "")
            )
            .group_by(WeeklyManagerShift.shift_date, WeeklyManagerShift.base_name)
            .all()
        )
        base_names: set[str] = set()
        for sd, base_name, n in mgr_rows:
            sd_s = str(sd)
            base_s = str(base_name or "").strip() or "(Unknown)"
            base_names.add(base_s)
            nn = int(n or 0)
            mgr_total_by_shift_date[sd_s] += nn
            mgr_by_shift_date_by_base[sd_s][base_s] += nn

        manager_line_shifts_breakdown_order = sorted(base_names, key=lambda s: s.lower())
        manager_line_shifts_breakdown_series = {
            b: [] for b in manager_line_shifts_breakdown_order
        }

    # Aggregate into buckets (simple arithmetic mean of week-level metrics).
    for b_start, b_end in buckets:
        in_bucket = [(d0, m) for d0, m in weekly_metrics if b_start <= d0 <= b_end]
        if not in_bucket:
            continue
        bucket_metrics = [m for _d0, m in in_bucket]
        rollups = compute_period_rollups(bucket_metrics)
        assert rollups is not None
        n = rollups.n_weeks
        avg_staffing = rollups.avg_staffing_rate
        avg_ot = rollups.avg_ot_dependency
        avg_exc = rollups.avg_leave_exposure
        avg_rw = rollups.avg_system_rw_pct
        avg_gr = rollups.avg_system_gr_pct

        week_start_min = min(d0 for d0, _m in in_bucket)
        week_start_max = max(d0 for d0, _m in in_bucket)
        week_start_range_label = (
            week_start_min.isoformat()
            if week_start_min == week_start_max
            else f"{week_start_min.isoformat()}–{week_start_max.isoformat()}"
        )

        label = bucket_label(
            granularity,
            b_start,
            b_end,
            fy_week1=fy_start,
        )
        labels.append(label)
        staffing_rate_series.append(round(100.0 * avg_staffing, 2))
        ot_dependency_series.append(round(100.0 * avg_ot, 2))
        shift_exception_series.append(round(100.0 * avg_exc, 2))
        system_rw_series.append(round(100.0 * avg_rw, 2))
        system_gr_series.append(round(100.0 * avg_gr, 2))

        drill_qs = serialize_filters_query_from_parts(
            {
                "fy": str(fy_label),
                "granularity": granularity,
                "date_start": b_start.isoformat(),
                "date_end": b_end.isoformat(),
            }
        )
        table_rows.append(
            {
                "label": label,
                "bucket_start": b_start.isoformat(),
                "bucket_end": b_end.isoformat(),
                "weeks_included": n,
                "week_start_min": week_start_min.isoformat(),
                "week_start_max": week_start_max.isoformat(),
                "week_start_range_label": week_start_range_label,
                "staffing_rate_pct": round(100.0 * avg_staffing, 2),
                "ot_dependency_pct": round(100.0 * avg_ot, 2),
                "shift_exception_pct": round(100.0 * avg_exc, 2),
                "system_rw_coverage_pct": round(100.0 * avg_rw, 2),
                "system_gr_coverage_pct": round(100.0 * avg_gr, 2),
                "staffing_rate_pooled_pct": round(
                    100.0 * rollups.pooled_staffing_rate, 2
                ),
                "ot_dependency_pooled_pct": round(
                    100.0 * rollups.pooled_ot_dependency, 2
                ),
                "shift_exception_pooled_pct": round(
                    100.0 * rollups.pooled_leave_exposure, 2
                ),
                "system_rw_coverage_pooled_pct": round(
                    100.0 * rollups.pooled_system_rw_pct, 2
                ),
                "system_gr_coverage_pooled_pct": round(
                    100.0 * rollups.pooled_system_gr_pct, 2
                ),
                "drill_qs": drill_qs,
            }
        )

        # Manager line shifts per bucket: sum counts by day (shift_date)
        mgr_bucket_total = 0
        mgr_bucket_by_base: dict[str, int] = {b: 0 for b in manager_line_shifts_breakdown_order}
        cur = b_start
        while cur <= b_end:
            sd = cur.isoformat()
            mgr_bucket_total += int(mgr_total_by_shift_date.get(sd, 0))
            by_base = mgr_by_shift_date_by_base.get(sd, {})
            for b in manager_line_shifts_breakdown_order:
                mgr_bucket_by_base[b] += int(by_base.get(b, 0))
            cur += timedelta(days=1)
        manager_line_shifts_total_series.append(mgr_bucket_total)
        for b in manager_line_shifts_breakdown_order:
            manager_line_shifts_breakdown_series[b].append(int(mgr_bucket_by_base.get(b, 0)))
        manager_line_shifts_table.append(
            {
                "label": label,
                "bucket_start": b_start.isoformat(),
                "bucket_end": b_end.isoformat(),
                "manager_line_shifts_total": mgr_bucket_total,
                "manager_line_shifts_by_base": mgr_bucket_by_base,
                "manager_line_shifts_by_base_list": [
                    int(mgr_bucket_by_base.get(b, 0))
                    for b in manager_line_shifts_breakdown_order
                ],
                "drill_manager_shifts_qs": serialize_filters_query_from_parts(
                    {
                        "fy": str(fy_label),
                        "date_start": b_start.isoformat(),
                        "date_end": b_end.isoformat(),
                    }
                ),
            }
        )

        # Exceptions per bucket: sum weekly totals where week_start falls in bucket
        totals = {k: 0 for k in EXC_BREAKDOWN_SERIES_ORDER}
        total_all = 0
        for ws, _m in [(d0.isoformat(), m) for d0, m in weekly_metrics]:
            d_ws = date.fromisoformat(ws)
            if not (b_start <= d_ws <= b_end):
                continue
            total_all += int(exc_total_by_week.get(ws, 0))
            by_group = exc_by_week_by_group.get(ws, {})
            for k in EXC_BREAKDOWN_SERIES_ORDER:
                totals[k] += int(by_group.get(k, 0))

        exc_total_series.append(int(total_all))
        for k in EXC_BREAKDOWN_SERIES_ORDER:
            exc_breakdown_series[k].append(int(totals[k]))
        exc_table_rows.append(
            {
                "label": label,
                "bucket_start": b_start.isoformat(),
                "bucket_end": b_end.isoformat(),
                "exceptions_total": int(total_all),
                "exceptions_lt": int(totals["LT"]),
                "exceptions_loa": int(totals["LOA"]),
                "exceptions_sick": int(totals["SICK"]),
                "exceptions_at": int(totals["AT"]),
                "exceptions_jury": int(totals["JURY"]),
                "exceptions_brev": int(totals["BREV"]),
                "exceptions_other": int(totals["Other"]),
                "drill_qs": drill_qs,
            }
        )

    return {
        "fy_label": fy_label,
        "fy_choices": fy_choices,
        "fy_start": fy_start.isoformat(),
        "fy_end": fy_end.isoformat(),
        "granularity": granularity,
        "date_start": date_start.isoformat(),
        "date_end": date_end.isoformat(),
        "data_through": default_end.isoformat(),
        "is_current_fy": is_current_fy,
        "latest_week_start": latest_week_start,
        "latest_updated_at": latest_updated_at,
        "last_import_week_start": last_import_week_start,
        "last_import_updated_at": last_import_updated_at,
        "labels_json": json.dumps(labels),
        "staffing_rate_series_json": json.dumps(staffing_rate_series),
        "ot_dependency_series_json": json.dumps(ot_dependency_series),
        "shift_exception_series_json": json.dumps(shift_exception_series),
        "system_rw_series_json": json.dumps(system_rw_series),
        "system_gr_series_json": json.dumps(system_gr_series),
        "table_rows": table_rows,
        "manager_line_shifts_total_series_json": json.dumps(manager_line_shifts_total_series),
        "manager_line_shifts_breakdown_order": manager_line_shifts_breakdown_order,
        "manager_line_shifts_breakdown_series_json": json.dumps(
            manager_line_shifts_breakdown_series
        ),
        "manager_line_shifts_table": manager_line_shifts_table,
        "exc_breakdown_order": EXC_BREAKDOWN_SERIES_ORDER,
        "exc_total_series_json": json.dumps(exc_total_series),
        "exc_breakdown_series_json": json.dumps(exc_breakdown_series),
        "exc_table_rows": exc_table_rows,
        "filters_qs": serialize_filters_query(fy_label, granularity, date_start, date_end),
        "no_data": False,
        "preset_links": preset_links,
        "data_quality_rows": data_quality_rows,
        "expected_week_starts_count": len(expected_week_starts),
    }


def staffing_dashboard(request):
    """
    Staffing dashboard with Quarter / Month / derived biweekly buckets and FY presets.
    """
    try:
        ctx = _build_staffing_dashboard_context(request)
    except Http404 as exc:
        messages.error(request, str(exc))
        return redirect("home")
    return render(request, "dashboard/staffing_dashboard.html", ctx)


def staffing_dashboard_export_csv(request):
    """Export the currently-selected staffing dashboard summary as CSV."""
    ctx = _build_staffing_dashboard_context(request)
    rows = cast(list[dict[str, object]], ctx.get("table_rows") or [])
    mgr_rows = cast(list[dict[str, object]], ctx.get("manager_line_shifts_table") or [])
    mgr_order = cast(list[str], ctx.get("manager_line_shifts_breakdown_order") or [])
    exc_rows = cast(list[dict[str, object]], ctx.get("exc_table_rows") or [])
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Metadata"])
    writer.writerow(["Generated (UTC)", _utc_now_iso()])
    writer.writerow(["FY label", f"FY{ctx.get('fy_label')}"])
    writer.writerow(["PP#1 week-1 Sunday (FY start)", ctx.get("fy_start")])
    writer.writerow(["FY end (inclusive)", ctx.get("fy_end")])
    writer.writerow(["FY / pay period policy", FY_AND_PAY_PERIOD_POLICY_NOTE])
    writer.writerow(["Granularity", ctx.get("granularity")])
    writer.writerow(["Date start", ctx.get("date_start")])
    writer.writerow(["Date end", ctx.get("date_end")])
    writer.writerow(["Data through (default)", ctx.get("data_through")])
    writer.writerow(["Latest week in DB", ctx.get("latest_week_start")])
    writer.writerow(["Last schedule import applied", ctx.get("last_import_week_start")])
    writer.writerow([])
    writer.writerow(
        [
            "Period",
            "Period start",
            "Period end",
            "Weeks included",
            "Week starts included",
            "Staffing rate (%)",
            "Staffing rate pooled (%)",
            "OT dependency (%)",
            "OT dependency pooled (%)",
            "Shift exception (%)",
            "Shift exception pooled (%)",
            "System RW coverage (%)",
            "System RW coverage pooled (%)",
            "System GR coverage (%)",
            "System GR coverage pooled (%)",
        ]
    )
    for r in rows:
        writer.writerow(
            [
                r.get("label"),
                r.get("bucket_start"),
                r.get("bucket_end"),
                r.get("weeks_included"),
                r.get("week_start_range_label"),
                r.get("staffing_rate_pct"),
                r.get("staffing_rate_pooled_pct"),
                r.get("ot_dependency_pct"),
                r.get("ot_dependency_pooled_pct"),
                r.get("shift_exception_pct"),
                r.get("shift_exception_pooled_pct"),
                r.get("system_rw_coverage_pct"),
                r.get("system_rw_coverage_pooled_pct"),
                r.get("system_gr_coverage_pct"),
                r.get("system_gr_coverage_pooled_pct"),
            ]
        )

    # Manager line shifts section
    writer.writerow([])
    writer.writerow(["Manager line shifts"])
    writer.writerow(
        ["Period", "Period start", "Period end", "Total (count)"] + [f"{b} (count)" for b in mgr_order]
    )
    for r in mgr_rows:
        by_base = cast(dict[str, int], r.get("manager_line_shifts_by_base") or {})
        writer.writerow(
            [
                r.get("label"),
                r.get("bucket_start"),
                r.get("bucket_end"),
                r.get("manager_line_shifts_total"),
            ]
            + [by_base.get(b, 0) for b in mgr_order]
        )

    # Exceptions section
    writer.writerow([])
    writer.writerow(["Shift exceptions (counts)"])
    writer.writerow(
        [
            "Period",
            "Period start",
            "Period end",
            "Exceptions total (count)",
            "LT (count)",
            "LOA (count)",
            "SICK/SL (count)",
            "AT (count)",
            "JURY (count)",
            "BREV (count)",
            "Other (count)",
        ]
    )
    for r in exc_rows:
        writer.writerow(
            [
                r.get("label"),
                r.get("bucket_start"),
                r.get("bucket_end"),
                r.get("exceptions_total"),
                r.get("exceptions_lt"),
                r.get("exceptions_loa"),
                r.get("exceptions_sick"),
                r.get("exceptions_at"),
                r.get("exceptions_jury"),
                r.get("exceptions_brev"),
                r.get("exceptions_other"),
            ]
        )
    csv_bytes = output.getvalue().encode("utf-8-sig")
    filename = f"staffing_dashboard_{ctx.get('granularity')}_{ctx.get('date_start')}_to_{ctx.get('date_end')}.csv"
    response = HttpResponse(csv_bytes, content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def staffing_dashboard_export_xlsx(request):
    """Export the currently-selected staffing dashboard summary as XLSX."""
    ctx = _build_staffing_dashboard_context(request)
    rows = cast(list[dict[str, object]], ctx.get("table_rows") or [])
    mgr_rows = cast(list[dict[str, object]], ctx.get("manager_line_shifts_table") or [])
    mgr_order = cast(list[str], ctx.get("manager_line_shifts_breakdown_order") or [])
    exc_rows = cast(list[dict[str, object]], ctx.get("exc_table_rows") or [])

    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "Metadata"
    ws_meta.append(["Key", "Value"])
    ws_meta.append(["Generated (UTC)", _utc_now_iso()])
    ws_meta.append(["FY label", f"FY{ctx.get('fy_label')}"])
    ws_meta.append(["PP#1 week-1 Sunday (FY start)", ctx.get("fy_start")])
    ws_meta.append(["FY end (inclusive)", ctx.get("fy_end")])
    ws_meta.append(["FY / pay period policy", FY_AND_PAY_PERIOD_POLICY_NOTE])
    ws_meta.append(["Granularity", ctx.get("granularity")])
    ws_meta.append(["Date start", ctx.get("date_start")])
    ws_meta.append(["Date end", ctx.get("date_end")])
    ws_meta.append(["Data through (default)", ctx.get("data_through")])
    ws_meta.append(["Latest week in DB", ctx.get("latest_week_start")])
    ws_meta.append(["Latest updated at", ctx.get("latest_updated_at")])
    ws_meta.append(["Last schedule import week_start", ctx.get("last_import_week_start")])
    ws_meta.append(["Last schedule import updated at", ctx.get("last_import_updated_at")])

    ws = wb.create_sheet("Summary", 1)
    ws.append(
        [
            "Period",
            "Period start",
            "Period end",
            "Weeks included",
            "Week starts included",
            "Staffing rate (%)",
            "Staffing rate pooled (%)",
            "OT dependency (%)",
            "OT dependency pooled (%)",
            "Shift exception (%)",
            "Shift exception pooled (%)",
            "System RW coverage (%)",
            "System RW coverage pooled (%)",
            "System GR coverage (%)",
            "System GR coverage pooled (%)",
        ]
    )
    for r in rows:
        ws.append(
            [
                r.get("label"),
                r.get("bucket_start"),
                r.get("bucket_end"),
                r.get("weeks_included"),
                r.get("week_start_range_label"),
                r.get("staffing_rate_pct"),
                r.get("staffing_rate_pooled_pct"),
                r.get("ot_dependency_pct"),
                r.get("ot_dependency_pooled_pct"),
                r.get("shift_exception_pct"),
                r.get("shift_exception_pooled_pct"),
                r.get("system_rw_coverage_pct"),
                r.get("system_rw_coverage_pooled_pct"),
                r.get("system_gr_coverage_pct"),
                r.get("system_gr_coverage_pooled_pct"),
            ]
        )

    ws_mgr = wb.create_sheet("Manager line shifts")
    ws_mgr.append(["Period", "Period start", "Period end", "Total (count)"] + [f"{b} (count)" for b in mgr_order])
    for r in mgr_rows:
        by_base = cast(dict[str, int], r.get("manager_line_shifts_by_base") or {})
        ws_mgr.append(
            [
                r.get("label"),
                r.get("bucket_start"),
                r.get("bucket_end"),
                r.get("manager_line_shifts_total"),
            ]
            + [by_base.get(b, 0) for b in mgr_order]
        )

    ws_exc = wb.create_sheet("Exceptions")
    ws_exc.append(
        [
            "Period",
            "Period start",
            "Period end",
            "Exceptions total (count)",
            "LT (count)",
            "LOA (count)",
            "SICK/SL (count)",
            "AT (count)",
            "JURY (count)",
            "BREV (count)",
            "Other (count)",
        ]
    )
    for r in exc_rows:
        ws_exc.append(
            [
                r.get("label"),
                r.get("bucket_start"),
                r.get("bucket_end"),
                r.get("exceptions_total"),
                r.get("exceptions_lt"),
                r.get("exceptions_loa"),
                r.get("exceptions_sick"),
                r.get("exceptions_at"),
                r.get("exceptions_jury"),
                r.get("exceptions_brev"),
                r.get("exceptions_other"),
            ]
        )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"staffing_dashboard_{ctx.get('granularity')}_{ctx.get('date_start')}_to_{ctx.get('date_end')}.xlsx"
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

