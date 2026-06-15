"""Manager line shifts listing view and exports."""

import csv
import io
import json
from collections import defaultdict
from datetime import date
from typing import cast

from django.contrib import messages
from django.http import Http404, HttpResponse
from django.shortcuts import redirect, render
from openpyxl import Workbook
from sqlalchemy import func
from staffing_tool.db import session_scope
from staffing_tool.fiscal_year import (
    fiscal_quarter_windows_for_fy,
    fy_end_date,
    fy_label_year,
    fy_week1_sunday_containing,
    pay_period_count_for_fy,
    pay_periods_for_fy,
)
from staffing_tool.manager_names import canonical_manager_name
from staffing_tool.models import WeeklyManagerShift
from staffing_tool.time_buckets import (
    bucket_label,
    bucket_label_short,
    buckets_for_range,
)
from staffing_tool.timeutil import utc_now_iso as _utc_now_iso

from .dashboard_filters import (
    fy_choice_rows,
    last_closed_pay_period_end_for_fy,
    parse_date_param,
    parse_fy_week1_from_request,
    serialize_filters_query,
)
from .helpers import (
    DB_PATH,
    FY_AND_PAY_PERIOD_POLICY_NOTE,
    _ensure_db,
    _manager_last_names_upper_for_parse,
)

# Manager line-shift minimums (policy): full FY and biweekly pay-period equivalent.
MANAGER_MIN_SHIFTS_PER_FY = 52
MANAGER_MIN_PER_PAY_PERIOD = 2
MANAGER_DETAIL_PAGE_SIZE = 50
MANAGER_AOC_DETAIL_PAGE_SIZE = 50


def _manager_row_event_type(row: WeeklyManagerShift) -> str:
    """Normalize stored event_type (legacy rows default to line_shift)."""
    et = (getattr(row, "event_type", None) or "").strip().lower()
    return et if et in {"line_shift", "aoc"} else "line_shift"


# Distinct colors for stacked period chart (BMF palette + extras).
MANAGER_CHART_COLORS = (
    "#2a4492",
    "#052c47",
    "#c12126",
    "#0b3d91",
    "#5c2d91",
    "#b31b1b",
    "#198754",
    "#fd7e14",
    "#6f42c1",
    "#20c997",
    "#495057",
    "#adb5bd",
)


def _prorated_manager_minimum(
    range_start: date, range_end: date
) -> tuple[float, date, date, int, int]:
    """
    Minimum expected manager line-shifts per person for ``range_start``..``range_end``,
    prorated from 52 per fiscal year over the overlap with the FY that contains ``range_end``.

    Returns (target_float, fy_start, fy_end, overlap_days, fy_total_days).
    """
    fy_start = fy_week1_sunday_containing(range_end)
    fy_end = fy_end_date(fy_start)
    fy_total_days = (fy_end - fy_start).days + 1
    overlap_s = max(range_start, fy_start)
    overlap_e = min(range_end, fy_end)
    if overlap_s > overlap_e:
        return 0.0, fy_start, fy_end, 0, fy_total_days
    overlap_days = (overlap_e - overlap_s).days + 1
    target = MANAGER_MIN_SHIFTS_PER_FY * overlap_days / fy_total_days
    return target, fy_start, fy_end, overlap_days, fy_total_days


def _status_for_count(n: int, prorated_min: float) -> tuple[str, bool, float, float]:
    if prorated_min < 1e-6:
        return "N/A (no FY overlap)", True, float(n), 0.0
    delta = n - prorated_min
    met = n >= prorated_min - 1e-6
    short_by = max(0.0, prorated_min - n)
    ahead_by = max(0.0, n - prorated_min)
    if met and ahead_by < 0.05:
        status_label = "Met"
    elif met:
        status_label = f"Ahead by {ahead_by:.1f}"
    else:
        status_label = f"Short by {short_by:.1f}"
    return status_label, met, round(delta, 1), round(prorated_min, 1)


def _build_manager_shifts_context(request) -> dict[str, object]:
    """Shared context for manager shifts page and export endpoints."""
    _ensure_db()
    if not DB_PATH:
        raise Http404("Database is not configured (STAFFING_DB_PATH).")

    today = date.today()
    roster_upper = _manager_last_names_upper_for_parse()
    fy_start = parse_fy_week1_from_request(request, today)
    fy_end = fy_end_date(fy_start)
    fy_label = fy_label_year(fy_start)
    fy_choices = fy_choice_rows(fy_label_year(fy_week1_sunday_containing(today)))

    granularity = (request.GET.get("granularity") or "pay_period").strip().lower()
    if granularity not in {"quarter", "month", "pay_period", "fy_total"}:
        granularity = "pay_period"

    is_current_fy = fy_start == fy_week1_sunday_containing(today)
    last_closed_in_fy = last_closed_pay_period_end_for_fy(today, fy_start)
    default_end = last_closed_in_fy if is_current_fy else fy_end
    default_start = fy_start

    date_start = parse_date_param(request.GET.get("date_start", ""), default_start)
    date_end = parse_date_param(request.GET.get("date_end", ""), default_end)
    date_start = max(date_start, fy_start)
    date_end = min(date_end, fy_end)
    if date_start > date_end:
        date_start, date_end = default_start, default_end

    date_start_s = date_start.isoformat()
    date_end_s = date_end.isoformat()

    totals: dict[str, int] = defaultdict(int)
    aoc_totals: dict[str, int] = defaultdict(int)
    shift_rows: list[dict[str, object]] = []
    aoc_rows: list[dict[str, object]] = []
    with session_scope(DB_PATH) as session:
        db_min, db_max = session.query(
            func.min(WeeklyManagerShift.shift_date),
            func.max(WeeklyManagerShift.shift_date),
        ).one()
        shifts_raw = (
            session.query(WeeklyManagerShift)
            .filter(
                WeeklyManagerShift.shift_date >= date_start_s,
                WeeklyManagerShift.shift_date <= date_end_s,
            )
            .order_by(
                WeeklyManagerShift.shift_date,
                WeeklyManagerShift.person_display,
                WeeklyManagerShift.role,
                WeeklyManagerShift.unit_code,
            )
            .all()
        )
        for m in shifts_raw:
            raw_name = (m.person_display or "").strip() or "(unknown)"
            canon = canonical_manager_name(raw_name, roster_upper)
            event_type = _manager_row_event_type(m)
            if event_type == "aoc":
                aoc_totals[canon] += 1
                aoc_rows.append(
                    {
                        "shift_date": m.shift_date,
                        "person_display": canon,
                        "raw_person_display": raw_name if raw_name != canon else "",
                        "role": m.role,
                        "raw_value": m.raw_value,
                        "week_start": m.week_start,
                        "source_tab": m.source_tab,
                        "source_cell": m.source_cell,
                    }
                )
                continue
            totals[canon] += 1
            shift_rows.append(
                {
                    "shift_date": m.shift_date,
                    "person_display": canon,
                    "raw_person_display": raw_name if raw_name != canon else "",
                    "role": m.role,
                    "base_name": m.base_name,
                    "service_type": m.service_type,
                    "day_night": m.day_night,
                    "unit_code": m.unit_code,
                    "overtime": m.overtime,
                    "raw_value": m.raw_value,
                    "week_start": m.week_start,
                    "source_tab": m.source_tab,
                    "source_cell": m.source_cell,
                }
            )

    grand_total = len(shift_rows)
    aoc_grand_total = len(aoc_rows)
    range_start_d = date_start
    range_end_d = date_end
    prorated_min, fy_anchor_start, fy_anchor_end, overlap_days, fy_total_days = (
        _prorated_manager_minimum(range_start_d, range_end_d)
    )

    cumulative_rows: list[dict[str, object]] = []
    running = 0
    all_manager_names = sorted(
        set(totals) | set(aoc_totals),
        key=lambda n: (-(totals.get(n, 0) + aoc_totals.get(n, 0)), n.lower()),
    )
    totals_by_person = sorted(totals.items(), key=lambda x: (-x[1], x[0].lower()))
    for name in all_manager_names:
        n = totals.get(name, 0)
        aoc_n = aoc_totals.get(name, 0)
        if n:
            running += n
        pct = round(100.0 * n / grand_total, 1) if grand_total else 0.0
        status_label, met, delta, target_disp = _status_for_count(n, prorated_min)
        cumulative_rows.append(
            {
                "name": name,
                "count": n,
                "aoc_count": aoc_n,
                "pct": pct,
                "running": running if n else None,
                "target": target_disp,
                "delta": delta,
                "met": met,
                "status_label": status_label if n else "—",
            }
        )

    chart_granularity = "fy_total" if granularity == "fy_total" else granularity
    buckets = (
        [(range_start_d, range_end_d)]
        if chart_granularity == "fy_total"
        else buckets_for_range(chart_granularity, range_start_d, range_end_d)
    )
    bucket_labels_full = [
        bucket_label(chart_granularity, bs, be, fy_week1=fy_start) for bs, be in buckets
    ]
    bucket_labels_short = [
        bucket_label_short(chart_granularity, bs, be, fy_week1=fy_start)
        for bs, be in buckets
    ]

    manager_bucket: dict[str, list[int]] = {
        name: [0] * len(buckets) for name, _ in totals_by_person
    }
    for row in shift_rows:
        sd = date.fromisoformat(str(row["shift_date"]))
        name = str(row["person_display"])
        if name not in manager_bucket:
            manager_bucket[name] = [0] * len(buckets)
        for i, (bs, be) in enumerate(buckets):
            if bs <= sd <= be:
                manager_bucket[name][i] += 1
                break

    period_table_rows: list[dict[str, object]] = []
    for name, _n in totals_by_person:
        counts = manager_bucket.get(name, [0] * len(buckets))
        period_table_rows.append({"name": name, "counts": counts, "total": sum(counts)})

    top_n = 10
    chart_managers = [name for name, _n in totals_by_person[:top_n]]
    if len(totals_by_person) > top_n:
        chart_managers.append("Other")
    stacked_series: dict[str, list[int]] = {}
    for mgr in chart_managers:
        if mgr == "Other":
            other_counts = [0] * len(buckets)
            for name, counts in manager_bucket.items():
                if name in chart_managers[:-1]:
                    continue
                for i, c in enumerate(counts):
                    other_counts[i] += c
            stacked_series[mgr] = other_counts
        else:
            stacked_series[mgr] = manager_bucket.get(mgr, [0] * len(buckets))

    progress_labels = [row["name"] for row in cumulative_rows]
    progress_shifts = [row["count"] for row in cumulative_rows]
    progress_targets = [
        row["target"] if row["target"] else 0 for row in cumulative_rows
    ]
    progress_met = [row["met"] for row in cumulative_rows]

    periods = pay_periods_for_fy(fy_start)
    end_anchor = default_end
    closed = [p for p in periods if p.end <= end_anchor]
    last6 = closed[-6:] if len(closed) >= 6 else closed
    last6_start = last6[0].start if last6 else fy_start
    last6_end = last6[-1].end if last6 else end_anchor
    quarter_presets: list[dict[str, str]] = []
    for qnum, qa, qb in fiscal_quarter_windows_for_fy(fy_start):
        rs = max(qa, fy_start)
        re = min(qb, fy_end)
        if rs > re:
            continue
        quarter_presets.append(
            {
                "label": f"FY{fy_label} Q{qnum}",
                "qs": serialize_filters_query(fy_label, "quarter", rs, re),
            }
        )
    preset_links = {
        "fy_ytd": {
            "label": "FY YTD (last closed PP)",
            "qs": serialize_filters_query(fy_label, "pay_period", fy_start, end_anchor),
        },
        "last_6_pp": {
            "label": "Last 6 pay periods",
            "qs": serialize_filters_query(
                fy_label, "pay_period", last6_start, last6_end
            ),
        },
        "full_fy": {
            "label": "Full FY",
            "qs": serialize_filters_query(fy_label, "quarter", fy_start, fy_end),
        },
        "fy_total": {
            "label": "FY total (single bar)",
            "qs": serialize_filters_query(fy_label, "fy_total", fy_start, end_anchor),
        },
    }

    fy_anchor = fy_week1_sunday_containing(today)
    fy_end_cur = fy_end_date(fy_anchor)
    fy_label_current = fy_label_year(fy_anchor)
    cy_start = date(today.year, 1, 1)
    pp_count = pay_period_count_for_fy(fy_anchor)

    page_raw = (request.GET.get("page") or "1").strip()
    try:
        detail_page = max(1, int(page_raw))
    except ValueError:
        detail_page = 1
    detail_total = len(shift_rows)
    detail_page_count = max(
        1, (detail_total + MANAGER_DETAIL_PAGE_SIZE - 1) // MANAGER_DETAIL_PAGE_SIZE
    )
    detail_page = min(detail_page, detail_page_count)
    detail_start = (detail_page - 1) * MANAGER_DETAIL_PAGE_SIZE
    detail_page_rows = shift_rows[
        detail_start : detail_start + MANAGER_DETAIL_PAGE_SIZE
    ]

    aoc_page_raw = (request.GET.get("aoc_page") or "1").strip()
    try:
        aoc_detail_page = max(1, int(aoc_page_raw))
    except ValueError:
        aoc_detail_page = 1
    aoc_detail_total = len(aoc_rows)
    aoc_detail_page_count = max(
        1,
        (aoc_detail_total + MANAGER_AOC_DETAIL_PAGE_SIZE - 1)
        // MANAGER_AOC_DETAIL_PAGE_SIZE,
    )
    aoc_detail_page = min(aoc_detail_page, aoc_detail_page_count)
    aoc_detail_start = (aoc_detail_page - 1) * MANAGER_AOC_DETAIL_PAGE_SIZE
    aoc_detail_page_rows = aoc_rows[
        aoc_detail_start : aoc_detail_start + MANAGER_AOC_DETAIL_PAGE_SIZE
    ]

    aoc_summary_rows = sorted(
        (
            {"name": name, "count": aoc_totals[name]}
            for name in all_manager_names
            if aoc_totals.get(name, 0)
        ),
        key=lambda row: (-row["count"], row["name"].lower()),
    )

    return {
        "date_start": date_start_s,
        "date_end": date_end_s,
        "shifts": detail_page_rows,
        "all_shifts": shift_rows,
        "aoc_rows": aoc_detail_page_rows,
        "all_aoc_rows": aoc_rows,
        "aoc_summary_rows": aoc_summary_rows,
        "aoc_grand_total": aoc_grand_total,
        "detail_page": detail_page,
        "detail_page_count": detail_page_count,
        "detail_page_size": MANAGER_DETAIL_PAGE_SIZE,
        "detail_total": detail_total,
        "aoc_detail_page": aoc_detail_page,
        "aoc_detail_page_count": aoc_detail_page_count,
        "aoc_detail_page_size": MANAGER_AOC_DETAIL_PAGE_SIZE,
        "aoc_detail_total": aoc_detail_total,
        "cumulative_rows": cumulative_rows,
        "period_table_rows": period_table_rows,
        "bucket_labels": bucket_labels_full,
        "shift_count": grand_total,
        "grand_total": grand_total,
        "today_iso": today.isoformat(),
        "fy_label": fy_label,
        "fy_label_current": fy_label_current,
        "fy_start": fy_start.isoformat(),
        "fy_end": fy_end.isoformat(),
        "fy_start_iso": fy_anchor.isoformat(),
        "fy_end_iso": fy_end_cur.isoformat(),
        "calendar_year_start_iso": cy_start.isoformat(),
        "db_date_min": db_min,
        "db_date_max": db_max,
        "manager_min_fy": MANAGER_MIN_SHIFTS_PER_FY,
        "manager_min_per_pp": MANAGER_MIN_PER_PAY_PERIOD,
        "manager_pp_per_fy": pp_count,
        "prorated_manager_min": round(prorated_min, 1),
        "fy_target_start": fy_anchor_start.isoformat(),
        "fy_target_end": fy_anchor_end.isoformat(),
        "fy_overlap_days": overlap_days,
        "fy_total_days": fy_total_days,
        "granularity": granularity,
        "fy_choices": fy_choices,
        "preset_links": preset_links,
        "quarter_presets": quarter_presets,
        "filters_qs": serialize_filters_query(
            fy_label, granularity, date_start, date_end
        ),
        "chart_labels_json": json.dumps(bucket_labels_short),
        "chart_stacked_json": json.dumps(stacked_series),
        "chart_colors_json": json.dumps(
            list(MANAGER_CHART_COLORS[: len(chart_managers)])
        ),
        "chart_progress_labels_json": json.dumps(progress_labels),
        "chart_progress_shifts_json": json.dumps(progress_shifts),
        "chart_progress_targets_json": json.dumps(progress_targets),
        "chart_progress_met_json": json.dumps(progress_met),
        "is_current_fy": is_current_fy,
        "data_through": end_anchor.isoformat(),
    }


def manager_shifts(request):
    """Manager line shifts stored from schedule import (date range filter)."""
    try:
        ctx = _build_manager_shifts_context(request)
    except Http404 as exc:
        messages.error(request, str(exc))
        return redirect("home")
    return render(request, "dashboard/manager_shifts.html", ctx)


def manager_shifts_export_csv(request):
    """Export manager shift summary and detail for the current filter selection."""
    ctx = _build_manager_shifts_context(request)
    cumulative_rows = cast(list[dict[str, object]], ctx.get("cumulative_rows") or [])
    period_rows = cast(list[dict[str, object]], ctx.get("period_table_rows") or [])
    bucket_labels = cast(list[str], ctx.get("bucket_labels") or [])
    shift_rows = cast(
        list[dict[str, object]], ctx.get("all_shifts") or ctx.get("shifts") or []
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Metadata"])
    writer.writerow(["Generated (UTC)", _utc_now_iso()])
    writer.writerow(["Report", "Manager line shifts"])
    writer.writerow(["FY label", f"FY{ctx.get('fy_label')}"])
    writer.writerow(["PP#1 week-1 Sunday (FY start)", ctx.get("fy_start")])
    writer.writerow(["FY end (inclusive)", ctx.get("fy_end")])
    writer.writerow(["FY / pay period policy", FY_AND_PAY_PERIOD_POLICY_NOTE])
    writer.writerow(["Minimum shifts per person per FY", ctx.get("manager_min_fy")])
    writer.writerow(
        ["Minimum shifts per pay period (policy)", ctx.get("manager_min_per_pp")]
    )
    writer.writerow(["Granularity", ctx.get("granularity")])
    writer.writerow(["Date start", ctx.get("date_start")])
    writer.writerow(["Date end", ctx.get("date_end")])
    writer.writerow(["Data through (FY YTD default)", ctx.get("data_through")])
    writer.writerow(["Prorated minimum (per person)", ctx.get("prorated_manager_min")])
    writer.writerow(["FY target window start", ctx.get("fy_target_start")])
    writer.writerow(["FY target window end", ctx.get("fy_target_end")])
    writer.writerow(["Total person-shifts", ctx.get("grand_total")])
    writer.writerow(["Total AOC days", ctx.get("aoc_grand_total")])
    writer.writerow([])

    writer.writerow(
        [
            "Manager (last name)",
            "Shifts",
            "AOC days",
            "Min (prorated)",
            "Delta",
            "Status",
            "% of total",
            "Running total",
        ]
    )
    for row in cumulative_rows:
        writer.writerow(
            [
                row.get("name"),
                row.get("count"),
                row.get("aoc_count"),
                row.get("target"),
                row.get("delta"),
                row.get("status_label"),
                row.get("pct"),
                row.get("running") if row.get("running") is not None else "",
            ]
        )
    writer.writerow(
        [
            "Total (all managers)",
            ctx.get("grand_total"),
            ctx.get("aoc_grand_total"),
            "",
            "",
            "",
            100 if ctx.get("grand_total") else "",
            ctx.get("grand_total"),
        ]
    )

    if period_rows and bucket_labels:
        writer.writerow([])
        writer.writerow(["Shifts by manager and period"])
        writer.writerow(["Manager"] + bucket_labels + ["Total"])
        for row in period_rows:
            counts = cast(list[int], row.get("counts") or [])
            writer.writerow(
                [row.get("name")]
                + [c if c else "" for c in counts]
                + [row.get("total")]
            )

    writer.writerow([])
    writer.writerow(["Detail (one row per staffed unit assignment)"])
    writer.writerow(
        [
            "Shift date",
            "Manager",
            "Legacy label",
            "Role",
            "Base",
            "RW/GR",
            "D/N",
            "Unit",
            "OT",
            "Source value",
            "Week start",
            "Source tab",
            "Source cell",
        ]
    )
    for row in shift_rows:
        writer.writerow(
            [
                row.get("shift_date"),
                row.get("person_display"),
                row.get("raw_person_display") or "",
                row.get("role"),
                row.get("base_name"),
                row.get("service_type"),
                row.get("day_night"),
                row.get("unit_code"),
                "Yes" if row.get("overtime") else "",
                row.get("raw_value"),
                row.get("week_start"),
                row.get("source_tab"),
                row.get("source_cell"),
            ]
        )

    aoc_rows = cast(
        list[dict[str, object]], ctx.get("all_aoc_rows") or ctx.get("aoc_rows") or []
    )
    if aoc_rows:
        writer.writerow([])
        writer.writerow(["AOC detail (one row per AOC cell on manager roster rows)"])
        writer.writerow(
            [
                "Date",
                "Manager",
                "Legacy label",
                "Role",
                "Source value",
                "Week start",
                "Source tab",
                "Source cell",
            ]
        )
        for row in aoc_rows:
            writer.writerow(
                [
                    row.get("shift_date"),
                    row.get("person_display"),
                    row.get("raw_person_display") or "",
                    row.get("role"),
                    row.get("raw_value"),
                    row.get("week_start"),
                    row.get("source_tab"),
                    row.get("source_cell"),
                ]
            )

    filename = (
        f"manager_shifts_{ctx.get('granularity')}_"
        f"{ctx.get('date_start')}_to_{ctx.get('date_end')}.csv"
    )
    response = HttpResponse(
        output.getvalue().encode("utf-8-sig"),
        content_type="text/csv; charset=utf-8",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def manager_shifts_export_xlsx(request):
    """Export manager shift summary and detail as Excel workbook."""
    ctx = _build_manager_shifts_context(request)
    cumulative_rows = cast(list[dict[str, object]], ctx.get("cumulative_rows") or [])
    period_rows = cast(list[dict[str, object]], ctx.get("period_table_rows") or [])
    bucket_labels = cast(list[str], ctx.get("bucket_labels") or [])
    shift_rows = cast(
        list[dict[str, object]], ctx.get("all_shifts") or ctx.get("shifts") or []
    )

    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "Metadata"
    ws_meta.append(["Key", "Value"])
    ws_meta.append(["Generated (UTC)", _utc_now_iso()])
    ws_meta.append(["Report", "Manager line shifts"])
    ws_meta.append(["FY label", f"FY{ctx.get('fy_label')}"])
    ws_meta.append(["PP#1 week-1 Sunday (FY start)", ctx.get("fy_start")])
    ws_meta.append(["FY end (inclusive)", ctx.get("fy_end")])
    ws_meta.append(["FY / pay period policy", FY_AND_PAY_PERIOD_POLICY_NOTE])
    ws_meta.append(["Minimum shifts per person per FY", ctx.get("manager_min_fy")])
    ws_meta.append(
        ["Minimum shifts per pay period (policy)", ctx.get("manager_min_per_pp")]
    )
    ws_meta.append(["Granularity", ctx.get("granularity")])
    ws_meta.append(["Date start", ctx.get("date_start")])
    ws_meta.append(["Date end", ctx.get("date_end")])
    ws_meta.append(["Data through (FY YTD default)", ctx.get("data_through")])
    ws_meta.append(["Prorated minimum (per person)", ctx.get("prorated_manager_min")])
    ws_meta.append(["FY target window start", ctx.get("fy_target_start")])
    ws_meta.append(["FY target window end", ctx.get("fy_target_end")])
    ws_meta.append(["Total person-shifts", ctx.get("grand_total")])
    ws_meta.append(["Total AOC days", ctx.get("aoc_grand_total")])

    ws_summary = wb.create_sheet("Summary", 1)
    ws_summary.append(
        [
            "Manager (last name)",
            "Shifts",
            "AOC days",
            "Min (prorated)",
            "Delta",
            "Status",
            "% of total",
            "Running total",
        ]
    )
    for row in cumulative_rows:
        ws_summary.append(
            [
                row.get("name"),
                row.get("count"),
                row.get("aoc_count"),
                row.get("target"),
                row.get("delta"),
                row.get("status_label"),
                row.get("pct"),
                row.get("running"),
            ]
        )
    ws_summary.append(
        [
            "Total (all managers)",
            ctx.get("grand_total"),
            ctx.get("aoc_grand_total"),
            None,
            None,
            None,
            100 if ctx.get("grand_total") else None,
            ctx.get("grand_total"),
        ]
    )

    if period_rows and bucket_labels:
        ws_period = wb.create_sheet("By period")
        ws_period.append(["Manager"] + bucket_labels + ["Total"])
        for row in period_rows:
            counts = cast(list[int], row.get("counts") or [])
            ws_period.append([row.get("name")] + list(counts) + [row.get("total")])

    ws_detail = wb.create_sheet("Detail")
    ws_detail.append(
        [
            "Shift date",
            "Manager",
            "Legacy label",
            "Role",
            "Base",
            "RW/GR",
            "D/N",
            "Unit",
            "OT",
            "Source value",
            "Week start",
            "Source tab",
            "Source cell",
        ]
    )
    for row in shift_rows:
        ws_detail.append(
            [
                row.get("shift_date"),
                row.get("person_display"),
                row.get("raw_person_display") or "",
                row.get("role"),
                row.get("base_name"),
                row.get("service_type"),
                row.get("day_night"),
                row.get("unit_code"),
                "Yes" if row.get("overtime") else "",
                row.get("raw_value"),
                row.get("week_start"),
                row.get("source_tab"),
                row.get("source_cell"),
            ]
        )

    aoc_rows = cast(
        list[dict[str, object]], ctx.get("all_aoc_rows") or ctx.get("aoc_rows") or []
    )
    if aoc_rows:
        ws_aoc = wb.create_sheet("AOC detail")
        ws_aoc.append(
            [
                "Date",
                "Manager",
                "Legacy label",
                "Role",
                "Source value",
                "Week start",
                "Source tab",
                "Source cell",
            ]
        )
        for row in aoc_rows:
            ws_aoc.append(
                [
                    row.get("shift_date"),
                    row.get("person_display"),
                    row.get("raw_person_display") or "",
                    row.get("role"),
                    row.get("raw_value"),
                    row.get("week_start"),
                    row.get("source_tab"),
                    row.get("source_cell"),
                ]
            )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = (
        f"manager_shifts_{ctx.get('granularity')}_"
        f"{ctx.get('date_start')}_to_{ctx.get('date_end')}.xlsx"
    )
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
