"""Workbook/worksheet plumbing: opening, sheet + header-row detection."""

from __future__ import annotations

import hashlib
import os
from collections.abc import Iterable, Iterator
from contextlib import contextmanager
from datetime import date, datetime, timedelta

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@contextmanager
def _open_workbook(path: str, wb: Workbook | None = None) -> Iterator[Workbook]:
    """Yield a workbook, reusing ``wb`` when provided instead of re-reading.

    A single schedule import parses the same file several times (records,
    OPS View daily/detail, raw-cell archive). Passing one already-loaded
    workbook through those passes avoids 3-4 redundant full-file reads.
    When ``wb`` is given the caller owns its lifecycle, so it is left open;
    otherwise the workbook opened here is closed on exit.
    """
    if wb is not None:
        yield wb
        return
    opened = load_workbook(path, data_only=True)
    try:
        yield opened
    finally:
        opened.close()


def schedule_file_sha256(path: str) -> str:
    """SHA-256 hex digest of a schedule workbook file."""
    digest = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _header_cell_to_date(value: object, default_year: int | None = None) -> date | None:
    """Parse a single header cell (Excel date or string) to a calendar date."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m/%d"):
        try:
            if fmt == "%m/%d":
                if default_year is None:
                    continue
                return datetime.strptime(s, fmt).replace(year=default_year).date()
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _iter_date_headers(
    row,
    week_start_date: date | None = None,
    week_end_date: date | None = None,
) -> list[tuple[int, date]]:
    """
    Given a header row (RN/Medic grid), return list of (column_index, date).
    Dates are taken from C1:P1 only (columns 3–16); row 2 is day labels
    (C2:P2).

    We expect each date header cell to be either a real Excel date or
    a string parseable as YYYY-MM-DD or similar. Cells with no valid
    date are skipped.
    If week_start_date and week_end_date are set, only columns whose
    date falls in [week_start_date, week_end_date] are included.
    """
    default_y = (
        week_start_date.year if week_start_date is not None else date.today().year
    )
    result: list[tuple[int, date]] = []
    for cell in row:
        if cell.column < 3 or cell.column > 16:  # only C–P (C1:P1)
            continue
        d = _header_cell_to_date(cell.value, default_y)
        if d is None:
            continue
        if week_start_date is not None and week_end_date is not None:
            if not week_start_date <= d <= week_end_date:
                continue
        result.append((cell.column, d))
    return result


def _normalize_sheet_title(name: str) -> str:
    s = " ".join(str(name).split())
    return s.lower().replace(" and ", " & ")


def resolve_workbook_sheet(wb: Workbook, *labels: str) -> str | None:
    """Return the workbook's actual sheet name matching any label (spacing/case/& vs and)."""
    for actual in wb.sheetnames:
        an = _normalize_sheet_title(actual)
        for lb in labels:
            if _normalize_sheet_title(lb) == an:
                return actual
    return None


def _best_header_row_in_ws(
    ws: Worksheet,
    week_start_date: date | None,
    week_end_date: date | None,
    prefer_row: int = 1,
) -> tuple[int, list[tuple[int, date]]]:
    """Use row 1–5; pick the row with the most date columns in the selected week."""
    order = [prefer_row] + [r for r in range(1, 6) if r != prefer_row]
    best_row = prefer_row
    best_cols: list[tuple[int, date]] = []
    for row_idx in order:
        col_dates = _iter_date_headers(
            ws[row_idx],
            week_start_date=week_start_date,
            week_end_date=week_end_date,
        )
        if len(col_dates) > len(best_cols):
            best_row = row_idx
            best_cols = col_dates
    return best_row, best_cols


def detect_schedule_week_starts(path: str) -> list[str]:
    """Sunday ``YYYY-MM-DD`` for each week present in schedule headers (cols C–P, rows 1–5).

    Scans **RN & Medic**, **EMT**, and **OPS View** (name variants) so two-week workbooks work.
    """
    wb = load_workbook(path, data_only=True)
    try:
        all_d: set[date] = set()
        default_y = date.today().year
        sheet_keys = [
            resolve_workbook_sheet(wb, "RN & Medic", "RN AND MEDIC", "RN/MEDIC"),
            resolve_workbook_sheet(wb, "EMT"),
            resolve_workbook_sheet(wb, "OPS View", "Ops View"),
        ]
        for sn in sheet_keys:
            if not sn:
                continue
            ws = wb[sn]
            for row_idx in range(1, 6):
                for cell in ws[row_idx]:
                    if cell.column < 3 or cell.column > 16:
                        continue
                    d = _header_cell_to_date(cell.value, default_y)
                    if d is not None:
                        all_d.add(d)
        if not all_d:
            return []
        sundays: set[date] = set()
        for d in all_d:
            sun = d - timedelta(days=(d.weekday() + 1) % 7)
            sundays.add(sun)
        return sorted(s_dt.isoformat() for s_dt in sorted(sundays))
    finally:
        wb.close()


def find_schedule_workbook_for_week(
    week_start: str,
    search_dirs: Iterable[str],
) -> str | None:
    """Return the newest .xlsx in search_dirs whose dates include week_start."""
    candidates: list[tuple[float, str]] = []
    for directory in search_dirs:
        if not directory or not os.path.isdir(directory):
            continue
        for name in os.listdir(directory):
            if not name.lower().endswith((".xlsx", ".xlsm")):
                continue
            path = os.path.join(directory, name)
            if not os.path.isfile(path):
                continue
            try:
                weeks = detect_schedule_week_starts(path)
            except Exception:
                continue
            if week_start in weeks:
                candidates.append((os.path.getmtime(path), path))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]
