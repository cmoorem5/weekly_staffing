"""
Monthly staffing Excel export: aggregates weeks in a date range (Sunday week_starts).
Reuses BMF styling from report.py.
"""

from __future__ import annotations

import os
from collections import defaultdict
from datetime import datetime
from typing import Any, cast

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from .db import session_scope
from .leave_grid import (
    EXCEPTION_COL_BREAKDOWN_KEYS,
    EXCEPTION_GRID_COLS,
    EXCEPTION_GRID_ROLES,
)
from .metrics import (
    SYSTEM_GR_MAX_SHIFTS_PER_WEEK,
    TOTAL_PERSON_SHIFTS,
    compute_period_rollups,
    compute_week_metrics,
)
from .models import (
    BaseConfig,
    WeeklyBaseCoverage,
    WeeklyLeaveDetail,
    WeeklyStaffing,
)
from .report import (
    ALIGN_CENTER,
    BOLD,
    DETAIL_BASE_ORDER,
    FILL_BMF_GRAY_BG,
    FILL_BMF_NAVY,
    FILL_HEADER_LIGHT,
    FONT_BMF_BODY_BOLD,
    FONT_BMF_SECTION,
    FONT_BMF_SUBTITLE,
    FONT_BMF_TITLE,
    THIN_BORDER,
    _add_logo,
    _bmf_border_block,
    _bmf_cell_border,
    _bmf_merge_band,
    _exc_count_breakdown,
    _new_workbook_from_template_or_empty,
    _project_root,
)

# By_Base “System total” row highlight (light blue, distinct from body rows)
FILL_BY_BASE_SYSTEM_TOTAL = PatternFill(
    start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
)


def _style_by_base_system_total_cell(
    cell, *, value, number_format: str | None = None
) -> None:
    cell.value = value
    cell.font = BOLD
    cell.fill = FILL_BY_BASE_SYSTEM_TOTAL
    cell.border = THIN_BORDER
    if number_format:
        cell.number_format = number_format


def _parse_iso(d: str) -> datetime | None:
    d = (d or "").strip()
    if not d:
        return None
    try:
        return datetime.strptime(d, "%Y-%m-%d")
    except ValueError:
        return None


def _weeks_in_range(session: Session, start_s: str, end_s: str) -> list[WeeklyStaffing]:
    """Week rows where week_start is between start_s and end_s (inclusive, ISO string compare)."""
    q = (
        session.query(WeeklyStaffing)
        .filter(WeeklyStaffing.week_start >= start_s)
        .filter(WeeklyStaffing.week_start <= end_s)
        .order_by(WeeklyStaffing.week_start)
    )
    return list(q.all())


def _week_start_key(row: object) -> str:
    """ORM instance week_start as str (avoid ColumnElement == in comprehensions for pyright)."""
    return str(getattr(row, "week_start", ""))


def _as_int(v: object, default: int = 0) -> int:
    """Coerce ORM / numeric values to int for typing and SQLAlchemy 1.x models."""
    if v is None:
        return default
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return v
    try:
        return int(cast(Any, v))
    except (TypeError, ValueError):
        return default


def export_monthly_report(
    db_path: str,
    date_start: str,
    date_end: str,
    output_dir: str | None = None,
) -> str:
    """
    Build Monthly_staffing_<start>_to_<end>.xlsx with Monthly_Summary, By_Week
    (vacancies, leave total, weekly rates), By_Base, and Exceptions
    (count + % person-shifts matrices with color-scale heat).

    date_start / date_end: YYYY-MM-DD (any day); weeks included are those whose
    Sunday week_start falls in [date_start, date_end] inclusive.
    """
    ds = _parse_iso(date_start)
    de = _parse_iso(date_end)
    if not ds or not de:
        raise ValueError("date_start and date_end must be YYYY-MM-DD.")
    if ds.date() > de.date():
        raise ValueError("date_start must be on or before date_end.")

    start_s = ds.strftime("%Y-%m-%d")
    end_s = de.strftime("%Y-%m-%d")

    with session_scope(db_path) as session:
        weeks = _weeks_in_range(session, start_s, end_s)
        bases_cfg = list(session.query(BaseConfig).order_by(BaseConfig.base_name).all())
        base_by_name: dict[str, BaseConfig] = {
            str(getattr(b, "base_name", "")): b for b in bases_cfg
        }

        cov_rows = (
            session.query(WeeklyBaseCoverage)
            .filter(WeeklyBaseCoverage.week_start >= start_s)
            .filter(WeeklyBaseCoverage.week_start <= end_s)
            .all()
        )
        leave_rows = (
            session.query(WeeklyLeaveDetail)
            .filter(WeeklyLeaveDetail.week_start >= start_s)
            .filter(WeeklyLeaveDetail.week_start <= end_s)
            .all()
        )

    # --- Aggregate base coverage (sums) — same day/night + legacy rules as compute_week_metrics ---
    base_rw_sum: dict[str, int] = defaultdict(int)
    base_gr_sum: dict[str, int] = defaultdict(int)
    for c in cov_rows:
        cc = cast(Any, c)
        rw_d = _as_int(getattr(cc, "rw_staffed_day", 0))
        rw_n = _as_int(getattr(cc, "rw_staffed_night", 0))
        rw_tot = _as_int(getattr(cc, "rw_staffed_unit_days", 0))
        if rw_d + rw_n == 0 and rw_tot > 0:
            rw_d = rw_tot
        gr_d = _as_int(getattr(cc, "gr_staffed_day", 0))
        gr_n = _as_int(getattr(cc, "gr_staffed_night", 0))
        gr_tot = _as_int(getattr(cc, "gr_staffed_unit_days", 0))
        if gr_d + gr_n == 0 and gr_tot > 0:
            gr_d = gr_tot
        bn = str(getattr(cc, "base_name", ""))
        base_rw_sum[bn] += rw_d + rw_n
        base_gr_sum[bn] += gr_d + gr_n

    # --- Aggregate leave detail ---
    leave_agg: dict[tuple[str, str], int] = defaultdict(int)
    for lr in leave_rows:
        lr_any = cast(Any, lr)
        k = (str(lr_any.role), str(lr_any.leave_type))
        leave_agg[k] += _as_int(getattr(lr_any, "count", 0))

    wb = _new_workbook_from_template_or_empty()
    safe_start = start_s.replace("-", "")
    safe_end = end_s.replace("-", "")
    out_name = f"Monthly_staffing_{safe_start}_to_{safe_end}.xlsx"
    out_dir = output_dir or os.path.join(_project_root(), "output")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, out_name)

    n_weeks = len(weeks)

    # Weekly metrics for period averages (dashboard KPIs)
    weekly_metrics: list[Any] = []
    for w in weeks:
        wk = _week_start_key(w)
        cov = [c for c in cov_rows if _week_start_key(c) == wk]
        weekly_metrics.append(compute_week_metrics(w, cov, bases_cfg))

    rollups = compute_period_rollups(weekly_metrics)
    avg_staffing = rollups.avg_staffing_rate if rollups else 0.0
    avg_ot_dep = rollups.avg_ot_dependency if rollups else 0.0
    avg_leave_exp = rollups.avg_leave_exposure if rollups else 0.0
    avg_sys_rw = rollups.avg_system_rw_pct if rollups else 0.0
    avg_sys_gr = rollups.avg_system_gr_pct if rollups else 0.0
    pooled_staffing = rollups.pooled_staffing_rate if rollups else 0.0
    pooled_ot_dep = rollups.pooled_ot_dependency if rollups else 0.0
    pooled_leave_exp = rollups.pooled_leave_exposure if rollups else 0.0
    pooled_sys_rw = rollups.pooled_system_rw_pct if rollups else 0.0
    pooled_sys_gr = rollups.pooled_system_gr_pct if rollups else 0.0

    # -------- Sheet: Summary (dashboard-style) --------
    ws0 = wb.create_sheet("Monthly_Summary", 0)
    r = 1
    logo = _add_logo(ws0, "A1", max_height_px=64)
    if logo:
        ws0.row_dimensions[1].height = 62
        ws0.cell(1, 1).fill = FILL_BMF_NAVY
        _bmf_merge_band(
            ws0,
            r,
            2,
            8,
            "Boston MedFlight — Monthly staffing dashboard",
            fill=FILL_BMF_NAVY,
            font=FONT_BMF_TITLE,
            alignment=ALIGN_CENTER,
        )
    else:
        _bmf_merge_band(
            ws0,
            r,
            1,
            8,
            "Boston MedFlight — Monthly Staffing Summary",
            fill=FILL_BMF_NAVY,
            font=FONT_BMF_TITLE,
            alignment=ALIGN_CENTER,
        )
    r = 2
    sub = f"Reporting period: {start_s} through {end_s} · Weeks included: {n_weeks}"
    # Full-width A:H so subtitle is centered like the no-logo layout (logo is row 1 only).
    _bmf_merge_band(
        ws0,
        r,
        1,
        8,
        sub,
        fill=FILL_BMF_GRAY_BG,
        font=FONT_BMF_SUBTITLE,
        alignment=ALIGN_CENTER,
    )
    r = 3
    _bmf_merge_band(
        ws0,
        r,
        1,
        8,
        "Average values for reporting period",
        fill=FILL_BMF_NAVY,
        font=FONT_BMF_SECTION,
        alignment=ALIGN_CENTER,
    )
    r += 1

    def kpi_cell(
        rr: int, c1: int, label: str, val: float, fmt: str, c_val: int = 2
    ) -> None:
        _bmf_cell_border(ws0, rr, c1, label, FONT_BMF_BODY_BOLD, fill=FILL_HEADER_LIGHT)
        vc = _bmf_cell_border(ws0, rr, c_val, val, align=ALIGN_CENTER)
        vc.number_format = fmt

    # Row 1 of KPIs: three pairs across A–F, then spacer G–H
    kpi_row = r
    kpi_cell(kpi_row, 1, "Avg staffing rate", avg_staffing, "0.0%", 2)
    kpi_cell(kpi_row, 3, "Avg OT dependency", avg_ot_dep, "0.0%", 4)
    kpi_cell(kpi_row, 5, "Avg shift exception %", avg_leave_exp, "0.0%", 6)
    _bmf_border_block(ws0, kpi_row, kpi_row, 1, 6)
    kpi_row += 1
    kpi_cell(kpi_row, 1, "Avg system RW %", avg_sys_rw, "0.0%", 2)
    kpi_cell(kpi_row, 3, "Avg system GR %", avg_sys_gr, "0.0%", 4)
    _bmf_cell_border(
        ws0,
        kpi_row,
        5,
        "Person-shifts / week (ref.)",
        FONT_BMF_BODY_BOLD,
        fill=FILL_HEADER_LIGHT,
    )
    _bmf_cell_border(ws0, kpi_row, 6, TOTAL_PERSON_SHIFTS, align=ALIGN_CENTER)
    _bmf_border_block(ws0, kpi_row, kpi_row, 1, 6)
    kpi_row += 1
    _bmf_merge_band(
        ws0,
        kpi_row,
        1,
        6,
        "Pooled period rates (sum ÷ sum)",
        fill=FILL_BMF_GRAY_BG,
        font=FONT_BMF_SUBTITLE,
        alignment=ALIGN_CENTER,
    )
    kpi_row += 1
    kpi_cell(kpi_row, 1, "Pooled staffing rate", pooled_staffing, "0.0%", 2)
    kpi_cell(kpi_row, 3, "Pooled OT dependency", pooled_ot_dep, "0.0%", 4)
    kpi_cell(kpi_row, 5, "Pooled shift exception %", pooled_leave_exp, "0.0%", 6)
    _bmf_border_block(ws0, kpi_row, kpi_row, 1, 6)
    kpi_row += 1
    kpi_cell(kpi_row, 1, "Pooled system RW %", pooled_sys_rw, "0.0%", 2)
    kpi_cell(kpi_row, 3, "Pooled system GR %", pooled_sys_gr, "0.0%", 4)
    _bmf_border_block(ws0, kpi_row, kpi_row, 1, 6)
    r = kpi_row + 1

    _bmf_merge_band(
        ws0,
        r,
        1,
        8,
        "Period volumes",
        fill=FILL_BMF_NAVY,
        font=FONT_BMF_SECTION,
    )
    r += 1

    sum_filled = sum(w.filled_day + w.filled_night for w in weeks)
    sum_ot_rn = sum((w.ot_rn_day or 0) + (w.ot_rn_night or 0) for w in weeks)
    sum_ot_med = sum((w.ot_medic_day or 0) + (w.ot_medic_night or 0) for w in weeks)
    sum_ot_emt = sum((w.ot_emt_day or 0) + (w.ot_emt_night or 0) for w in weeks)
    sum_leave = sum(
        (w.leave_at or 0)
        + (w.leave_lt or 0)
        + (w.leave_sick or 0)
        + (w.leave_loa or 0)
        + (getattr(w, "leave_jury", 0) or 0)
        + (getattr(w, "leave_brev", 0) or 0)
        for w in weeks
    )

    def period_vol_row(label: str, value: Any, rr: int) -> int:
        _bmf_cell_border(ws0, rr, 1, label, FONT_BMF_BODY_BOLD)
        _bmf_cell_border(ws0, rr, 2, value, align=ALIGN_CENTER)
        _bmf_border_block(ws0, rr, rr, 1, 4)
        return rr + 1

    _bmf_merge_band(
        ws0,
        r,
        1,
        4,
        "Crew shifts & exceptions",
        fill=FILL_BMF_GRAY_BG,
        font=FONT_BMF_SUBTITLE,
        alignment=ALIGN_CENTER,
    )
    r += 1
    r = period_vol_row("Filled crew shifts (total)", sum_filled, r)
    r = period_vol_row("Shift exceptions (total)", sum_leave, r)
    _bmf_merge_band(
        ws0,
        r,
        1,
        4,
        "Overtime",
        fill=FILL_BMF_GRAY_BG,
        font=FONT_BMF_SUBTITLE,
        alignment=ALIGN_CENTER,
    )
    r += 1
    r = period_vol_row("RN OT shifts (total)", sum_ot_rn, r)
    r = period_vol_row("Medic OT shifts (total)", sum_ot_med, r)
    r = period_vol_row("EMT OT shifts (total)", sum_ot_emt, r)

    ws0.column_dimensions["A"].width = 26
    for c in ("B", "C", "D", "E", "F", "G", "H"):
        ws0.column_dimensions[c].width = 13

    # -------- Sheet: By_Week --------
    ws1 = wb.create_sheet("By_Week", 1)
    headers = [
        "Week start",
        "Filled total",
        "Vacancies",
        "Leave total",
        "Staffing rate",
        "OT dependency",
        "Shift exc. %",
        "System RW %",
        "System GR %",
    ]
    hr = 1
    n_bw_cols = len(headers)
    for ci, h in enumerate(headers, 1):
        cell = ws1.cell(hr, ci, h)
        cell.font = BOLD
        cell.fill = FILL_HEADER_LIGHT
        cell.border = THIN_BORDER
    row = 2
    for i, w in enumerate(weeks):
        m = weekly_metrics[i]
        ws1.cell(row, 1, w.week_start)
        ws1.cell(row, 2, m.filled_total)
        ws1.cell(row, 3, m.vacancies)
        ws1.cell(row, 4, m.leave_total)
        ws1.cell(row, 5, m.staffing_rate).number_format = "0.0%"
        ws1.cell(row, 6, m.ot_dependency).number_format = "0.0%"
        ws1.cell(row, 7, m.leave_exposure).number_format = "0.0%"
        ws1.cell(row, 8, m.system_rw_pct).number_format = "0.0%"
        ws1.cell(row, 9, m.system_gr_pct).number_format = "0.0%"
        for c in range(1, n_bw_cols + 1):
            ws1.cell(row, c).border = THIN_BORDER
        row += 1
    ws1.freeze_panes = "A2"
    ws1.column_dimensions["A"].width = 14
    for c in range(2, n_bw_cols + 1):
        ws1.column_dimensions[get_column_letter(c)].width = 12

    # -------- Sheet: By_Base (sums + staffed % of period capacity) --------
    ws2 = wb.create_sheet("By_Base", 2)
    r = 1
    _bmf_merge_band(
        ws2,
        r,
        1,
        5,
        "Staffed Units and Staffing %",
        fill=FILL_BMF_NAVY,
        font=FONT_BMF_SECTION,
    )
    r += 1
    bh = [
        "Base",
        "RW staffed (sum)",
        "GR staffed (sum)",
        "RW %",
        "GR %",
    ]
    for c, h in enumerate(bh, 1):
        cell = ws2.cell(r, c, h)
        cell.font = BOLD
        cell.fill = FILL_HEADER_LIGHT
        cell.border = THIN_BORDER
        if c >= 4:
            cell.alignment = ALIGN_CENTER
    r += 1
    for base_name in DETAIL_BASE_ORDER:
        cfg = base_by_name.get(base_name)
        rw_t = _as_int(getattr(cfg, "rw_total_unit_days", 0)) if cfg is not None else 0
        gr_t = _as_int(getattr(cfg, "gr_total_unit_days", 0)) if cfg is not None else 0
        rw_sum = base_rw_sum.get(base_name, 0)
        gr_sum = base_gr_sum.get(base_name, 0)
        rw_cap_period = n_weeks * rw_t if n_weeks else 0
        gr_cap_period = n_weeks * gr_t if n_weeks else 0
        rw_pct = (rw_sum / rw_cap_period) if rw_cap_period > 0 else None
        gr_pct = (gr_sum / gr_cap_period) if gr_cap_period > 0 else None
        ws2.cell(r, 1, base_name).border = THIN_BORDER
        ws2.cell(r, 2, rw_sum).border = THIN_BORDER
        ws2.cell(r, 3, gr_sum).border = THIN_BORDER
        c_rw = ws2.cell(r, 4)
        c_rw.border = THIN_BORDER
        c_rw.alignment = ALIGN_CENTER
        if rw_pct is not None:
            c_rw.value = rw_pct
            c_rw.number_format = "0.0%"
        else:
            c_rw.value = "—"
        c_gr = ws2.cell(r, 5)
        c_gr.border = THIN_BORDER
        c_gr.alignment = ALIGN_CENTER
        if gr_pct is not None:
            c_gr.value = gr_pct
            c_gr.number_format = "0.0%"
        else:
            c_gr.value = "—"
        r += 1

    sys_rw_sum = sum(base_rw_sum.get(b, 0) for b in DETAIL_BASE_ORDER)
    sys_gr_sum = sum(base_gr_sum.get(b, 0) for b in DETAIL_BASE_ORDER)
    sys_rw_cap = 0
    for b in DETAIL_BASE_ORDER:
        cfg_b = base_by_name.get(b)
        if cfg_b is not None:
            sys_rw_cap += _as_int(getattr(cfg_b, "rw_total_unit_days", 0)) * n_weeks
    sys_gr_ops_cap = (
        SYSTEM_GR_MAX_SHIFTS_PER_WEEK * n_weeks
        if n_weeks and SYSTEM_GR_MAX_SHIFTS_PER_WEEK
        else 0
    )
    sys_rw_pct = (sys_rw_sum / sys_rw_cap) if sys_rw_cap else None
    sys_gr_pct = (sys_gr_sum / sys_gr_ops_cap) if sys_gr_ops_cap else None

    _style_by_base_system_total_cell(ws2.cell(r, 1), value="System total")
    _style_by_base_system_total_cell(ws2.cell(r, 2), value=sys_rw_sum)
    _style_by_base_system_total_cell(ws2.cell(r, 3), value=sys_gr_sum)
    c_srw = ws2.cell(r, 4)
    _style_by_base_system_total_cell(
        c_srw,
        value=sys_rw_pct if sys_rw_pct is not None else "—",
        number_format="0.0%" if sys_rw_pct is not None else None,
    )
    c_srw.alignment = ALIGN_CENTER
    c_sgr = ws2.cell(r, 5)
    _style_by_base_system_total_cell(
        c_sgr,
        value=sys_gr_pct if sys_gr_pct is not None else "—",
        number_format="0.0%" if sys_gr_pct is not None else None,
    )
    c_sgr.alignment = ALIGN_CENTER
    r += 1

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 16
    for c in ("D", "E"):
        ws2.column_dimensions[c].width = 12
    ws2.freeze_panes = "A2"

    # -------- Sheet: Exceptions (counts + % person-shifts, color-scaled “heat”) --------
    ws3 = wb.create_sheet("Exceptions", 3)
    n_exc_cols = 1 + len(EXCEPTION_GRID_COLS)
    last_type_col = get_column_letter(n_exc_cols)

    exc_counts: list[list[int]] = []
    for role in EXCEPTION_GRID_ROLES:
        row_counts: list[int] = []
        for col in EXCEPTION_GRID_COLS:
            keys = EXCEPTION_COL_BREAKDOWN_KEYS[col]
            row_counts.append(_exc_count_breakdown(leave_agg, role, keys))
        exc_counts.append(row_counts)
    col_sums = [0] * len(EXCEPTION_GRID_COLS)
    for ci, col in enumerate(EXCEPTION_GRID_COLS):
        keys = EXCEPTION_COL_BREAKDOWN_KEYS[col]
        col_sums[ci] = sum(
            _exc_count_breakdown(leave_agg, role, keys) for role in EXCEPTION_GRID_ROLES
        )

    r = 1
    _bmf_merge_band(
        ws3,
        r,
        1,
        n_exc_cols,
        "Schedule Exceptions",
        fill=FILL_BMF_NAVY,
        font=FONT_BMF_SECTION,
    )
    r += 1
    _bmf_cell_border(ws3, r, 1, "Role", FONT_BMF_BODY_BOLD, fill=FILL_HEADER_LIGHT)
    for c, lt in enumerate(EXCEPTION_GRID_COLS, start=2):
        cell = ws3.cell(r, c, lt)
        cell.font = FONT_BMF_BODY_BOLD
        cell.fill = FILL_HEADER_LIGHT
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER
    r += 1
    count_data_first = r
    for ri, role in enumerate(EXCEPTION_GRID_ROLES):
        _bmf_cell_border(ws3, r, 1, role, FONT_BMF_BODY_BOLD)
        for ci, col in enumerate(EXCEPTION_GRID_COLS, start=2):
            v = exc_counts[ri][ci - 2]
            _bmf_cell_border(ws3, r, ci, v)
        _bmf_border_block(ws3, r, r, 1, n_exc_cols)
        r += 1
    count_data_last = r - 1
    _bmf_cell_border(ws3, r, 1, "Total", FONT_BMF_BODY_BOLD)
    for ci, total in enumerate(col_sums, start=2):
        _bmf_cell_border(ws3, r, ci, total)
    _bmf_border_block(ws3, r, r, 1, n_exc_cols)

    count_heat = f"B{count_data_first}:{last_type_col}{count_data_last}"
    ws3.conditional_formatting.add(
        count_heat,
        ColorScaleRule(
            start_type="min",
            start_color="FFF9F9F9",
            end_type="max",
            end_color="FF2A4492",
        ),
    )

    person_shifts_period = TOTAL_PERSON_SHIFTS * n_weeks if n_weeks else 0
    r += 2
    _bmf_merge_band(
        ws3,
        r,
        1,
        n_exc_cols,
        "% of Period",
        fill=FILL_BMF_NAVY,
        font=FONT_BMF_SECTION,
    )
    r += 1
    _bmf_cell_border(ws3, r, 1, "Role", FONT_BMF_BODY_BOLD, fill=FILL_HEADER_LIGHT)
    for c, lt in enumerate(EXCEPTION_GRID_COLS, start=2):
        cell = ws3.cell(r, c, lt)
        cell.font = FONT_BMF_BODY_BOLD
        cell.fill = FILL_HEADER_LIGHT
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER
    r += 1
    denom_ps = float(person_shifts_period) if person_shifts_period else None
    ps_data_first = r
    for ri, role in enumerate(EXCEPTION_GRID_ROLES):
        _bmf_cell_border(ws3, r, 1, role, FONT_BMF_BODY_BOLD)
        for ci, col in enumerate(EXCEPTION_GRID_COLS, start=2):
            cnt = exc_counts[ri][ci - 2]
            cell = ws3.cell(r, ci)
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
            if denom_ps:
                cell.value = cnt / denom_ps
                cell.number_format = "0.00%"
            else:
                cell.value = "—"
        _bmf_border_block(ws3, r, r, 1, n_exc_cols)
        r += 1
    ps_data_last = r - 1
    _bmf_cell_border(ws3, r, 1, "Column %", FONT_BMF_BODY_BOLD)
    for ci, cs in enumerate(col_sums, start=2):
        cell = ws3.cell(r, ci)
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER
        if denom_ps:
            cell.value = cs / denom_ps
            cell.number_format = "0.00%"
        else:
            cell.value = "—"
    _bmf_border_block(ws3, r, r, 1, n_exc_cols)

    ps_heat = f"B{ps_data_first}:{last_type_col}{ps_data_last}"
    ws3.conditional_formatting.add(
        ps_heat,
        ColorScaleRule(
            start_type="min",
            start_color="FFF9F9F9",
            end_type="max",
            end_color="FFC12126",
        ),
    )

    ws3.freeze_panes = "A2"
    for col in range(1, n_exc_cols + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 11

    wb.save(out_path)
    return out_path
