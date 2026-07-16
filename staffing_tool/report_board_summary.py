"""Board_Summary sheet writer (no cross-sheet formulas — values only)."""

from __future__ import annotations

from datetime import UTC, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .metrics import (
    PeriodRollups,
    WeekMetrics,
    get_metric_value,
    get_pooled_metric_value,
)
from .models import (
    KpiThreshold,
)
from .rag import RAG, direction_for_metric, evaluate_rag
from .report_data import _rag_for_metric, _status_display
from .report_excel_style import (
    ALIGN_LEFT,
    BMF_BLACK,
    BOLD,
    FONT_BMF_BODY,
    FONT_FOOTER_META,
    FONT_NAME,
    THIN_BORDER,
    _add_logo,
    _base_gr_cell_notable,
    _base_rw_cell_notable,
    _fill_and_font_for_status,
    _generator_version_label,
    _kpi_notable,
    _target_display,
    _write_pct_cell,
    _write_pct_or_num,
)
from .validation import ot_action_ceiling, shift_exception_monitor_ceiling


def _generate_narrative(
    this_week: WeekMetrics,
    prior_week: WeekMetrics | None,
    avg_4w: WeekMetrics | None,
    thresholds: dict[str, KpiThreshold],
    rag_statuses: dict[str, RAG],
) -> dict[str, list[str]]:
    """Produce Key Takeaways, Drivers, Risks, Actions placeholders."""
    takeaways = []
    drivers = []
    risks = []
    actions = []

    # Overall status
    overall = rag_statuses.get("Staffing Rate", "Green")
    if overall == "Green":
        takeaways.append("Overall staffing rate is on target.")
    elif overall == "Yellow":
        takeaways.append("Overall staffing rate is below target; status is Monitor.")
    else:
        takeaways.append(
            "Overall staffing rate is below acceptable level; status is Action needed."
        )

    ot_ceil = ot_action_ceiling(thresholds)
    exc_monitor = shift_exception_monitor_ceiling(thresholds)

    # Week-over-week comparisons need a prior week…
    if prior_week:
        sr_now, sr_prior = this_week.staffing_rate, prior_week.staffing_rate
        if sr_now > sr_prior:
            takeaways.append(
                f"Staffing rate improved week-over-week ({sr_prior:.1%} → {sr_now:.1%})."
            )
        elif sr_now < sr_prior:
            takeaways.append(
                f"Staffing rate declined week-over-week ({sr_prior:.1%} → {sr_now:.1%})."
            )
        ot_now, ot_prior = this_week.ot_dependency, prior_week.ot_dependency
        if ot_now > ot_ceil and ot_now > ot_prior:
            drivers.append(
                f"OT dependency increased ({ot_prior:.1%} → {ot_now:.1%}); overtime filling gaps."
            )
    else:
        takeaways.append("No prior week for comparison.")

    # …but current-week drivers and risks do not.
    if this_week.leave_exposure > exc_monitor:
        drivers.append(
            f"Shift exception % at {this_week.leave_exposure:.1%}; contributes to coverage pressure."
        )
    if this_week.ot_dependency > ot_ceil:
        risks.append("High OT dependency; fatigue and sustainability risk.")
    if (
        rag_statuses.get("System RW Coverage %") == "Red"
        or rag_statuses.get("System GR Coverage %") == "Red"
    ):
        risks.append(
            "RW or GR system coverage below threshold; readiness and capacity risk."
        )

    # Actions (placeholders + suggestions from flags)
    if rag_statuses.get("Staffing Rate") == "Red":
        actions.append("Address staffing shortfall (scheduling/recruiting).")
    if rag_statuses.get("OT Dependency") == "Red":
        actions.append("Reduce OT dependency; review scheduling and capacity.")
    if not actions:
        actions.append("Maintain current staffing and leave monitoring.")

    return {
        "key_takeaways": takeaways[:4],
        "drivers": drivers[:4],
        "risks": risks[:3],
        "actions": actions,
    }


def _write_board_summary(
    wb: Workbook,
    week_start: str,
    week_end: str,
    this_metrics: WeekMetrics,
    prior_metrics: WeekMetrics | None,
    avg_4w: WeekMetrics | None,
    avg_12w: WeekMetrics | None,
    rollups_4w: PeriodRollups | None,
    rollups_12w: PeriodRollups | None,
    trend_list: list[tuple[str, WeekMetrics]],
    thresholds: dict[str, KpiThreshold],
    narrative: dict[str, list[str]],
    metadata: dict | None = None,
) -> None:
    ws = wb.create_sheet("Board_Summary", 0)
    row = 1

    logo_ok = _add_logo(ws, "A1", max_height_px=76)
    if logo_ok:
        ws.row_dimensions[1].height = 78
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=7)
        t = ws.cell(row, 2, "Weekly staffing summary")
        t.font = Font(name=FONT_NAME, bold=True, size=14, color=BMF_BLACK)
        t.alignment = ALIGN_LEFT
        row += 1
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
        p = ws.cell(row, 2, f"Period: {week_start} to {week_end}")
        p.font = FONT_BMF_BODY
        p.alignment = ALIGN_LEFT
        row += 2
    else:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        t = ws.cell(row, 1, "Weekly staffing summary")
        t.font = Font(name=FONT_NAME, bold=True, size=14, color=BMF_BLACK)
        t.alignment = ALIGN_LEFT
        row += 1
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
        ws.cell(row, 1, f"Period: {week_start} to {week_end}").font = FONT_BMF_BODY
        row += 2

    # KPI Panel: (display_label, internal_metric_name)
    board_metrics = [
        ("Staffing Rate", "Staffing Rate"),
        ("Backfill Rate", "OT Dependency"),
        ("Shift Exception %", "Shift Exception %"),
        ("System RW Coverage %", "System RW Coverage %"),
        ("System GR Coverage %", "System GR Coverage %"),
    ]
    headers = [
        "Metric",
        "This Week",
        "Prior Week",
        "4-Week Avg",
        "4-Week Pooled",
        "12-Week Avg",
        "12-Week Pooled",
        "Target",
        "Status",
        "Direction",
    ]
    kpi_header_row = row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = BOLD
        cell.border = THIN_BORDER
    row += 1

    for display_name, metric_key in board_metrics:
        val_this = get_metric_value(this_metrics, metric_key)
        val_prior = (
            get_metric_value(prior_metrics, metric_key) if prior_metrics else None
        )
        val_4 = get_metric_value(avg_4w, metric_key) if avg_4w else None
        val_4_pool = (
            get_pooled_metric_value(rollups_4w, metric_key) if rollups_4w else None
        )
        val_12 = get_metric_value(avg_12w, metric_key) if avg_12w else None
        val_12_pool = (
            get_pooled_metric_value(rollups_12w, metric_key) if rollups_12w else None
        )
        rag = (
            _rag_for_metric(metric_key, val_this or 0, thresholds)
            if val_this is not None
            else "Green"
        )
        direction = direction_for_metric(metric_key, val_this or 0, val_prior)
        thr = thresholds.get(metric_key)
        notable = _kpi_notable(metric_key, val_this or 0, this_metrics)
        ws.cell(row, 1, display_name).border = THIN_BORDER
        _write_pct_or_num(ws, row, 2, val_this, metric_key)
        _write_pct_or_num(ws, row, 3, val_prior, metric_key)
        _write_pct_or_num(ws, row, 4, val_4, metric_key)
        _write_pct_or_num(ws, row, 5, val_4_pool, metric_key)
        _write_pct_or_num(ws, row, 6, val_12, metric_key)
        _write_pct_or_num(ws, row, 7, val_12_pool, metric_key)
        tcell = ws.cell(row, 8, _target_display(metric_key, thr))
        tcell.border = THIN_BORDER
        status_cell = ws.cell(row, 9, _status_display(rag))
        status_cell.border = THIN_BORDER
        fill, font = _fill_and_font_for_status(
            rag,
            notable=notable,
            value=val_this or 0,
            thr=thr,
        )
        status_cell.fill = fill
        status_cell.font = font
        ws.cell(row, 10, direction).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row, 1, "Narrative").font = BOLD
    row += 1
    ws.cell(row, 1, "Key Takeaways").font = BOLD
    row += 1
    for bullet in narrative["key_takeaways"]:
        ws.cell(row, 1, "• " + bullet)
        row += 1
    row += 1
    ws.cell(row, 1, "Drivers").font = BOLD
    row += 1
    for bullet in narrative["drivers"]:
        ws.cell(row, 1, "• " + bullet)
        row += 1
    row += 1
    ws.cell(row, 1, "Risks (fatigue/coverage/readiness)").font = BOLD
    row += 1
    for bullet in narrative["risks"]:
        ws.cell(row, 1, "• " + bullet)
        row += 1
    row += 1
    ws.cell(row, 1, "Actions / Decisions Needed").font = BOLD
    row += 1
    for bullet in narrative["actions"]:
        ws.cell(row, 1, "• " + bullet)
        row += 1

    row += 2
    ws.cell(row, 1, "Base Coverage").font = BOLD
    row += 1
    ws.cell(row, 1, "Base").font = BOLD
    ws.cell(row, 2, "RW %").font = BOLD
    ws.cell(row, 3, "GR %").font = BOLD
    ws.cell(row, 4, "Notes").font = BOLD
    row += 1
    if this_metrics.base_metrics:
        t_rw = thresholds.get("System RW Coverage %")
        t_gr = thresholds.get("System GR Coverage %")
        for base_name in sorted(this_metrics.base_metrics.keys()):
            pcts = this_metrics.base_metrics[base_name]
            rw_pct, gr_pct = pcts.get("rw_pct", 0), pcts.get("gr_pct", 0)
            rw_rag = evaluate_rag(rw_pct, t_rw) if t_rw else "Green"
            gr_rag = evaluate_rag(gr_pct, t_gr) if t_gr else "Green"
            notes = ""
            if rw_rag != "Green" or gr_rag != "Green":
                notes = (
                    "Below threshold"
                    if (rw_rag == "Red" or gr_rag == "Red")
                    else "Monitor"
                )
            ws.cell(row, 1, base_name)
            _write_pct_cell(
                ws,
                row,
                2,
                rw_pct,
                rw_rag,
                thr=t_rw,
                notable=_base_rw_cell_notable(rw_pct),
            )
            _write_pct_cell(
                ws,
                row,
                3,
                gr_pct,
                gr_rag,
                thr=t_gr,
                notable=_base_gr_cell_notable(gr_pct),
            )
            ws.cell(row, 4, notes)
            row += 1

    row += 1
    meta = metadata or {}
    gen_iso = meta.get("generated_utc")
    if not gen_iso:
        gen_iso = datetime.now(UTC).isoformat().replace("+00:00", "Z")
    src = meta.get("source_filename")
    src_s = src if src else "(unknown)"
    nrows = meta.get("source_rows")
    nrow_s = str(nrows) if nrows is not None else "(unknown)"
    ver = meta.get("generator_version") or _generator_version_label()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    fc = ws.cell(
        row,
        1,
        f"Generated: {gen_iso}\n"
        f"Generator version: {ver}\n"
        f"Source: {src_s}\n"
        f"Source rows: {nrow_s}",
    )
    fc.font = FONT_FOOTER_META
    fc.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths and freeze (keep KPI header row visible when scrolling)
    ws.column_dimensions["A"].width = 28
    for c in range(2, 11):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = f"A{kpi_header_row + 1}"
