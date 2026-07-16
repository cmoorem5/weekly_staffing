"""OPS View parsing (weekly, daily, and per-person detail sheets)."""

from __future__ import annotations

from datetime import date, datetime, timedelta

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .schedule_cells import (
    LEGACY_UNIT_ALIASES,
    MAX_GR_UNIT_DAYS_PER_WEEK,
    MAX_RW_UNIT_DAYS_PER_WEEK,
    RETIRED_UNIT_CODES,
    UNIT_MAP,
    _normalize_cell_value,
)
from .schedule_types import OpsViewAssignment, OpsViewDayBase, ServiceType
from .schedule_workbook import (
    _best_header_row_in_ws,
    _open_workbook,
    resolve_workbook_sheet,
)

# --- OPS View: vehicle-level RW/GR staffed unit-days ----------------------

# OPS View: unit codes in column A, role labels in column B (rows 4-66).
# GR2 EMT, NG2 EMT, PG EMT, NP EMT, NL EMT count as EMT; Orientee/RAL ignored.
_OPS_ROLE_RN = {"RN"}
_OPS_ROLE_MEDIC = {"MEDIC"}
_OPS_ROLE_PIC = {"PIC"}
_OPS_ROLE_EMT = {"EMT", "GR2 EMT", "NG2 EMT", "PG EMT", "NP EMT", "NL EMT"}


def _ops_parse_dates_row(
    ws: Worksheet,
    week_start_date: date,
    week_end_date: date,
) -> list[tuple[int, date]]:
    """Columns C–P; try rows 1–5 for the date header (templates vary)."""
    _, best = _best_header_row_in_ws(
        ws,
        week_start_date=week_start_date,
        week_end_date=week_end_date,
        prefer_row=1,
    )
    return best


def _ops_vehicle_blocks(
    ws: Worksheet,
) -> list[tuple[str, int, dict[str, list[int]]]]:
    """
    Column A = unit code, column B = role label (rows 4-66). When A matches
    UNIT_MAP we start a new vehicle block; subsequent rows with A empty use B
    for role.
    Return (unit_code, first_row, role_rows) with role_rows mapping role to
    row indices.
    """
    blocks: list[tuple[str, int, dict[str, list[int]]]] = []
    current_unit: str | None = None
    current_start = 0
    role_rows: dict[str, list[int]] = {}

    # Row 3 is often weekday labels; unit blocks start row 4 (BMF 29 Mar 2026 layout).
    for row_idx in range(4, 67):
        cell_a = ws.cell(row=row_idx, column=1)
        cell_b = ws.cell(row=row_idx, column=2)
        raw_a = cell_a.value
        raw_b = cell_b.value
        text_a = (_normalize_cell_value(raw_a) or "").strip()
        text_b = (_normalize_cell_value(raw_b) or "").strip()
        # Unit code in A starts a new block. Any other section label in A
        # (FLOAT, OPEN, Extra, footer text) ENDS the current block — its role
        # rows belong to that section, not to the previous vehicle. It must
        # not merely reset role_rows while current_unit stays set: that
        # discarded the previous unit's rows and re-filled them from the
        # non-unit section (a FLOAT block after PG erased Plymouth ground
        # coverage this way). Parenthesized annotations like "(Badged)" are
        # still ignored without closing the block.
        if text_a and text_a not in RETIRED_UNIT_CODES:
            canonical_a = LEGACY_UNIT_ALIASES.get(text_a, text_a)
            if canonical_a in UNIT_MAP:
                if current_unit is not None:
                    blocks.append((current_unit, current_start, role_rows))
                current_unit = canonical_a
                current_start = row_idx
                role_rows = {}
            elif not (text_a.startswith("(") and text_a.endswith(")")):
                if current_unit is not None:
                    blocks.append((current_unit, current_start, role_rows))
                current_unit = None
                role_rows = {}
        if current_unit is None:
            continue
        if not text_b:
            continue
        if text_b in _OPS_ROLE_RN:
            role_rows.setdefault("RN", []).append(row_idx)
        elif text_b in _OPS_ROLE_MEDIC:
            role_rows.setdefault("Medic", []).append(row_idx)
        elif text_b in _OPS_ROLE_PIC:
            role_rows.setdefault("PIC", []).append(row_idx)
        elif text_b in _OPS_ROLE_EMT:
            role_rows.setdefault("EMT", []).append(row_idx)
    if current_unit is not None:
        blocks.append((current_unit, current_start, role_rows))
    return blocks


def _ops_cell_staffed(cell_value: object) -> bool:
    """True if cell has a non-empty value that looks like a name."""
    s = _normalize_cell_value(cell_value)
    if not s or s == "OPEN":
        return False
    return True


def _cap_base_coverage_split(
    base_rw_day: dict[str, int],
    base_rw_night: dict[str, int],
    base_gr_day: dict[str, int],
    base_gr_night: dict[str, int],
) -> None:
    """Enforce MAX_RW / MAX_GR per base on (day+night) totals.

    Reduce night counts first.
    """
    for base in set(base_rw_day) | set(base_rw_night):
        cap = MAX_RW_UNIT_DAYS_PER_WEEK.get(base)
        if cap is None:
            continue
        d, n = base_rw_day.get(base, 0), base_rw_night.get(base, 0)
        total = d + n
        if total <= cap:
            continue
        over = total - cap
        nn = min(over, n)
        base_rw_night[base] = n - nn
        over -= nn
        if over > 0:
            base_rw_day[base] = max(0, d - over)
    for base in set(base_gr_day) | set(base_gr_night):
        cap = MAX_GR_UNIT_DAYS_PER_WEEK.get(base)
        if cap is None:
            continue
        d, n = base_gr_day.get(base, 0), base_gr_night.get(base, 0)
        total = d + n
        if total <= cap:
            continue
        over = total - cap
        nn = min(over, n)
        base_gr_night[base] = n - nn
        over -= nn
        if over > 0:
            base_gr_day[base] = max(0, d - over)


def _parse_ops_view_worksheet(
    ws: Worksheet,
    week_start: str,
) -> tuple[dict[str, int], dict[str, int], dict[str, int], dict[str, int]]:
    """
    Parse an already-loaded OPS View worksheet (same rules as parse_ops_view).
    Returns (rw_day, rw_night, gr_day, gr_night) per base_name.
    """
    try:
        week_start_date = datetime.strptime(week_start, "%Y-%m-%d").date()
    except ValueError:
        return {}, {}, {}, {}
    week_end_date = week_start_date + timedelta(days=6)

    col_dates = _ops_parse_dates_row(ws, week_start_date, week_end_date)
    blocks = _ops_vehicle_blocks(ws)

    base_rw_day: dict[str, int] = {}
    base_rw_night: dict[str, int] = {}
    base_gr_day: dict[str, int] = {}
    base_gr_night: dict[str, int] = {}

    for unit_code, _first_row, role_rows in blocks:
        unit_info = UNIT_MAP.get(unit_code)
        if not unit_info:
            continue
        base, service_type, dn = unit_info
        rn_rows = role_rows.get("RN") or []
        medic_rows = role_rows.get("Medic") or []
        pic_rows = role_rows.get("PIC") or []
        emt_rows = role_rows.get("EMT") or []

        for col_idx, _d in col_dates:
            staffed = _ops_staffed_for_column(
                ws,
                col_idx,
                service_type,
                rn_rows,
                medic_rows,
                pic_rows,
                emt_rows,
            )
            if service_type == "RW":
                if staffed:
                    if dn == "N":
                        base_rw_night[base] = base_rw_night.get(base, 0) + 1
                    else:
                        base_rw_day[base] = base_rw_day.get(base, 0) + 1
            elif service_type == "GR":
                if staffed:
                    if dn == "N":
                        base_gr_night[base] = base_gr_night.get(base, 0) + 1
                    else:
                        base_gr_day[base] = base_gr_day.get(base, 0) + 1

    _cap_base_coverage_split(
        base_rw_day,
        base_rw_night,
        base_gr_day,
        base_gr_night,
    )

    return base_rw_day, base_rw_night, base_gr_day, base_gr_night


def parse_ops_view(
    path: str,
    week_start: str,
) -> tuple[dict[str, int], dict[str, int], dict[str, int], dict[str, int]]:
    """
    Parse the OPS View sheet from a workbook path (loads the file once).

    Day vs night follows the unit code in UNIT_MAP
    (e.g. D7B=day RW, N7B=night RW).

    - Dates: row 1, columns C–P; only dates in
      [week_start, week_start+6] are used.
    - Vehicles: column A rows 4–66; unit codes from UNIT_MAP; role rows
      RN, Medic, PIC, EMT (GR2/NG2 EMT = EMT).
    - RW staffed: RN + Medic + (PIC or EMT); GR staffed: RN + Medic + EMT.

    Returns (rw_day, rw_night, gr_day, gr_night) per base_name.
    """
    wb = load_workbook(path, data_only=True)
    try:
        sn = resolve_workbook_sheet(wb, "OPS View", "Ops View")
        if not sn:
            return {}, {}, {}, {}
        return _parse_ops_view_worksheet(wb[sn], week_start)
    finally:
        wb.close()


def _ops_staffed_for_column(
    ws: Worksheet,
    col_idx: int,
    service_type: ServiceType,
    rn_rows: list[int],
    medic_rows: list[int],
    pic_rows: list[int],
    emt_rows: list[int],
) -> bool:
    rn_ok = any(
        _ops_cell_staffed(ws.cell(row=r, column=col_idx).value) for r in rn_rows
    )
    medic_ok = any(
        _ops_cell_staffed(ws.cell(row=r, column=col_idx).value) for r in medic_rows
    )
    pic_ok = any(
        _ops_cell_staffed(ws.cell(row=r, column=col_idx).value) for r in pic_rows
    )
    emt_ok = any(
        _ops_cell_staffed(ws.cell(row=r, column=col_idx).value) for r in emt_rows
    )
    if service_type == "RW":
        return rn_ok and medic_ok and (pic_ok or emt_ok)
    return rn_ok and medic_ok and emt_ok


def _parse_ops_view_daily_worksheet(
    ws: Worksheet,
    week_start: str,
) -> dict[date, tuple[int, int]]:
    """Per calendar day: (RW staffed unit-days, GR staffed unit-days) from OPS View."""
    try:
        week_start_date = datetime.strptime(week_start, "%Y-%m-%d").date()
    except ValueError:
        return {}
    week_end_date = week_start_date + timedelta(days=6)

    col_dates = _ops_parse_dates_row(ws, week_start_date, week_end_date)
    if not col_dates:
        return {}

    daily: dict[date, list[int]] = {d: [0, 0] for _, d in col_dates}
    for unit_code, _first_row, role_rows in _ops_vehicle_blocks(ws):
        unit_info = UNIT_MAP.get(unit_code)
        if not unit_info:
            continue
        base, service_type, _dn = unit_info
        rn_rows = role_rows.get("RN") or []
        medic_rows = role_rows.get("Medic") or []
        pic_rows = role_rows.get("PIC") or []
        emt_rows = role_rows.get("EMT") or []

        for col_idx, day_date in col_dates:
            if not _ops_staffed_for_column(
                ws,
                col_idx,
                service_type,
                rn_rows,
                medic_rows,
                pic_rows,
                emt_rows,
            ):
                continue
            if service_type == "RW":
                daily[day_date][0] += 1
            else:
                daily[day_date][1] += 1

    return {d: (vals[0], vals[1]) for d, vals in daily.items()}


def parse_ops_view_daily(
    path: str, week_start: str, wb: Workbook | None = None
) -> dict[date, tuple[int, int]]:
    """Load workbook and return per-day RW/GR staffed counts from OPS View."""
    with _open_workbook(path, wb) as wb:
        sn = resolve_workbook_sheet(wb, "OPS View", "Ops View")
        if not sn:
            return {}
        return _parse_ops_view_daily_worksheet(wb[sn], week_start)


def _ops_role_label(text_b: str) -> str | None:
    if text_b in _OPS_ROLE_RN:
        return "RN"
    if text_b in _OPS_ROLE_MEDIC:
        return "Medic"
    if text_b in _OPS_ROLE_PIC:
        return "PIC"
    if text_b in _OPS_ROLE_EMT:
        return "EMT"
    return None


def _parse_ops_view_detail_worksheet(
    ws: Worksheet,
    week_start: str,
) -> tuple[list[OpsViewDayBase], list[OpsViewAssignment]]:
    """Per-day per-base counts and name-level OPS View assignments."""
    try:
        week_start_date = datetime.strptime(week_start, "%Y-%m-%d").date()
    except ValueError:
        return [], []
    week_end_date = week_start_date + timedelta(days=6)

    col_dates = _ops_parse_dates_row(ws, week_start_date, week_end_date)
    if not col_dates:
        return [], []

    day_base: dict[tuple[date, str], list[int]] = {
        (d, base): [0, 0]
        for _, d in col_dates
        for base in {info[0] for info in UNIT_MAP.values()}
    }
    assignments: list[OpsViewAssignment] = []

    for unit_code, _first_row, role_rows in _ops_vehicle_blocks(ws):
        unit_info = UNIT_MAP.get(unit_code)
        if not unit_info:
            continue
        base, service_type, _dn = unit_info
        for role_key, row_list in role_rows.items():
            role_label = role_key
            for row_idx in row_list:
                for col_idx, day_date in col_dates:
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    staffed = _ops_cell_staffed(cell_val)
                    raw_s = _normalize_cell_value(cell_val)
                    if raw_s:
                        assignments.append(
                            OpsViewAssignment(
                                day_date=day_date,
                                unit_code=unit_code,
                                role=role_label,
                                excel_row=row_idx,
                                person_display=raw_s,
                                raw_value=raw_s,
                                is_staffed=staffed,
                            )
                        )
        for col_idx, day_date in col_dates:
            staffed_slot = _ops_staffed_for_column(
                ws,
                col_idx,
                service_type,
                role_rows.get("RN") or [],
                role_rows.get("Medic") or [],
                role_rows.get("PIC") or [],
                role_rows.get("EMT") or [],
            )
            if not staffed_slot:
                continue
            key = (day_date, base)
            if key not in day_base:
                day_base[key] = [0, 0]
            if service_type == "RW":
                day_base[key][0] += 1
            else:
                day_base[key][1] += 1

    day_rows = [
        OpsViewDayBase(
            day_date=day_date,
            base_name=base_name,
            rw_count=vals[0],
            gr_count=vals[1],
        )
        for (day_date, base_name), vals in sorted(day_base.items())
        if vals[0] or vals[1]
    ]
    return day_rows, assignments


def parse_ops_view_detail(
    path: str,
    week_start: str,
    wb: Workbook | None = None,
) -> tuple[list[OpsViewDayBase], list[OpsViewAssignment]]:
    """Load workbook and return OPS View day/base counts and assignments."""
    with _open_workbook(path, wb) as wb:
        sn = resolve_workbook_sheet(wb, "OPS View", "Ops View")
        if not sn:
            return [], []
        return _parse_ops_view_detail_worksheet(wb[sn], week_start)
