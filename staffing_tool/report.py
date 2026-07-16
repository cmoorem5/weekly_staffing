"""
Weekly staffing summary Excel export: Board_Summary, Weekly_Detail, Trend_12_Weeks, Data_Dump.
Uses openpyxl; status fills are computed in Python (not Excel conditional formatting).

This module is the public façade: it keeps ``export_board_pack`` /
``export_week_excel`` and the Trend/Data_Dump sheet writers, and re-exports
every name from the modules the generator was split into — import from
here, not the submodules:

- ``report_excel_style``   brand fills/fonts/borders + shared cell writers
- ``report_data``          week loading, metric prep, status text, config checks
- ``report_board_summary`` Board_Summary sheet writer + narrative
- ``report_weekly_detail`` Weekly_Detail sheet writer + §7 self-checks
"""

import os

from openpyxl import Workbook
from sqlalchemy.orm import Session

from .db import session_scope
from .metrics import (
    WeekMetrics,
    compute_period_rollups,
    compute_week_metrics,
    get_metric_value,
)
from .models import (
    BaseConfig,
    KpiThreshold,
    WeeklyLeaveDetail,
    WeeklyStaffing,
)
from .report_board_summary import (  # noqa: F401
    _generate_narrative,
    _write_board_summary,
)
from .report_data import (  # noqa: F401
    BASE_UNIT_CELL_CONFIGURED,
    DETAIL_BASE_ORDER,
    EXCEPTION_ROLES,
    LEAVE_TYPE_COLS,
    RW_SYSTEM_WEEKLY_DENOMINATOR,
    _assert_rw_config_rw_cap_56,
    _averages,
    _exc_count_breakdown,
    _leave_totals_from_breakdown,
    _load_week_with_coverage,
    _load_weeks_ordered,
    _metrics_for_weeks,
    _parse_week,
    _rag_for_metric,
    _status_display,
    _week_end,
)
from .report_excel_style import (  # noqa: F401
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    BMF_BLACK,
    BMF_BLUE,
    BMF_GRAY,
    BMF_MEDIUM_GRAY,
    BMF_NAVY,
    BMF_RED,
    BMF_WHITE,
    BOLD,
    FILL_BAND_ALT,
    FILL_BAND_TOTAL,
    FILL_BMF_BLUE,
    FILL_BMF_GRAY_BG,
    FILL_BMF_NAVY,
    FILL_GREEN,
    FILL_GREEN_FULL,
    FILL_GREEN_SOFT,
    FILL_HEADER_DARK,
    FILL_HEADER_LIGHT,
    FILL_RAG_RED,
    FILL_RED,
    FILL_RED_FULL,
    FILL_RED_SOFT,
    FILL_WHITE_SOLID,
    FILL_YELLOW,
    FONT_BMF_BODY,
    FONT_BMF_BODY_BOLD,
    FONT_BMF_RAG_VALUE,
    FONT_BMF_RAG_VALUE_ON_RED,
    FONT_BMF_SECTION,
    FONT_BMF_SUBTITLE,
    FONT_BMF_TITLE,
    FONT_DETAIL_TITLE,
    FONT_FOOTER_META,
    FONT_GENERATED,
    FONT_LEGEND_LABEL,
    FONT_NA,
    FONT_NAME,
    FONT_SUBTITLE_MUTED,
    FONT_WEEK_BADGE,
    FONT_ZERO_MUTED,
    SIDE_SEP,
    THIN_BORDER,
    WHITE_BOLD,
    _add_logo,
    _base_gr_cell_notable,
    _base_rw_cell_notable,
    _bmf_border_block,
    _bmf_cell_border,
    _bmf_merge_band,
    _border_right_sep_cell,
    _detail_section_banner,
    _fill_and_font_for_status,
    _generator_version_label,
    _generator_version_string,
    _iso_week_label,
    _kpi_notable,
    _moderate_red_soft,
    _new_workbook_from_template_or_empty,
    _project_root,
    _report_template_path,
    _resolve_logo_path,
    _target_display,
    _write_na_or_int,
    _write_pct_cell,
    _write_pct_or_num,
)
from .report_weekly_detail import (  # noqa: F401
    _verify_weekly_detail_checks,
    _write_weekly_detail,
)


def _write_trend_sheet(
    wb: Workbook,
    trend_list: list[tuple[str, WeekMetrics]],
    thresholds: dict[str, KpiThreshold],
) -> None:
    ws = wb.create_sheet("Trend_12_Weeks", 2)
    headers = [
        "Week Start",
        "Staffing Rate",
        "Backfill Rate",
        "Shift Exception %",
        "System RW %",
        "System GR %",
        "Status",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h).font = BOLD
    thr_sr = thresholds.get("Staffing Rate")
    row = 2
    for week_start, m in trend_list:
        ws.cell(row, 1, week_start)
        ws.cell(row, 2, m.staffing_rate).number_format = "0.0%"
        ws.cell(row, 3, m.ot_dependency).number_format = "0.0%"
        ws.cell(row, 4, m.leave_exposure).number_format = "0.0%"
        ws.cell(row, 5, m.system_rw_pct).number_format = "0.0%"
        ws.cell(row, 6, m.system_gr_pct).number_format = "0.0%"
        rag = _rag_for_metric("Staffing Rate", m.staffing_rate, thresholds)
        st = ws.cell(row, 7, _status_display(rag))
        notable = _kpi_notable("Staffing Rate", m.staffing_rate, m)
        fill, font = _fill_and_font_for_status(
            rag,
            notable=notable,
            value=m.staffing_rate,
            thr=thr_sr,
        )
        st.fill = fill
        st.font = font
        st.border = THIN_BORDER
        row += 1
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12


def _write_data_dump(wb: Workbook, session: Session) -> None:
    ws = wb.create_sheet("Data_Dump", 3)
    rows = session.query(WeeklyStaffing).order_by(WeeklyStaffing.week_start).all()
    if not rows:
        return
    # Headers from first row
    attrs = [
        "week_start",
        "day_target",
        "night_min",
        "filled_day",
        "filled_night",
        "ot_shifts",
        "leave_at",
        "leave_lt",
        "leave_sick",
        "leave_loa",
        "leave_pfml",
        "medic_unpartnered",
        "rn_unpartnered_staff",
        "overnights_below",
        "pilot_vacancies",
        "notes",
        "entered_by",
        "created_at",
        "updated_at",
    ]
    for c, a in enumerate(attrs, 1):
        ws.cell(1, c, a).font = BOLD
    for r, row in enumerate(rows, 2):
        for c, a in enumerate(attrs, 1):
            val = getattr(row, a, None)
            ws.cell(r, c, val)
    ws.freeze_panes = "A2"


def export_board_pack(
    db_path: str | None,
    week_start: str,
    trend_weeks: int = 12,
    output_dir: str = "output",
    metadata: dict | None = None,
) -> str:
    """
    Generate Weekly_staffing_summary_<week_start>_to_<week_end>.xlsx.
    Returns path to the written file.
    """
    os.makedirs(output_dir, exist_ok=True)
    week_end = _week_end(week_start)
    filename = f"Weekly_staffing_summary_{week_start}_to_{week_end}.xlsx"
    filepath = os.path.join(output_dir, filename)

    with session_scope(db_path) as session:
        thresholds = {t.metric_name: t for t in session.query(KpiThreshold).all()}
        bases_all = session.query(BaseConfig).order_by(BaseConfig.base_name).all()
        _assert_rw_config_rw_cap_56(bases_all)
        data = _load_week_with_coverage(session, week_start)
        if not data:
            raise ValueError(f"No weekly data for week_start={week_start}")

        row, coverages, bases = data
        leave_details = (
            session.query(WeeklyLeaveDetail)
            .filter(WeeklyLeaveDetail.week_start == week_start)
            .all()
        )
        leave_breakdown = {(r.role, r.leave_type): r.count for r in leave_details}
        this_metrics = compute_week_metrics(row, coverages, bases)

        week_starts = _load_weeks_ordered(
            session, trend_weeks, through_week_start=week_start
        )
        trend_list = _metrics_for_weeks(
            session, week_starts
        )  # list of (week_start, WeekMetrics, RAG)
        trend_metrics = [m for _, m, _ in trend_list]
        prior_metrics = None
        idx_this = next(
            (i for i, (ws, _, _) in enumerate(trend_list) if ws == week_start), None
        )
        if idx_this is not None and idx_this > 0:
            prior_metrics = trend_list[idx_this - 1][1]

        last_4 = trend_metrics[-4:] if len(trend_metrics) >= 4 else trend_metrics
        avg_4w = _averages(last_4) if last_4 else None
        avg_12w = _averages(trend_metrics) if trend_metrics else None
        rollups_4w = compute_period_rollups(last_4) if last_4 else None
        rollups_12w = compute_period_rollups(trend_metrics) if trend_metrics else None

        rag_statuses = {}
        for name in [
            "Staffing Rate",
            "OT Dependency",
            "Shift Exception %",
            "System RW Coverage %",
            "System GR Coverage %",
        ]:
            v = get_metric_value(this_metrics, name)
            if v is not None:
                rag_statuses[name] = _rag_for_metric(name, v, thresholds)

        narrative = _generate_narrative(
            this_metrics, prior_metrics, avg_4w, thresholds, rag_statuses
        )

        wb = _new_workbook_from_template_or_empty()

        trend_data = [(ws, m) for (ws, m, _) in trend_list]
        _write_board_summary(
            wb,
            week_start,
            week_end,
            this_metrics,
            prior_metrics,
            avg_4w,
            avg_12w,
            rollups_4w,
            rollups_12w,
            trend_data,
            thresholds,
            narrative,
            metadata=metadata,
        )
        _write_weekly_detail(
            wb,
            week_start,
            week_end,
            row,
            this_metrics,
            bases,
            leave_breakdown=leave_breakdown,
            thresholds=thresholds,
        )
        _write_trend_sheet(wb, trend_data, thresholds)
        _write_data_dump(wb, session)

        wb.save(filepath)

        # Verification: reconciliation checks (console output)
        print("Verification (Weekly Staffing Detail):")
        for name, passed, msg in _verify_weekly_detail_checks(this_metrics, row):
            status = "PASS" if passed else "FAIL"
            print(f"  [{status}] {name}: {msg}")
        print(f"  -> Saved: {filepath}")

    return filepath


def export_week_excel(
    db_path: str | None,
    week_start: str,
    output_dir: str = "output",
) -> str:
    """Single-week export (simplified one-sheet or same structure). Reuse board pack with 1 week trend."""
    return export_board_pack(db_path, week_start, trend_weeks=1, output_dir=output_dir)
