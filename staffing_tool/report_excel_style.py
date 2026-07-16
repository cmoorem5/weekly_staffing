"""BMF brand styles + Excel style helpers for the board-pack workbook.

Status colors are computed in code — never Excel conditional formatting
(docs/report-generator-spec.md).
"""

from __future__ import annotations

import os
import subprocess
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from . import __version__ as STAFFING_TOOL_VERSION
from .metrics import (
    WeekMetrics,
)
from .models import (
    KpiThreshold,
)
from .rag import RAG

# ----- Boston MedFlight brand (Clinical Operations guidelines) -----
# Colors: Blue #2a4492, Navy #052c47, Gray #e6e6e6, Medium Gray #cbc7d1, White #ffffff, Black #000000, Red #c12126
# Typography: Barlow; if unavailable Excel falls back (often Calibri) — Arial/Helvetica acceptable per brand.
BMF_BLUE = "2A4492"
BMF_NAVY = "052C47"
BMF_GRAY = "E6E6E6"
BMF_MEDIUM_GRAY = "CBC7D1"
BMF_WHITE = "FFFFFF"
BMF_BLACK = "000000"
BMF_RED = "C12126"

FONT_NAME = "Barlow"

# §2.2 severity gradient (fills)
FILL_GREEN_SOFT = PatternFill(
    start_color="EAF5E9", end_color="EAF5E9", fill_type="solid"
)
FILL_GREEN_FULL = PatternFill(
    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
)
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RED_SOFT = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
FILL_RED_FULL = PatternFill(start_color=BMF_RED, end_color=BMF_RED, fill_type="solid")
# Back-compat aliases
FILL_GREEN = FILL_GREEN_FULL
FILL_RAG_RED = FILL_RED_FULL
FILL_RED = FILL_RED_FULL

FILL_BMF_NAVY = PatternFill(start_color=BMF_NAVY, end_color=BMF_NAVY, fill_type="solid")
FILL_BMF_BLUE = PatternFill(start_color=BMF_BLUE, end_color=BMF_BLUE, fill_type="solid")
FILL_BMF_GRAY_BG = PatternFill(
    start_color=BMF_GRAY, end_color=BMF_GRAY, fill_type="solid"
)
# Table column headers: BMF Gray (guidelines); section bars use Navy or Blue
FILL_HEADER_LIGHT = FILL_BMF_GRAY_BG

FILL_HEADER_DARK = FILL_BMF_NAVY

FONT_BMF_TITLE = Font(name=FONT_NAME, size=18, bold=True, color=BMF_WHITE)
FONT_BMF_SUBTITLE = Font(name=FONT_NAME, size=11, bold=True, color=BMF_NAVY)
FONT_BMF_SECTION = Font(name=FONT_NAME, size=11, bold=True, color=BMF_WHITE)
FONT_BMF_BODY = Font(name=FONT_NAME, size=11, color=BMF_BLACK)
FONT_BMF_BODY_BOLD = Font(name=FONT_NAME, size=11, bold=True, color=BMF_BLACK)
# RAG-filled cells: strong contrast (percent values)
FONT_BMF_RAG_VALUE = Font(name=FONT_NAME, size=11, bold=True, color=BMF_BLACK)
FONT_BMF_RAG_VALUE_ON_RED = Font(name=FONT_NAME, size=11, bold=True, color=BMF_WHITE)

BOLD = Font(name=FONT_NAME, bold=True, color=BMF_BLACK)
WHITE_BOLD = Font(name=FONT_NAME, bold=True, color=BMF_WHITE)
FONT_NA = Font(name=FONT_NAME, size=11, italic=True, color="999999")
FONT_SUBTITLE_MUTED = Font(name=FONT_NAME, size=11, italic=True, color=BMF_MEDIUM_GRAY)
FONT_GENERATED = Font(name=FONT_NAME, size=9, italic=True, color=BMF_MEDIUM_GRAY)
FONT_FOOTER_META = Font(name=FONT_NAME, size=9, italic=True, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin", color=BMF_MEDIUM_GRAY),
    right=Side(style="thin", color=BMF_MEDIUM_GRAY),
    top=Side(style="thin", color=BMF_MEDIUM_GRAY),
    bottom=Side(style="thin", color=BMF_MEDIUM_GRAY),
)

FILL_BAND_ALT = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
FILL_BAND_TOTAL = PatternFill(
    start_color=BMF_GRAY, end_color=BMF_GRAY, fill_type="solid"
)
SIDE_SEP = Side(style="thin", color=BMF_MEDIUM_GRAY)


def _moderate_red_soft(value: float, t: KpiThreshold) -> bool:
    """
    Soft red vs full red: 'moderate' when the value is not in the worst part of the red band
    (roughly 40–80% of the way from yellow boundary toward the worst edge → soft).
    """
    hi = (t.higher_is_better or 0) != 0
    if hi:
        # Higher is better → red on the low side
        edge = t.yellow_min
        if edge is None or edge <= 0:
            return True
        if value >= edge:
            return True
        depth = (edge - value) / edge
        return depth < 0.8
    # Lower is better → red on the high side
    edge = t.yellow_max
    if edge is None:
        return True
    if value <= edge:
        return True
    span = max((t.red_max or 1.0) - edge, 1e-9)
    depth = (value - edge) / span
    return depth < 0.8


def _kpi_notable(metric_key: str, val: float, this_metrics: WeekMetrics) -> bool:
    """§2.2: notable → full green when On target."""
    if metric_key == "System GR Coverage %":
        return val >= 0.95
    if metric_key == "System RW Coverage %":
        return val >= 0.95
    if metric_key in ("Staffing Rate", "OT Dependency", "Shift Exception %"):
        return False
    return False


def _base_rw_cell_notable(rw_pct_val: float) -> bool:
    """Per-base RW% cell: notable at 100% coverage."""
    return rw_pct_val >= 1.0 - 1e-9


def _base_gr_cell_notable(gr_pct_val: float) -> bool:
    """Per-base GR% cell: notable at ≥95% (spec) or 100%."""
    return gr_pct_val >= 0.95 - 1e-9


def _fill_and_font_for_status(
    rag: RAG,
    *,
    notable: bool,
    value: float,
    thr: KpiThreshold | None,
) -> tuple[PatternFill, Font]:
    if rag == "Green":
        fill = FILL_GREEN_FULL if notable else FILL_GREEN_SOFT
        return fill, FONT_BMF_RAG_VALUE
    if rag == "Yellow":
        return FILL_YELLOW, FONT_BMF_RAG_VALUE
    if thr is not None and _moderate_red_soft(value, thr):
        return FILL_RED_SOFT, FONT_BMF_RAG_VALUE
    return FILL_RED_FULL, FONT_BMF_RAG_VALUE_ON_RED


def _target_display(metric_key: str, t: KpiThreshold | None) -> str:
    """Green-threshold hint for Target column (§1.4 display)."""
    if not t:
        return "—"
    hi = (t.higher_is_better or 0) != 0
    if hi:
        g = t.green_min
        if g is not None:
            return f"≥ {g * 100:.0f}%"
    else:
        g = t.green_max
        if g is not None:
            return f"≤ {g * 100:.0f}%"
    return "—"


def _generator_version_label() -> str:
    if STAFFING_TOOL_VERSION:
        return STAFFING_TOOL_VERSION
    return _generator_version_string()


FILL_WHITE_SOLID = PatternFill(
    start_color=BMF_WHITE, end_color=BMF_WHITE, fill_type="solid"
)
FONT_ZERO_MUTED = Font(name=FONT_NAME, size=11, color=BMF_MEDIUM_GRAY)
FONT_DETAIL_TITLE = Font(name=FONT_NAME, size=16, bold=True, color=BMF_WHITE)
FONT_WEEK_BADGE = Font(name=FONT_NAME, size=10, color=BMF_WHITE)
FONT_LEGEND_LABEL = Font(name=FONT_NAME, size=10, bold=True, color=BMF_NAVY)


def _iso_week_label(week_start: str) -> str:
    d = datetime.strptime(week_start, "%Y-%m-%d").date()
    ic = d.isocalendar()
    week = ic.week if hasattr(ic, "week") else ic[1]
    year = ic.year if hasattr(ic, "year") else ic[0]
    return f"Week {week} · {year}"


def _detail_section_banner(ws, row: int, title: str, subtitle: str) -> None:
    """Weekly_Detail: merge A:F (title) + G (subtitle), navy bar."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    a = ws.cell(row, 1, title)
    a.fill = FILL_BMF_NAVY
    a.font = Font(name=FONT_NAME, size=12, bold=True, color=BMF_WHITE)
    a.alignment = ALIGN_LEFT
    a.border = THIN_BORDER
    g = ws.cell(row, 7, subtitle)
    g.fill = FILL_BMF_NAVY
    g.font = Font(name=FONT_NAME, size=10, italic=True, color=BMF_MEDIUM_GRAY)
    g.alignment = ALIGN_RIGHT
    g.border = THIN_BORDER


def _border_right_sep_cell(ws, row: int, col: int) -> None:
    """§5.5: thin vertical separator on the right edge of `col`."""
    c = ws.cell(row, col)
    prev = c.border
    c.border = Border(
        left=prev.left if prev else THIN_BORDER.left,
        right=SIDE_SEP,
        top=prev.top if prev else THIN_BORDER.top,
        bottom=prev.bottom if prev else THIN_BORDER.bottom,
    )


def _write_na_or_int(ws, row: int, col: int, configured: bool, n: int) -> None:
    """§4.1: unconfigured base/unit/shift cells show gray italic N/A."""
    if not configured:
        c = _bmf_cell_border(ws, row, col, "N/A", align=ALIGN_CENTER)
        c.font = FONT_NA
    else:
        _bmf_cell_border(ws, row, col, n, align=ALIGN_CENTER)


def _project_root() -> str:
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _generator_version_string() -> str:
    try:
        root = _project_root()
        return subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=root,
            stderr=subprocess.DEVNULL,
            text=True,
        ).strip()
    except (OSError, subprocess.CalledProcessError):
        return "unknown"


def _report_template_path() -> str:
    """Optional shell workbook: output/Weekly_staffing _report template.xlsx (project root)."""
    return os.path.join(
        _project_root(), "output", "Weekly_staffing _report template.xlsx"
    )


def _resolve_logo_path() -> str | None:
    """
    Boston MedFlight coastal logo PNG for Excel (optional).
    Set WEEKLY_STAFFING_LOGO to a file path, or place assets/bmf_coastal_logo.png in the project root.
    """
    env = os.environ.get("WEEKLY_STAFFING_LOGO", "").strip()
    if env and os.path.isfile(env):
        return env
    root = _project_root()
    candidates = [
        os.path.join(root, "assets", "bmf_coastal_logo.png"),
        os.path.join(
            root,
            "bmf_staffing",
            "dashboard",
            "static",
            "dashboard",
            "images",
            "BMF_Coastal_Logos.png",
        ),
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    return None


def _add_logo(ws, anchor: str = "A1", max_height_px: int = 88) -> bool:
    """Embed logo image if Pillow + file exist; returns True when an image was added."""
    path = _resolve_logo_path()
    if not path:
        return False
    try:
        from openpyxl.drawing.image import Image as XLImage
    except ImportError:
        return False
    try:
        img = XLImage(path)
        if img.height and img.height > max_height_px:
            scale = max_height_px / float(img.height)
            img.width = int(img.width * scale)
            img.height = max_height_px
        ws.add_image(img, anchor)
        return True
    except Exception:
        return False


def _new_workbook_from_template_or_empty() -> Workbook:
    """
    If the template file exists, load it and remove all sheets (use as branded shell / theme).
    Otherwise a fresh Workbook. Caller adds Board_Summary, Weekly_Detail, etc.
    """
    path = os.environ.get("WEEKLY_STAFFING_REPORT_TEMPLATE", _report_template_path())
    if os.path.isfile(path):
        try:
            wb = load_workbook(path)
            for name in list(wb.sheetnames):
                wb.remove(wb[name])
            return wb
        except Exception:
            pass
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def _bmf_merge_band(
    ws,
    row: int,
    col_start: int,
    col_end: int,
    text: str,
    *,
    fill: PatternFill,
    font: Font,
    alignment: Alignment | None = None,
) -> None:
    """Merge a row band and apply BMF header styling."""
    top_left = ws.cell(row, col_start, text)
    if col_end > col_start:
        ws.merge_cells(
            start_row=row, start_column=col_start, end_row=row, end_column=col_end
        )
    top_left.fill = fill
    top_left.font = font
    top_left.alignment = alignment or ALIGN_CENTER


def _bmf_cell_border(ws, row: int, col: int, value, font=None, fill=None, align=None):
    c = ws.cell(row, col, value)
    c.font = font or FONT_BMF_BODY
    c.border = THIN_BORDER
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    else:
        c.alignment = ALIGN_LEFT
    return c


def _bmf_border_block(ws, r1: int, r2: int, c1: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = THIN_BORDER


def _write_pct_or_num(
    ws, row: int, col: int, value: float | None, metric_name: str
) -> None:
    if value is None:
        ws.cell(row, col, "").border = THIN_BORDER
        return
    cell = ws.cell(row, col)
    cell.border = THIN_BORDER
    percent_metrics = {
        "Staffing Rate",
        "OT Dependency",
        "Backfill Rate",
        "Shift Exception %",
        "System RW Coverage %",
        "System GR Coverage %",
    }
    if metric_name in percent_metrics:
        cell.value = value
        cell.number_format = "0.0%"
    else:
        cell.value = value if isinstance(value, (int, float)) else value
        if isinstance(value, float) and value == int(value):
            cell.value = int(value)


def _write_pct_cell(
    ws,
    row: int,
    col: int,
    value: float,
    rag: RAG,
    *,
    thr: KpiThreshold | None = None,
    notable: bool = False,
) -> None:
    cell = ws.cell(row, col)
    cell.value = value
    cell.number_format = "0.0%"
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_CENTER
    fill, font = _fill_and_font_for_status(rag, notable=notable, value=value, thr=thr)
    cell.fill = fill
    cell.font = font
