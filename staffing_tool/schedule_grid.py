"""RN/Medic/EMT grid parsing: walk each sheet block cell by cell."""

from __future__ import annotations

from datetime import date, timedelta

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .manager_roster import default_manager_last_names_upper
from .person_names import person_displays_for_role
from .schedule_cells import (
    AT_ALIASES,
    IGNORE_UNIT_CODES,
    LEAVE_CODES,
    RETIRED_UNIT_CODES,
    SKIP_CELL_VALUES,
    SKIP_TRAINING_VALUES,
    UNIT_LEAVE_MERGE,
    _canonical_unit_code,
    _classify_skip_reason,
    _classify_unit,
    _is_resolvable_unit,
    _normalize_cell_value,
    _split_unit_suffix,
)
from .schedule_types import ParseIssue, Role, ShiftRecord, SkipReason
from .schedule_workbook import _best_header_row_in_ws, _header_cell_to_date

# Training / admin / manager markers: skip cell entirely
# (no staffed, no leave, no issue).
# RN & Medic / EMT grids end with an "OPEN" row (unfilled units per day) and
# an "EXTRA" row (float staff names per day), followed by footer/notes rows.
# None of that is a person's schedule, so once we hit OPEN or EXTRA in a
# row's name column, that row and everything below it (within the block) is
# skipped. Detected by label text rather than a fixed row number: adding
# staff above these rows pushes them down the sheet every time a roster
# grows, so a hardcoded row number silently goes stale and lets the OPEN/
# EXTRA row's contents (unit-code lists, staff names) leak into the
# "unknown unit codes" review as if they were shift codes.
NON_PERSON_ROW_LABELS = frozenset({"OPEN", "EXTRA"})
_SCHEDULE_COL_B = 2
_SCHEDULE_COL_P = 16


def _find_non_person_skip_row(
    ws: Worksheet, first_row_idx: int, last_row_idx: int
) -> int | None:
    """First row in range labeled OPEN or EXTRA in its name column (A or B).

    Everything from that row through the end of the block is summary/
    footer content, not a person's schedule.
    """
    for row_idx in range(first_row_idx, last_row_idx + 1):
        raw_a, raw_b = _grid_name_cells(ws, row_idx)
        label = (raw_a or raw_b or "").strip().upper()
        if label in NON_PERSON_ROW_LABELS:
            return row_idx
    return None


def _append_skipped_shift(
    records: list[ShiftRecord],
    *,
    ws: Worksheet,
    row_idx: int,
    col_idx: int,
    d: date,
    role: Role,
    sheet_label: str,
    text: str,
    person_displays: tuple[str, ...],
    person_display: str,
    is_manager_row: bool,
    skip_reason: SkipReason,
) -> None:
    cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
    # Training is the one skip category that still counts toward a weekly
    # total (WeeklyStaffing.training_shifts) -- everything else here is
    # truly dropped. Manager-row cells are excluded, matching how manager
    # leave/OT are tracked separately (weekly_manager_shifts) rather than
    # folded into the staff weekly totals.
    included_in_aggregates = skip_reason == "training" and not is_manager_row
    records.append(
        ShiftRecord(
            date=d,
            base="",
            service_type="",
            day_night="D",
            role=role,
            filled=False,
            overtime=False,
            leave_type=None,
            source_tab=sheet_label,
            source_cell=cell_ref,
            raw_value=text,
            person_display=person_display,
            person_displays=person_displays,
            is_manager_row=is_manager_row,
            skip_reason=skip_reason,
            included_in_aggregates=included_in_aggregates,
            excel_row=row_idx,
            excel_col=col_idx,
        )
    )


def _name_tokens_for_grid_row(ws: Worksheet, row_idx: int) -> set[str]:
    """Tokens from columns A–B (last/first names) for manager filtering."""
    tokens: set[str] = set()
    for col in (1, 2):
        name_cell = ws.cell(row=row_idx, column=col).value
        name_key = (_normalize_cell_value(name_cell) or "").strip()
        if not name_key:
            continue
        tokens.update(t.strip(".,").upper() for t in name_key.split() if t.strip(".,"))
    return tokens


def _grid_name_cells(ws: Worksheet, row_idx: int) -> tuple[str, str]:
    def _cell_str(v: object) -> str:
        if v is None:
            return ""
        return str(v).strip()

    return (
        _cell_str(ws.cell(row=row_idx, column=1).value),
        _cell_str(ws.cell(row=row_idx, column=2).value),
    )


def _row_person_displays_and_manager_flag(
    ws: Worksheet,
    row_idx: int,
    role: Role,
    manager_last_names_upper: frozenset[str],
) -> tuple[tuple[str, ...], bool]:
    """Clean person label(s) from columns A–B and manager-roster flag."""
    tokens = _name_tokens_for_grid_row(ws, row_idx)
    roster_hit = (
        tokens & manager_last_names_upper if manager_last_names_upper else set()
    )
    if roster_hit:
        return (min(roster_hit).title(),), True

    raw_a, raw_b = _grid_name_cells(ws, row_idx)
    persons = person_displays_for_role(role, raw_a, raw_b)
    return persons, False


def _parse_grid(
    ws: Worksheet,
    header_row_idx: int,
    first_row_idx: int,
    last_row_idx: int,
    role: Role,
    sheet_label: str,
    week_start_date: date | None = None,
    week_end_date: date | None = None,
    unit_overrides: dict[str, str] | None = None,
    manager_last_names_upper: frozenset[str] | None = None,
    extra_training_codes: frozenset[str] | None = None,
) -> tuple[list[ShiftRecord], list[ParseIssue]]:
    """Parse a rectangular RN/Medic/EMT grid into shift records + issues.

    unit_overrides: map raw_value -> replacement
    (e.g. {"D7BCP": "D7BC", "TYPO": "D7B/LT"}).
    Use at import to fix typos or merge unknown codes.

    manager_last_names_upper: roster from staffing.db (or caller default); when
    omitted, uses the built-in default list.

    extra_training_codes: admin-added codes (Settings > Training codes) on
    top of the built-in SKIP_TRAINING_VALUES -- e.g. a new class name that
    would otherwise show up as an unknown unit code.
    """
    mgr_upper = (
        manager_last_names_upper
        if manager_last_names_upper is not None
        else default_manager_last_names_upper()
    )
    records: list[ShiftRecord] = []
    issues: list[ParseIssue] = []
    overrides = unit_overrides or {}
    training_values = SKIP_TRAINING_VALUES | (extra_training_codes or frozenset())
    skip_cell_values = SKIP_CELL_VALUES | (extra_training_codes or frozenset())

    header_row_idx, col_dates = _best_header_row_in_ws(
        ws,
        week_start_date=week_start_date,
        week_end_date=week_end_date,
        prefer_row=header_row_idx,
    )
    if not col_dates:
        default_y = (
            week_start_date.year if week_start_date is not None else date.today().year
        )
        raw_dates: list[date] = []
        for scan_r in range(1, 6):
            for cell in ws[scan_r]:
                if cell.column < 3 or cell.column > 16:
                    continue
                d = _header_cell_to_date(cell.value, default_y)
                if d is not None:
                    raw_dates.append(d)
        if week_start_date is not None and week_end_date is not None and raw_dates:
            lo, hi = min(raw_dates), max(raw_dates)
            sun = lo - timedelta(days=(lo.weekday() + 1) % 7)
            suggested = sun.isoformat()
            issues.append(
                ParseIssue(
                    sheet=ws.title,
                    cell=f"row {header_row_idx} C:P",
                    raw_value="",
                    issue_type="week_mismatch",
                    message=(
                        f"No columns fall in the selected week "
                        f"{week_start_date.isoformat()} through "
                        f"{week_end_date.isoformat()}. "
                        f"This sheet shows dates {lo.isoformat()} through "
                        f"{hi.isoformat()}. "
                        f"Use Sunday week_start={suggested} for this file, "
                        f"or choose a schedule export whose header dates "
                        f"include your week."
                    ),
                )
            )
        else:
            issues.append(
                ParseIssue(
                    sheet=ws.title,
                    cell=f"A{header_row_idx}",
                    raw_value="",
                    issue_type="no_dates",
                    message="No date headers found in RN/Medic grid header row.",
                )
            )
        return records, issues

    skip_from_row_idx = _find_non_person_skip_row(ws, first_row_idx, last_row_idx)

    for row_idx in range(first_row_idx, last_row_idx + 1):
        person_displays, is_manager_row = _row_person_displays_and_manager_flag(
            ws, row_idx, role, mgr_upper
        )
        person_display = person_displays[0] if person_displays else ""
        for col_idx, d in col_dates:
            cell = ws.cell(row=row_idx, column=col_idx)
            raw = cell.value
            text = _normalize_cell_value(raw)
            if not text:
                continue
            if (
                skip_from_row_idx is not None
                and row_idx >= skip_from_row_idx
                and _SCHEDULE_COL_B <= col_idx <= _SCHEDULE_COL_P
            ):
                _append_skipped_shift(
                    records,
                    ws=ws,
                    row_idx=row_idx,
                    col_idx=col_idx,
                    d=d,
                    role=role,
                    sheet_label=sheet_label,
                    text=text,
                    person_displays=person_displays,
                    person_display=person_display,
                    is_manager_row=is_manager_row,
                    skip_reason="schedule_row",
                )
                continue
            # Apply import mapping: unknown code -> treat as this value
            # (e.g. D7BCP -> D7BC)
            if text in overrides:
                text = _normalize_cell_value(overrides[text])
                if not text:
                    continue
            cell_ref = f"{get_column_letter(col_idx)}{row_idx}"

            # EMT: NL = Lawrence RW night (N9L) when EMT is operator on line.
            if role == "EMT" and text == "NL":
                text = "N9L"

            # Non-staffing / training / admin markers — persist as skipped.
            if text in skip_cell_values:
                if is_manager_row and text == "AOC":
                    records.append(
                        ShiftRecord(
                            date=d,
                            base="",
                            service_type="",
                            day_night="",
                            role=role,
                            filled=False,
                            overtime=False,
                            leave_type=None,
                            source_tab=sheet_label,
                            source_cell=cell_ref,
                            raw_value=text,
                            person_display=person_display,
                            person_displays=person_displays,
                            is_manager_row=True,
                            included_in_aggregates=False,
                            manager_event_type="aoc",
                            excel_row=row_idx,
                            excel_col=col_idx,
                        )
                    )
                    continue
                _append_skipped_shift(
                    records,
                    ws=ws,
                    row_idx=row_idx,
                    col_idx=col_idx,
                    d=d,
                    role=role,
                    sheet_label=sheet_label,
                    text=text,
                    person_displays=person_displays,
                    person_display=person_display,
                    is_manager_row=is_manager_row,
                    skip_reason=_classify_skip_reason(text, training_values),
                )
                continue

            if text == "OPEN":
                _append_skipped_shift(
                    records,
                    ws=ws,
                    row_idx=row_idx,
                    col_idx=col_idx,
                    d=d,
                    role=role,
                    sheet_label=sheet_label,
                    text=text,
                    person_displays=person_displays,
                    person_display=person_display,
                    is_manager_row=is_manager_row,
                    skip_reason="open",
                )
                continue

            # Weekday labels (e.g. row 2: SUN, MON, TUE, …) — ignore.
            if text in {"SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"}:
                continue

            # "<UNIT> EMT" (e.g. "MG EMT") on an RN/Medic row: the person is
            # covering that unit's EMT seat — a manager or medic filling in.
            # The seat itself is already recorded on the EMT sheet (and OPS
            # View), so counting this cross-reference as a staffed shift
            # would double-count the person. Persist as skipped admin. (On
            # the EMT sheet itself such a value would be the real seat
            # assignment, so this only applies to RN/Medic rows.)
            if (
                role != "EMT"
                and text.endswith(" EMT")
                and _is_resolvable_unit(text[:-4].strip())
            ):
                _append_skipped_shift(
                    records,
                    ws=ws,
                    row_idx=row_idx,
                    col_idx=col_idx,
                    d=d,
                    role=role,
                    sheet_label=sheet_label,
                    text=text,
                    person_displays=person_displays,
                    person_display=person_display,
                    is_manager_row=is_manager_row,
                    skip_reason="admin",
                )
                continue

            # Ignored unit-like codes — persist as skipped admin.
            if text in IGNORE_UNIT_CODES:
                _append_skipped_shift(
                    records,
                    ws=ws,
                    row_idx=row_idx,
                    col_idx=col_idx,
                    d=d,
                    role=role,
                    sheet_label=sheet_label,
                    text=text,
                    person_displays=person_displays,
                    person_display=person_display,
                    is_manager_row=is_manager_row,
                    skip_reason="admin",
                )
                continue

            # EMT: PER → LT (same bucket as leave time for exception grid).
            if role == "EMT" and text == "PER":
                if is_manager_row:
                    _append_skipped_shift(
                        records,
                        ws=ws,
                        row_idx=row_idx,
                        col_idx=col_idx,
                        d=d,
                        role=role,
                        sheet_label=sheet_label,
                        text=text,
                        person_displays=person_displays,
                        person_display=person_display,
                        is_manager_row=True,
                        skip_reason="manager_row",
                    )
                    continue
                records.append(
                    ShiftRecord(
                        date=d,
                        base="",
                        service_type="",
                        day_night="D",
                        role=role,
                        filled=False,
                        overtime=False,
                        leave_type="LT",
                        source_tab=sheet_label,
                        source_cell=cell_ref,
                        raw_value=text,
                        person_display=person_display,
                        person_displays=person_displays,
                        is_manager_row=is_manager_row,
                        excel_row=row_idx,
                        excel_col=col_idx,
                    )
                )
                continue

            # Merge: UNIT/LEAVETYPE → leave (any unit + any leave type)
            if "/" in text:
                parts = text.split("/", 1)
                suffix = (parts[1] or "").strip()
                leave_display = UNIT_LEAVE_MERGE.get(suffix)
                if leave_display is not None:
                    if is_manager_row:
                        _append_skipped_shift(
                            records,
                            ws=ws,
                            row_idx=row_idx,
                            col_idx=col_idx,
                            d=d,
                            role=role,
                            sheet_label=sheet_label,
                            text=text,
                            person_displays=person_displays,
                            person_display=person_display,
                            is_manager_row=True,
                            skip_reason="manager_row",
                        )
                        continue
                    records.append(
                        ShiftRecord(
                            date=d,
                            base="",
                            service_type="",
                            day_night="D",
                            role=role,
                            filled=False,
                            overtime=False,
                            leave_type=leave_display,
                            source_tab=sheet_label,
                            source_cell=cell_ref,
                            raw_value=text,
                            person_display=person_display,
                            person_displays=person_displays,
                            is_manager_row=is_manager_row,
                            excel_row=row_idx,
                            excel_col=col_idx,
                        )
                    )
                    continue

            # Leave/absence codes (LT-D, LT-N kept; SM/AT counts as AT.)
            if text in AT_ALIASES:
                leave_code = "AT"
                leave_display = "AT"
            elif text.startswith("LT-"):
                leave_code = "LT"
                leave_display = text  # LT-D or LT-N
            else:
                leave_code = text
                leave_display = text
            if leave_code in LEAVE_CODES:
                if is_manager_row:
                    _append_skipped_shift(
                        records,
                        ws=ws,
                        row_idx=row_idx,
                        col_idx=col_idx,
                        d=d,
                        role=role,
                        sheet_label=sheet_label,
                        text=text,
                        person_displays=person_displays,
                        person_display=person_display,
                        is_manager_row=True,
                        skip_reason="manager_row",
                    )
                    continue
                records.append(
                    ShiftRecord(
                        date=d,
                        base="",
                        service_type="",
                        day_night="D",
                        role=role,
                        filled=False,
                        overtime=False,
                        leave_type=leave_display,
                        source_tab=sheet_label,
                        source_cell=cell_ref,
                        raw_value=text,
                        person_display=person_display,
                        person_displays=person_displays,
                        is_manager_row=is_manager_row,
                        excel_row=row_idx,
                        excel_col=col_idx,
                    )
                )
                continue

            base_code, is_ot, is_dual = _split_unit_suffix(text)
            if base_code in RETIRED_UNIT_CODES:
                _append_skipped_shift(
                    records,
                    ws=ws,
                    row_idx=row_idx,
                    col_idx=col_idx,
                    d=d,
                    role=role,
                    sheet_label=sheet_label,
                    text=text,
                    person_displays=person_displays,
                    person_display=person_display,
                    is_manager_row=is_manager_row,
                    skip_reason="retired_unit",
                )
                continue

            unit_info = _classify_unit(base_code)
            if not unit_info:
                issues.append(
                    ParseIssue(
                        sheet=ws.title,
                        cell=cell_ref,
                        raw_value=text,
                        issue_type="unknown_unit",
                        message=(
                            f"Unknown unit code '{text}' (base '{base_code}') "
                            f"for role {role}."
                        ),
                    )
                )
                continue

            base, service_type, dn = unit_info
            canonical_unit = _canonical_unit_code(base_code) or base_code
            effective_role: Role = role
            if role == "RN" and is_dual:
                effective_role = "MEDIC"

            records.append(
                ShiftRecord(
                    date=d,
                    base=base,
                    service_type=service_type,
                    day_night=dn,
                    role=effective_role,
                    filled=True,
                    overtime=is_ot,
                    leave_type=None,
                    source_tab=sheet_label,
                    source_cell=cell_ref,
                    raw_value=text,
                    unit_code=canonical_unit,
                    person_display=person_display,
                    person_displays=person_displays,
                    is_manager_row=is_manager_row,
                    excel_row=row_idx,
                    excel_col=col_idx,
                )
            )

    return records, issues
