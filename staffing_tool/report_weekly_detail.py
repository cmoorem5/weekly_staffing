"""Weekly_Detail sheet writer (no cross-sheet formulas — values only)."""

from __future__ import annotations

from datetime import UTC, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .metrics import (
    TOTAL_PERSON_SHIFTS,
    WeekMetrics,
)
from .models import (
    BaseConfig,
    KpiThreshold,
    WeeklyStaffing,
)
from .rag import evaluate_rag
from .report_data import (
    BASE_UNIT_CELL_CONFIGURED,
    DETAIL_BASE_ORDER,
    EXCEPTION_ROLES,
    LEAVE_TYPE_COLS,
    _exc_count_breakdown,
    _leave_totals_from_breakdown,
    _rag_for_metric,
    _status_display,
)
from .report_excel_style import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    BMF_BLACK,
    BMF_NAVY,
    BOLD,
    FILL_BAND_ALT,
    FILL_BAND_TOTAL,
    FILL_BMF_GRAY_BG,
    FILL_BMF_NAVY,
    FILL_GREEN_FULL,
    FILL_RED_FULL,
    FILL_WHITE_SOLID,
    FILL_YELLOW,
    FONT_BMF_BODY,
    FONT_BMF_BODY_BOLD,
    FONT_BMF_RAG_VALUE,
    FONT_BMF_RAG_VALUE_ON_RED,
    FONT_DETAIL_TITLE,
    FONT_LEGEND_LABEL,
    FONT_NA,
    FONT_NAME,
    FONT_WEEK_BADGE,
    FONT_ZERO_MUTED,
    THIN_BORDER,
    _base_gr_cell_notable,
    _base_rw_cell_notable,
    _border_right_sep_cell,
    _detail_section_banner,
    _fill_and_font_for_status,
    _iso_week_label,
    _kpi_notable,
    _target_display,
    _write_na_or_int,
    _write_pct_cell,
)


def _verify_weekly_detail_checks(
    this_metrics: WeekMetrics, row_data: WeeklyStaffing
) -> list[tuple[str, bool, str]]:
    """Run verification rules for Weekly Staffing Detail. Returns list of (check_name, passed, message)."""
    checks = []
    # Required Total = Required Day + Required Night
    rt = this_metrics.required_day + this_metrics.required_night
    ok = rt == this_metrics.required_total
    checks.append(
        ("Required Total = Day + Night", ok, f"{rt} vs {this_metrics.required_total}")
    )
    # Filled Total = Filled Day + Filled Night
    ft = this_metrics.filled_day + this_metrics.filled_night
    ok = ft == this_metrics.filled_total
    checks.append(
        ("Filled Total = Day + Night", ok, f"{ft} vs {this_metrics.filled_total}")
    )
    # Vacancies = Required Total - Filled Total
    vac = max(0, this_metrics.required_total - this_metrics.filled_total)
    ok = vac == this_metrics.vacancies
    checks.append(
        ("Vacancies = Required - Filled", ok, f"{vac} vs {this_metrics.vacancies}")
    )
    # OT Total: mirror compute_week_metrics's fallback chain (day/night split,
    # then per-role legacy totals, then ot_shifts) so the check cross-checks
    # the same source the metrics actually use.
    total_ot = sum(
        int(getattr(row_data, f, 0) or 0)
        for f in (
            "ot_rn_day",
            "ot_rn_night",
            "ot_medic_day",
            "ot_medic_night",
            "ot_emt_day",
            "ot_emt_night",
        )
    )
    if total_ot == 0:
        total_ot = sum(
            int(getattr(row_data, f, 0) or 0) for f in ("ot_rn", "ot_medic", "ot_emt")
        )
    if total_ot == 0:
        total_ot = int(getattr(row_data, "ot_shifts", 0) or 0)
    ok = total_ot == this_metrics.ot_shifts
    checks.append(
        ("OT Total matches metrics", ok, f"{total_ot} vs {this_metrics.ot_shifts}")
    )
    return checks


def _write_weekly_detail(
    wb: Workbook,
    week_start: str,
    week_end: str,
    row_data: WeeklyStaffing,
    this_metrics: WeekMetrics,
    base_configs: list[BaseConfig],
    leave_breakdown: dict | None = None,
    thresholds: dict[str, KpiThreshold] | None = None,
) -> None:
    """Weekly_Detail: fixed cell map (columns A–G: label, Day, Night, Total, Target, Status, Notes)."""
    ws = wb.create_sheet("Weekly_Detail", 1)
    base_by_name = {b.base_name: b for b in base_configs}
    base_metrics = this_metrics.base_metrics or {}
    breakdown = leave_breakdown or {}
    grid_leave_total, col_totals_grid = (
        _leave_totals_from_breakdown(breakdown) if breakdown else (0, [0] * 6)
    )
    use_grid_leave = bool(breakdown)
    display_leave_total = (
        grid_leave_total if use_grid_leave else this_metrics.leave_total
    )
    display_leave_exp = (
        (grid_leave_total / TOTAL_PERSON_SHIFTS)
        if use_grid_leave
        else this_metrics.leave_exposure
    )
    thresholds = thresholds if thresholds is not None else {}
    thr_sr = thresholds.get("Staffing Rate")
    thr_exc = thresholds.get("Shift Exception %")
    thr_ot = thresholds.get("OT Dependency")
    t_rw = thresholds.get("System RW Coverage %")
    t_gr = thresholds.get("System GR Coverage %")

    medic_u = getattr(row_data, "medic_unpartnered", 0) or 0
    rn_u = getattr(row_data, "rn_unpartnered_staff", 0) or 0
    note_medic = (getattr(row_data, "unpartnered_note_medic", None) or "").strip()
    note_rn = (getattr(row_data, "unpartnered_note_rn", None) or "").strip()

    ot_rn_day = getattr(row_data, "ot_rn_day", 0) or 0
    ot_rn_night = getattr(row_data, "ot_rn_night", 0) or 0
    ot_medic_day = getattr(row_data, "ot_medic_day", 0) or 0
    ot_medic_night = getattr(row_data, "ot_medic_night", 0) or 0
    ot_emt_day = getattr(row_data, "ot_emt_day", 0) or 0
    ot_emt_night = getattr(row_data, "ot_emt_night", 0) or 0
    total_ot_day = ot_rn_day + ot_medic_day + ot_emt_day
    total_ot_night = ot_rn_night + ot_medic_night + ot_emt_night
    ot_dep = this_metrics.ot_dependency

    if breakdown:
        col_totals = col_totals_grid
    else:
        col_totals = [
            row_data.leave_at or 0,
            row_data.leave_lt or 0,
            row_data.leave_sick or 0,
            row_data.leave_loa or 0,
            getattr(row_data, "leave_jury", 0) or 0,
            getattr(row_data, "leave_brev", 0) or 0,
        ]

    def _exc_count(role: str, keys: list[str]) -> int:
        return _exc_count_breakdown(breakdown, role, keys)

    def _exec_striped(r: int) -> bool:
        return ((r - 7) % 2) == 1

    def _ot_striped(r: int) -> bool:
        return ((r - 18) % 2) == 1

    def _exc_striped(r: int) -> bool:
        return ((r - 26) % 2) == 1

    def _base_striped(r: int) -> bool:
        return ((r - 34) % 2) == 1

    def _row_bg(r: int, striped: bool) -> PatternFill:
        return FILL_BAND_ALT if striped else FILL_WHITE_SOLID

    def _paint_row(
        r: int, c1: int, c2: int, striped: bool, border: bool = True
    ) -> None:
        bg = _row_bg(r, striped)
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            cell.fill = bg
            if border:
                cell.border = THIN_BORDER

    def _zero_font(n: int) -> Font:
        return FONT_ZERO_MUTED if n == 0 else FONT_BMF_BODY

    # --- Row 1–2: title & period (pinned merges) ---
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=7)
    t1 = ws.cell(1, 1, "Boston MedFlight — Weekly Staffing Detail")
    t1.font = FONT_DETAIL_TITLE
    t1.fill = FILL_BMF_NAVY
    t1.alignment = ALIGN_LEFT
    t1.border = THIN_BORDER
    wk = ws.cell(1, 5, _iso_week_label(week_start))
    wk.font = FONT_WEEK_BADGE
    wk.fill = FILL_BMF_NAVY
    wk.alignment = ALIGN_RIGHT
    wk.border = THIN_BORDER

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=7)
    p2 = ws.cell(
        2,
        1,
        f"Reporting period: Sunday {week_start} — Saturday {week_end}",
    )
    p2.font = Font(name=FONT_NAME, size=10, color=BMF_NAVY)
    p2.fill = FILL_BMF_GRAY_BG
    p2.alignment = ALIGN_LEFT
    p2.border = THIN_BORDER
    g2 = ws.cell(
        2,
        5,
        f"Generated: {datetime.now(UTC).strftime('%Y-%m-%d')}",
    )
    g2.font = Font(name=FONT_NAME, size=10, color=BMF_NAVY)
    g2.fill = FILL_BMF_GRAY_BG
    g2.alignment = ALIGN_RIGHT
    g2.border = THIN_BORDER

    # --- Row 3: legend ---
    a3 = ws.cell(3, 1, "Status legend:")
    a3.font = FONT_LEGEND_LABEL
    a3.fill = FILL_BMF_GRAY_BG
    a3.alignment = ALIGN_RIGHT
    a3.border = THIN_BORDER
    b3 = ws.cell(3, 2, "On target")
    b3.font = Font(name=FONT_NAME, size=10, bold=True, color=BMF_BLACK)
    b3.fill = FILL_GREEN_FULL
    b3.alignment = ALIGN_CENTER
    b3.border = THIN_BORDER
    c3 = ws.cell(3, 3, "Monitor")
    c3.font = Font(name=FONT_NAME, size=10, bold=True, color=BMF_BLACK)
    c3.fill = FILL_YELLOW
    c3.alignment = ALIGN_CENTER
    c3.border = THIN_BORDER
    d3 = ws.cell(3, 4, "Action needed")
    d3.font = FONT_BMF_RAG_VALUE_ON_RED
    d3.fill = FILL_RED_FULL
    d3.alignment = ALIGN_CENTER
    d3.border = THIN_BORDER
    e3 = ws.cell(3, 5, "N/A = unit not staffed at that base/shift")
    e3.font = Font(name=FONT_NAME, size=10, italic=True, color="666666")
    e3.fill = FILL_BMF_GRAY_BG
    e3.alignment = ALIGN_LEFT
    e3.border = THIN_BORDER
    ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=7)
    fg3 = ws.cell(3, 6)
    fg3.fill = FILL_BMF_GRAY_BG
    fg3.border = THIN_BORDER

    # --- Row 4: spacer ---
    for c in range(1, 8):
        ws.cell(4, c).fill = FILL_WHITE_SOLID
        ws.cell(4, c).border = THIN_BORDER

    # --- Row 5: Executive summary banner ---
    _detail_section_banner(ws, 5, "Executive summary", "This week vs target")

    # --- Row 6: column headers ---
    hdr_fill = FILL_BMF_GRAY_BG
    h6 = [
        (1, "Metric", ALIGN_LEFT),
        (2, "Day", ALIGN_CENTER),
        (3, "Night", ALIGN_CENTER),
        (4, "Total", ALIGN_CENTER),
        (5, "Target", ALIGN_CENTER),
        (6, "Status", ALIGN_CENTER),
        (7, "Notes", ALIGN_LEFT),
    ]
    for col, text, al in h6:
        cell = ws.cell(6, col, text)
        cell.font = BOLD
        cell.fill = hdr_fill
        cell.border = THIN_BORDER
        cell.alignment = al

    # --- Executive data rows 7–14 ---
    sr_day = (
        this_metrics.filled_day / this_metrics.required_day
        if this_metrics.required_day
        else 0.0
    )
    sr_night = (
        this_metrics.filled_night / this_metrics.required_night
        if this_metrics.required_night
        else 0.0
    )
    sr_rag = _rag_for_metric("Staffing Rate", this_metrics.staffing_rate, thresholds)
    sr_notable = _kpi_notable("Staffing Rate", this_metrics.staffing_rate, this_metrics)
    exc_rag = _rag_for_metric("Shift Exception %", display_leave_exp, thresholds)
    exc_notable = _kpi_notable("Shift Exception %", display_leave_exp, this_metrics)
    ot_rag = _rag_for_metric("OT Dependency", ot_dep, thresholds)
    ot_notable = _kpi_notable("OT Dependency", ot_dep, this_metrics)

    ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=3)

    for r in range(7, 15):
        st = _exec_striped(r)
        _paint_row(r, 1, 7, st)

    # Row 7 Required shifts
    ws.cell(7, 1, "Required shifts").font = FONT_BMF_BODY
    ws.cell(7, 1).alignment = ALIGN_LEFT
    ws.cell(7, 2, this_metrics.required_day).alignment = ALIGN_CENTER
    ws.cell(7, 3, this_metrics.required_night).alignment = ALIGN_CENTER
    d7 = ws.cell(7, 4, this_metrics.required_total)
    d7.font = FONT_BMF_BODY_BOLD
    d7.alignment = ALIGN_CENTER
    for c in (5, 6, 7):
        ws.cell(7, c, None)

    # Row 8 Filled shifts
    ws.cell(8, 1, "Filled shifts").font = FONT_BMF_BODY
    ws.cell(8, 1).alignment = ALIGN_LEFT
    ws.cell(8, 2, this_metrics.filled_day).alignment = ALIGN_CENTER
    ws.cell(8, 3, this_metrics.filled_night).alignment = ALIGN_CENTER
    d8 = ws.cell(8, 4, this_metrics.filled_total)
    d8.font = FONT_BMF_BODY_BOLD
    d8.alignment = ALIGN_CENTER
    for c in (5, 6, 7):
        ws.cell(8, c, None)

    # Row 9 Staffing rate
    ws.cell(9, 1, "Staffing rate").font = FONT_BMF_BODY_BOLD
    ws.cell(9, 1).alignment = ALIGN_LEFT
    b9 = ws.cell(9, 2, sr_day)
    b9.number_format = "0.0%"
    b9.alignment = ALIGN_CENTER
    c9 = ws.cell(9, 3, sr_night)
    c9.number_format = "0.0%"
    c9.alignment = ALIGN_CENTER
    d9 = ws.cell(9, 4, this_metrics.staffing_rate)
    d9.number_format = "0.0%"
    d9.font = FONT_BMF_RAG_VALUE
    d9.alignment = ALIGN_CENTER
    dfill, dfont = _fill_and_font_for_status(
        sr_rag, notable=sr_notable, value=this_metrics.staffing_rate, thr=thr_sr
    )
    d9.fill = dfill
    d9.font = dfont
    e9 = ws.cell(9, 5, _target_display("Staffing Rate", thr_sr))
    e9.font = Font(name=FONT_NAME, size=11, color="666666")
    e9.number_format = "@"
    e9.alignment = ALIGN_CENTER
    f9 = ws.cell(9, 6, _status_display(sr_rag))
    f9.font = FONT_BMF_RAG_VALUE
    f9.alignment = ALIGN_CENTER
    ff9, ffont9 = _fill_and_font_for_status(
        sr_rag, notable=sr_notable, value=this_metrics.staffing_rate, thr=thr_sr
    )
    f9.fill = ff9
    f9.font = ffont9
    ws.cell(9, 7, None)

    # Row 10 Vacancies
    ws.cell(10, 1, "Vacancies").font = FONT_BMF_BODY
    ws.cell(10, 1).alignment = ALIGN_LEFT
    ws.cell(10, 2, None)
    ws.cell(10, 3, None)
    d10 = ws.cell(10, 4, this_metrics.vacancies)
    d10.font = FONT_BMF_BODY_BOLD
    d10.alignment = ALIGN_CENTER
    for c in (5, 6, 7):
        ws.cell(10, c, None)

    # Row 11 Shift exceptions (total)
    ws.cell(11, 1, "Shift exceptions (total)").font = FONT_BMF_BODY
    ws.cell(11, 1).alignment = ALIGN_LEFT
    ws.cell(11, 2, None)
    ws.cell(11, 3, None)
    d11 = ws.cell(11, 4, display_leave_total)
    d11.font = FONT_BMF_BODY_BOLD
    d11.alignment = ALIGN_CENTER
    for c in (5, 6, 7):
        ws.cell(11, c, None)

    # Row 12 Shift exception %
    ws.cell(12, 1, "Shift exception %").font = FONT_BMF_BODY_BOLD
    ws.cell(12, 1).alignment = ALIGN_LEFT
    ws.cell(12, 2, None)
    ws.cell(12, 3, None)
    d12 = ws.cell(12, 4, display_leave_exp)
    d12.number_format = "0.0%"
    d12.font = FONT_BMF_RAG_VALUE
    d12.alignment = ALIGN_CENTER
    df12, dt12 = _fill_and_font_for_status(
        exc_rag,
        notable=exc_notable,
        value=display_leave_exp,
        thr=thr_exc,
    )
    d12.fill = df12
    d12.font = dt12
    e12 = ws.cell(12, 5, _target_display("Shift Exception %", thr_exc))
    e12.font = Font(name=FONT_NAME, size=11, color="666666")
    e12.number_format = "@"
    e12.alignment = ALIGN_CENTER
    f12 = ws.cell(12, 6, _status_display(exc_rag))
    ff12, ft12 = _fill_and_font_for_status(
        exc_rag,
        notable=exc_notable,
        value=display_leave_exp,
        thr=thr_exc,
    )
    f12.fill = ff12
    f12.font = ft12
    f12.alignment = ALIGN_CENTER
    ws.cell(12, 7, None)

    # Row 13 Unpartnered Medic (A:C merged above)
    ws.cell(13, 1, "Unpartnered — Medic").font = FONT_BMF_BODY_BOLD
    ws.cell(13, 1).alignment = ALIGN_LEFT
    d13 = ws.cell(13, 4, medic_u)
    d13.font = FONT_BMF_BODY_BOLD
    d13.alignment = ALIGN_CENTER
    for c in (5, 6):
        ws.cell(13, c, None)
    g13 = ws.cell(13, 7, note_medic)
    g13.font = Font(name=FONT_NAME, size=11, color="666666")
    g13.alignment = ALIGN_LEFT

    # Row 14 Unpartnered RN
    ws.cell(14, 1, "Unpartnered — RN").font = FONT_BMF_BODY_BOLD
    ws.cell(14, 1).alignment = ALIGN_LEFT
    ws.cell(14, 2, None)
    ws.cell(14, 3, None)
    ws.cell(14, 4, rn_u).alignment = ALIGN_CENTER
    for c in (5, 6):
        ws.cell(14, c, None)
    g14 = ws.cell(14, 7, note_rn)
    g14.font = Font(name=FONT_NAME, size=11, italic=True, color="666666")
    g14.alignment = ALIGN_LEFT

    # Row 15 spacer
    for c in range(1, 8):
        ws.cell(15, c).fill = FILL_WHITE_SOLID
        ws.cell(15, c).border = THIN_BORDER

    # --- Section 2 Overtime rows 16–23 ---
    _detail_section_banner(ws, 16, "Overtime", "Shift counts")
    h_ot = [
        (1, "Role", ALIGN_LEFT),
        (2, "Day", ALIGN_CENTER),
        (3, "Night", ALIGN_CENTER),
        (4, "Total", ALIGN_CENTER),
        (5, "Target", ALIGN_CENTER),
        (6, "Status", ALIGN_CENTER),
        (7, "Notes", ALIGN_LEFT),
    ]
    for col, text, al in h_ot:
        cell = ws.cell(17, col, text)
        cell.font = BOLD
        cell.fill = hdr_fill
        cell.border = THIN_BORDER
        cell.alignment = al

    ot_rows = [
        ("RN", ot_rn_day, ot_rn_night),
        ("Medic", ot_medic_day, ot_medic_night),
        ("EMT", ot_emt_day, ot_emt_night),
    ]
    r = 18
    for label, bd, bn in ot_rows:
        st = _ot_striped(r)
        _paint_row(r, 1, 7, st)
        ws.cell(r, 1, label).font = FONT_BMF_BODY_BOLD
        ws.cell(r, 1).alignment = ALIGN_LEFT
        b = ws.cell(r, 2, bd)
        b.font = _zero_font(bd)
        b.alignment = ALIGN_CENTER
        c = ws.cell(r, 3, bn)
        c.font = _zero_font(bn)
        c.alignment = ALIGN_CENTER
        ws.cell(r, 4, bd + bn).alignment = ALIGN_CENTER
        for cc in (5, 6, 7):
            ws.cell(r, cc, None)
        r += 1

    # Row 21 Total OT
    for c in range(1, 8):
        cell = ws.cell(21, c)
        cell.fill = FILL_BAND_TOTAL
        cell.border = THIN_BORDER
        cell.font = FONT_BMF_BODY_BOLD
    ws.cell(21, 1, "Total").alignment = ALIGN_LEFT
    ws.cell(21, 2, total_ot_day).alignment = ALIGN_CENTER
    ws.cell(21, 3, total_ot_night).alignment = ALIGN_CENTER
    ws.cell(21, 4, total_ot_day + total_ot_night).alignment = ALIGN_CENTER
    for c in (5, 6, 7):
        ws.cell(21, c, None)

    # Row 22 Backfill rate
    _paint_row(22, 1, 7, False)
    ws.cell(22, 1, "Backfill rate (OT / filled)").font = FONT_BMF_BODY_BOLD
    ws.cell(22, 1).alignment = ALIGN_LEFT
    ws.cell(22, 2, None)
    ws.cell(22, 3, None)
    d22 = ws.cell(22, 4, ot_dep)
    d22.number_format = "0.0%"
    d22.font = FONT_BMF_RAG_VALUE
    d22.alignment = ALIGN_CENTER
    df22, dt22 = _fill_and_font_for_status(
        ot_rag, notable=ot_notable, value=ot_dep, thr=thr_ot
    )
    d22.fill = df22
    d22.font = dt22
    e22 = ws.cell(22, 5, _target_display("OT Dependency", thr_ot))
    e22.font = Font(name=FONT_NAME, size=11, color="666666")
    e22.number_format = "@"
    e22.alignment = ALIGN_CENTER
    f22 = ws.cell(22, 6, _status_display(ot_rag))
    ff22, ft22 = _fill_and_font_for_status(
        ot_rag, notable=ot_notable, value=ot_dep, thr=thr_ot
    )
    f22.fill = ff22
    f22.font = ft22
    f22.alignment = ALIGN_CENTER
    ws.cell(22, 7, None)

    # Row 23 spacer
    for c in range(1, 8):
        ws.cell(23, c).fill = FILL_WHITE_SOLID
        ws.cell(23, c).border = THIN_BORDER

    # --- Section 3 Exceptions rows 24–31 ---
    _detail_section_banner(
        ws, 24, "Schedule exceptions by role", "Shift counts by type"
    )
    ws.cell(25, 1, "Role").font = BOLD
    ws.cell(25, 1).fill = hdr_fill
    ws.cell(25, 1).border = THIN_BORDER
    ws.cell(25, 1).alignment = ALIGN_LEFT
    for ci, lt in enumerate(LEAVE_TYPE_COLS, start=2):
        cell = ws.cell(25, ci, lt)
        cell.font = BOLD
        cell.fill = hdr_fill
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER

    role_labels = EXCEPTION_ROLES
    rr = 26
    for role in role_labels:
        st = _exc_striped(rr)
        _paint_row(rr, 1, 7, st)
        ws.cell(rr, 1, role).font = FONT_BMF_BODY_BOLD
        ws.cell(rr, 1).alignment = ALIGN_LEFT
        vals = [
            _exc_count(role, ["AT"]),
            _exc_count(role, ["LT-D", "LT-N", "LT"]),
            _exc_count(role, ["SICK"]),
            _exc_count(role, ["LOA", "PFML"]),
            _exc_count(role, ["JURY"]),
            _exc_count(role, ["BREV"]),
        ]
        for j, v in enumerate(vals, start=2):
            cell = ws.cell(rr, j, v)
            cell.alignment = ALIGN_CENTER
            cell.font = _zero_font(v)
        rr += 1

    for c in range(1, 8):
        cell = ws.cell(30, c)
        cell.fill = FILL_BAND_TOTAL
        cell.border = THIN_BORDER
        cell.font = FONT_BMF_BODY_BOLD
    ws.cell(30, 1, "Total").alignment = ALIGN_LEFT
    for j, tot in enumerate(col_totals, start=2):
        cell = ws.cell(30, j, tot)
        cell.alignment = ALIGN_CENTER
        if j in (6, 7) and tot == 0:
            cell.font = FONT_ZERO_MUTED
        else:
            cell.font = FONT_BMF_BODY_BOLD

    # Row 31 spacer
    for c in range(1, 8):
        ws.cell(31, c).fill = FILL_WHITE_SOLID
        ws.cell(31, c).border = THIN_BORDER

    # --- Section 4 Base coverage ---
    _detail_section_banner(ws, 32, "Base coverage", "Rotor-Wing / Ground")
    base_headers = [
        "Base",
        "RW/D (of 7)",
        "RW/N (of 7)",
        "GR/D (of 7)",
        "GR/N (of 7)",
        "RW %",
        "GR %",
    ]
    for c, h in enumerate(base_headers, start=1):
        cell = ws.cell(33, c, h)
        cell.font = BOLD
        cell.fill = hdr_fill
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER if c > 1 else ALIGN_LEFT
    _border_right_sep_cell(ws, 33, 3)
    _border_right_sep_cell(ws, 33, 5)

    row_num = 34
    for base_name in DETAIL_BASE_ORDER:
        cfg = base_by_name.get(base_name)
        pct = base_metrics.get(base_name, {})
        rw_staffed = pct.get("rw_staffed", 0)
        gr_staffed = pct.get("gr_staffed", 0)
        rw_total = cfg.rw_total_unit_days if cfg else 0
        gr_total = cfg.gr_total_unit_days if cfg else 0
        rw_d = int(pct.get("rw_d", rw_staffed))
        rw_n = int(pct.get("rw_n", 0))
        gr_d = int(pct.get("gr_d", gr_staffed))
        gr_n = int(pct.get("gr_n", 0))
        umap = BASE_UNIT_CELL_CONFIGURED.get(base_name, {})
        st = _base_striped(row_num)
        _paint_row(row_num, 1, 7, st)
        ws.cell(row_num, 1, base_name).font = FONT_BMF_BODY_BOLD
        ws.cell(row_num, 1).alignment = ALIGN_LEFT
        _write_na_or_int(ws, row_num, 2, umap.get("rw_d", False), rw_d)
        _write_na_or_int(ws, row_num, 3, umap.get("rw_n", False), rw_n)
        _write_na_or_int(ws, row_num, 4, umap.get("gr_d", False), gr_d)
        _write_na_or_int(ws, row_num, 5, umap.get("gr_n", False), gr_n)
        if rw_total:
            rw_pct_val = rw_staffed / rw_total
            rw_rag = evaluate_rag(rw_pct_val, t_rw) if t_rw else "Green"
            c6 = ws.cell(row_num, 6, rw_pct_val)
            c6.number_format = "0.0%"
            c6.alignment = ALIGN_CENTER
            c6.border = THIN_BORDER
            fill6, font6 = _fill_and_font_for_status(
                rw_rag,
                notable=_base_rw_cell_notable(rw_pct_val),
                value=rw_pct_val,
                thr=t_rw,
            )
            c6.fill = fill6
            c6.font = font6
        else:
            c6 = ws.cell(row_num, 6, "N/A")
            c6.font = FONT_NA
            c6.alignment = ALIGN_CENTER
            c6.border = THIN_BORDER
        if gr_total:
            gr_pct_val = gr_staffed / gr_total
            gr_rag = evaluate_rag(gr_pct_val, t_gr) if t_gr else "Green"
            c7 = ws.cell(row_num, 7, gr_pct_val)
            c7.number_format = "0.0%"
            c7.alignment = ALIGN_CENTER
            c7.border = THIN_BORDER
            fill7, font7 = _fill_and_font_for_status(
                gr_rag,
                notable=_base_gr_cell_notable(gr_pct_val),
                value=gr_pct_val,
                thr=t_gr,
            )
            c7.fill = fill7
            c7.font = font7
        else:
            c7 = ws.cell(row_num, 7, "N/A")
            c7.font = FONT_NA
            c7.alignment = ALIGN_CENTER
            c7.border = THIN_BORDER
        _border_right_sep_cell(ws, row_num, 3)
        _border_right_sep_cell(ws, row_num, 5)
        row_num += 1

    for c in range(1, 8):
        cell = ws.cell(row_num, c)
        cell.fill = FILL_BAND_TOTAL
        cell.border = THIN_BORDER
        cell.font = FONT_BMF_BODY_BOLD
    ws.cell(row_num, 1, "System total").alignment = ALIGN_LEFT
    # §3.3: do not sum per-base raw counts (B–E); system view is weighted % only.
    for c in range(2, 6):
        cell = ws.cell(row_num, c)
        cell.value = None
        cell.alignment = ALIGN_CENTER
    sys_rw_rag = _rag_for_metric(
        "System RW Coverage %", this_metrics.system_rw_pct, thresholds
    )
    sys_rw_notable = _kpi_notable(
        "System RW Coverage %", this_metrics.system_rw_pct, this_metrics
    )
    _write_pct_cell(
        ws,
        row_num,
        6,
        this_metrics.system_rw_pct,
        sys_rw_rag,
        thr=t_rw,
        notable=sys_rw_notable,
    )
    sys_gr_rag = _rag_for_metric(
        "System GR Coverage %", this_metrics.system_gr_pct, thresholds
    )
    sys_gr_notable = _kpi_notable(
        "System GR Coverage %", this_metrics.system_gr_pct, this_metrics
    )
    _write_pct_cell(
        ws,
        row_num,
        7,
        this_metrics.system_gr_pct,
        sys_gr_rag,
        thr=t_gr,
        notable=sys_gr_notable,
    )
    _border_right_sep_cell(ws, row_num, 3)
    _border_right_sep_cell(ws, row_num, 5)

    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 28
    for letter, w in [("B", 12), ("C", 12), ("D", 12), ("E", 12), ("F", 14), ("G", 28)]:
        ws.column_dimensions[letter].width = w
