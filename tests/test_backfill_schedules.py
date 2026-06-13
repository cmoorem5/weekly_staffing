"""Tests for bulk schedule backfill script."""

import importlib.util
import shutil
import sys
import tempfile
import unittest
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from staffing_tool.db import (
    _get_engine_cached,
    _sessionmaker_for_path,
    get_engine,
    init_db,
    session_scope,
)
from staffing_tool.models import (
    ScheduleImport,
    WeeklyManagerShift,
    WeeklyPersonShift,
    WeeklyStaffing,
)
from staffing_tool.schedule_apply import (
    apply_schedule_workbook,
    week_already_imported,
    week_has_current_import,
)
from staffing_tool.schedule_import import PARSER_VERSION, _parse_grid
from staffing_tool.unit_mappings import (
    load_unit_mappings_from_csv,
    resolve_unit_overrides,
    save_unit_mappings,
)

_REPO_ROOT = Path(__file__).resolve().parents[1]
_SAMPLE = _REPO_ROOT / "uploads/schedule_upload_20260611T025001Z.xlsx"

_spec = importlib.util.spec_from_file_location(
    "backfill_schedules",
    _REPO_ROOT / "scripts/backfill_schedules.py",
)
assert _spec and _spec.loader
_backfill = importlib.util.module_from_spec(_spec)
sys.modules["backfill_schedules"] = _backfill
_spec.loader.exec_module(_backfill)


class BackfillPlanningTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.db_path = str(Path(self.tmp.name) / "test.db")

    def tearDown(self):
        import staffing_tool.db as db_mod

        resolved = db_mod._resolve_db_path(self.db_path)
        get_engine(self.db_path).dispose()
        _get_engine_cached.cache_clear()
        _sessionmaker_for_path.cache_clear()
        db_mod._DB_READY_PATHS.discard(resolved)
        self.tmp.cleanup()

    def test_week_start_from_filename(self):
        self.assertEqual(
            _backfill._week_start_from_filename(Path("BMF_Weekly_2026-05-31.xlsx")),
            "2026-05-31",
        )
        self.assertIsNone(
            _backfill._week_start_from_filename(Path("schedule.xlsx")),
        )

    def test_plan_skips_any_existing_import(self):
        if not _SAMPLE.is_file():
            self.skipTest("sample upload not present")
        init_db(self.db_path)
        inbox = Path(self.tmp.name) / "inbox"
        inbox.mkdir()
        shutil.copy(_SAMPLE, inbox / _SAMPLE.name)

        planned, errors = _backfill.plan_imports(
            [inbox / _SAMPLE.name],
            db_path=self.db_path,
            force=False,
            force_week="2026-05-31",
        )
        self.assertEqual(errors, [])
        self.assertEqual(len(planned), 1)
        self.assertIsNone(planned[0].skip_reason)

        with session_scope(self.db_path) as session:
            result, err = apply_schedule_workbook(
                session,
                week_start="2026-05-31",
                upload_path=str(inbox / _SAMPLE.name),
                source_filename=_SAMPLE.name,
            )
        self.assertIsNone(err)
        assert result is not None

        replanned, _ = _backfill.plan_imports(
            [inbox / _SAMPLE.name],
            db_path=self.db_path,
            force=False,
            force_week="2026-05-31",
        )
        self.assertEqual(len(replanned), 1)
        self.assertEqual(replanned[0].skip_reason, "already imported")

        forced, _ = _backfill.plan_imports(
            [inbox / _SAMPLE.name],
            db_path=self.db_path,
            force=True,
            force_week="2026-05-31",
        )
        self.assertIsNone(forced[0].skip_reason)

    def test_plan_respects_from_date_filter(self):
        if not _SAMPLE.is_file():
            self.skipTest("sample upload not present")
        init_db(self.db_path)
        inbox = Path(self.tmp.name) / "inbox"
        inbox.mkdir()
        shutil.copy(_SAMPLE, inbox / _SAMPLE.name)

        planned, _ = _backfill.plan_imports(
            [inbox / _SAMPLE.name],
            db_path=self.db_path,
            force=False,
            force_week="2026-05-31",
            from_date="2026-06-01",
        )
        self.assertEqual(planned, [])

        planned_ok, _ = _backfill.plan_imports(
            [inbox / _SAMPLE.name],
            db_path=self.db_path,
            force=False,
            force_week="2026-05-31",
            from_date="2026-05-01",
        )
        self.assertEqual(len(planned_ok), 1)


class BackfillRunTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.db_path = str(Path(self.tmp.name) / "test.db")
        init_db(self.db_path)
        self.inbox = Path(self.tmp.name) / "inbox"
        self.inbox.mkdir()

    def tearDown(self):
        import staffing_tool.db as db_mod

        resolved = db_mod._resolve_db_path(self.db_path)
        get_engine(self.db_path).dispose()
        _get_engine_cached.cache_clear()
        _sessionmaker_for_path.cache_clear()
        db_mod._DB_READY_PATHS.discard(resolved)
        self.tmp.cleanup()

    def test_dry_run_lists_week_without_db_writes(self):
        if not _SAMPLE.is_file():
            self.skipTest("sample upload not present")
        shutil.copy(_SAMPLE, self.inbox / _SAMPLE.name)
        code = _backfill.run_backfill(
            [self.inbox],
            db_path=self.db_path,
            dry_run=True,
            force=False,
            force_week=None,
        )
        self.assertEqual(code, 0)
        with session_scope(self.db_path) as session:
            self.assertEqual(session.query(ScheduleImport).count(), 0)

    def test_run_imports_person_shift_detail(self):
        if not _SAMPLE.is_file():
            self.skipTest("sample upload not present")
        shutil.copy(_SAMPLE, self.inbox / _SAMPLE.name)
        code = _backfill.run_backfill(
            [self.inbox],
            db_path=self.db_path,
            dry_run=False,
            force=False,
            force_week=None,
        )
        self.assertEqual(code, 0)
        with session_scope(self.db_path) as session:
            imp = (
                session.query(ScheduleImport)
                .filter(ScheduleImport.week_start == "2026-05-31")
                .first()
            )
            self.assertIsNotNone(imp)
            assert imp is not None
            self.assertEqual(imp.parser_version, PARSER_VERSION)
            self.assertGreater(imp.person_event_count, 0)
            self.assertTrue(week_has_current_import(session, "2026-05-31"))
            self.assertTrue(week_already_imported(session, "2026-05-31"))
            count = (
                session.query(WeeklyPersonShift)
                .filter(WeeklyPersonShift.week_start == "2026-05-31")
                .count()
            )
            self.assertEqual(count, imp.person_event_count)


class UnitMappingBackfillTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.db_path = str(Path(self.tmp.name) / "test.db")
        init_db(self.db_path)

    def tearDown(self):
        import staffing_tool.db as db_mod

        resolved = db_mod._resolve_db_path(self.db_path)
        get_engine(self.db_path).dispose()
        _get_engine_cached.cache_clear()
        _sessionmaker_for_path.cache_clear()
        db_mod._DB_READY_PATHS.discard(resolved)
        self.tmp.cleanup()

    def _parse_unknown_with_overrides(
        self, overrides: dict[str, str]
    ) -> tuple[list, list]:
        wb = Workbook()
        ws = wb.active
        ws.title = "RN & Medic"
        ws["C1"] = date(2024, 1, 7)
        ws["A4"] = "Smith"
        ws["B4"] = "Jane"
        ws["C4"] = "MYSTERY"
        return _parse_grid(
            ws=ws,
            header_row_idx=1,
            first_row_idx=4,
            last_row_idx=4,
            role="RN",
            sheet_label="RN & Medic (RN)",
            week_start_date=date(2024, 1, 7),
            week_end_date=date(2024, 1, 13),
            unit_overrides=overrides,
        )

    def test_saved_mapping_resolves_unknown_code(self):
        with session_scope(self.db_path) as session:
            save_unit_mappings(session, {"MYSTERY": "D7B"}, source="dashboard")
            overrides = resolve_unit_overrides(session)
        self.assertIn("MYSTERY", overrides)
        records, issues = self._parse_unknown_with_overrides(overrides)
        self.assertEqual(issues, [])
        staffed = [r for r in records if r.filled]
        self.assertEqual(len(staffed), 1)
        self.assertEqual(staffed[0].base, "Bedford")

    def test_csv_unit_map_merged_with_db(self):
        csv_path = Path(self.tmp.name) / "units.csv"
        csv_path.write_text("raw,maps_to\nMYSTERY,D7B\n", encoding="utf-8")
        loaded = load_unit_mappings_from_csv(csv_path)
        self.assertEqual(loaded["MYSTERY"], "D7B")
        records, issues = self._parse_unknown_with_overrides(loaded)
        self.assertEqual(issues, [])
        self.assertEqual(len([r for r in records if r.filled]), 1)

    def test_force_replaces_existing_import(self):
        if not _SAMPLE.is_file():
            self.skipTest("sample upload not present")
        inbox = Path(self.tmp.name) / "inbox"
        inbox.mkdir()
        shutil.copy(_SAMPLE, inbox / _SAMPLE.name)
        code = _backfill.run_backfill(
            [inbox],
            db_path=self.db_path,
            dry_run=False,
            force=False,
            force_week="2026-05-31",
        )
        self.assertEqual(code, 0)
        with session_scope(self.db_path) as session:
            first = (
                session.query(ScheduleImport)
                .filter(ScheduleImport.week_start == "2026-05-31")
                .one()
            )
            first_imported_at = first.imported_at
            first_count = first.person_event_count

        planned_skip, _ = _backfill.plan_imports(
            [inbox / _SAMPLE.name],
            db_path=self.db_path,
            force=False,
            force_week="2026-05-31",
        )
        self.assertEqual(planned_skip[0].skip_reason, "already imported")

        code = _backfill.run_backfill(
            [inbox],
            db_path=self.db_path,
            dry_run=False,
            force=True,
            force_week="2026-05-31",
        )
        self.assertEqual(code, 0)
        with session_scope(self.db_path) as session:
            second = (
                session.query(ScheduleImport)
                .filter(ScheduleImport.week_start == "2026-05-31")
                .one()
            )
        self.assertNotEqual(second.imported_at, first_imported_at)
        self.assertEqual(second.person_event_count, first_count)


def _seed_legacy_weekly_staffing(
    db_path: str,
    week_start: str,
    *,
    notes: str = "CEO reviewed — keep this",
) -> None:
    with session_scope(db_path) as session:
        session.add(
            WeeklyStaffing(
                week_start=week_start,
                day_target=9,
                night_min=5,
                filled_day=40,
                filled_night=20,
                overnights_below=2,
                pilot_vacancies=1,
                medic_unpartnered=3,
                rn_unpartnered_staff=1,
                unpartnered_note_medic="coverage gap",
                unpartnered_note_rn="training",
                notes=notes,
                entered_by="manual",
                created_at="2025-10-01T00:00:00Z",
            )
        )


def _write_aoc_workbook(path: Path, week_sunday: date) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RN & Medic"
    ws["C1"] = week_sunday
    ws["D1"] = week_sunday + timedelta(days=1)
    ws["A4"] = "Bowman"
    ws["B4"] = "m"
    ws["D4"] = "AOC"
    ws["A5"] = "Smith"
    ws["B5"] = "Jane"
    ws["C5"] = "D7B"
    ws["A52"] = "Jones"
    ws["B52"] = "Bob"
    ws["C52"] = "D7B"
    wb.save(path)
    wb.close()


class UpgradeDetailTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.db_path = str(Path(self.tmp.name) / "test.db")
        init_db(self.db_path)
        self.inbox = Path(self.tmp.name) / "inbox"
        self.inbox.mkdir()

    def tearDown(self):
        import staffing_tool.db as db_mod

        resolved = db_mod._resolve_db_path(self.db_path)
        get_engine(self.db_path).dispose()
        _get_engine_cached.cache_clear()
        _sessionmaker_for_path.cache_clear()
        db_mod._DB_READY_PATHS.discard(resolved)
        self.tmp.cleanup()

    def test_upgrade_preserves_manual_weekly_fields(self):
        if not _SAMPLE.is_file():
            self.skipTest("sample upload not present")
        week = "2026-05-31"
        _seed_legacy_weekly_staffing(self.db_path, week)
        shutil.copy(_SAMPLE, self.inbox / _SAMPLE.name)

        code = _backfill.run_backfill(
            [self.inbox],
            db_path=self.db_path,
            dry_run=False,
            force=False,
            force_week=None,
            upgrade_detail=True,
        )
        self.assertEqual(code, 0)

        with session_scope(self.db_path) as session:
            row = (
                session.query(WeeklyStaffing)
                .filter(WeeklyStaffing.week_start == week)
                .one()
            )
            self.assertEqual(row.notes, "CEO reviewed — keep this")
            self.assertEqual(row.day_target, 9)
            self.assertEqual(row.night_min, 5)
            self.assertEqual(row.overnights_below, 2)
            self.assertEqual(row.pilot_vacancies, 1)
            self.assertEqual(row.medic_unpartnered, 3)
            self.assertEqual(row.rn_unpartnered_staff, 1)
            self.assertEqual(row.unpartnered_note_medic, "coverage gap")
            self.assertEqual(row.unpartnered_note_rn, "training")
            self.assertEqual(row.entered_by, "manual")

    def test_upgrade_populates_person_shifts(self):
        if not _SAMPLE.is_file():
            self.skipTest("sample upload not present")
        week = "2026-05-31"
        _seed_legacy_weekly_staffing(self.db_path, week)
        shutil.copy(_SAMPLE, self.inbox / _SAMPLE.name)

        with session_scope(self.db_path) as session:
            self.assertEqual(
                session.query(WeeklyPersonShift)
                .filter(WeeklyPersonShift.week_start == week)
                .count(),
                0,
            )

        code = _backfill.run_backfill(
            [self.inbox],
            db_path=self.db_path,
            dry_run=False,
            force=False,
            force_week=None,
            upgrade_detail=True,
        )
        self.assertEqual(code, 0)

        with session_scope(self.db_path) as session:
            imp = (
                session.query(ScheduleImport)
                .filter(ScheduleImport.week_start == week)
                .one()
            )
            self.assertEqual(imp.parser_version, PARSER_VERSION)
            self.assertGreater(imp.person_event_count, 0)
            count = (
                session.query(WeeklyPersonShift)
                .filter(WeeklyPersonShift.week_start == week)
                .count()
            )
            self.assertEqual(count, imp.person_event_count)

    def test_upgrade_creates_aoc_rows(self):
        week_sunday = date(2026, 6, 7)
        week = week_sunday.isoformat()
        _seed_legacy_weekly_staffing(self.db_path, week)
        workbook = self.inbox / "aoc_week.xlsx"
        _write_aoc_workbook(workbook, week_sunday)

        code = _backfill.run_backfill(
            [self.inbox],
            db_path=self.db_path,
            dry_run=False,
            force=False,
            force_week=None,
            upgrade_detail=True,
        )
        self.assertEqual(code, 0)

        with session_scope(self.db_path) as session:
            aoc_rows = (
                session.query(WeeklyManagerShift)
                .filter(
                    WeeklyManagerShift.week_start == week,
                    WeeklyManagerShift.event_type == "aoc",
                )
                .all()
            )
            self.assertEqual(len(aoc_rows), 1)
            self.assertEqual(aoc_rows[0].person_display, "Bowman")
            self.assertEqual(aoc_rows[0].raw_value, "AOC")

    def test_upgrade_reports_missing_workbook(self):
        week = "2026-05-31"
        _seed_legacy_weekly_staffing(self.db_path, week)

        code = _backfill.run_backfill(
            [self.inbox],
            db_path=self.db_path,
            dry_run=True,
            force=False,
            force_week=None,
            upgrade_detail=True,
        )
        self.assertEqual(code, 1)

        planned, errors = _backfill.plan_imports(
            [],
            db_path=self.db_path,
            force=False,
            force_week=None,
            upgrade_detail=True,
        )
        self.assertEqual(planned, [])
        self.assertEqual(len(errors), 1)
        self.assertIn("no workbook found", errors[0])


if __name__ == "__main__":
    unittest.main()
