"""Training codes (EDU, CCT, Neo Sim, Clinical/PER, SM/SIM) are classified
as skip_reason="training", counted toward a weekly total (unlike other
skipped cells), and never flagged as unknown unit codes.
"""

import unittest
from datetime import date

from openpyxl import Workbook
from staffing_tool.schedule_import import (
    ShiftRecord,
    _parse_grid,
    _person_shift_event_type,
    aggregate_week_from_records,
)


def _parse_cell(
    cell_value: str,
    *,
    role: str = "RN",
    manager: bool = False,
    extra_training_codes: frozenset[str] | None = None,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "RN & Medic"
    ws["C1"] = date(2024, 1, 7)
    ws["A4"] = "MGRLAST" if manager else "Smith"
    ws["B4"] = "" if manager else "Jane"
    ws["C4"] = cell_value
    mgr_upper = frozenset({"MGRLAST"}) if manager else frozenset()
    records, issues = _parse_grid(
        ws=ws,
        header_row_idx=1,
        first_row_idx=4,
        last_row_idx=4,
        role=role,
        sheet_label="RN & Medic (RN)",
        week_start_date=date(2024, 1, 7),
        week_end_date=date(2024, 1, 13),
        manager_last_names_upper=mgr_upper,
        extra_training_codes=extra_training_codes,
    )
    return records, issues


class TrainingCellClassificationTests(unittest.TestCase):
    def test_training_codes_not_unknown_unit(self):
        for code in ("EDU", "CCT", "NEO SIM", "CLINICAL/PER", "CLINICAL/ PER"):
            records, issues = _parse_cell(code)
            self.assertEqual(issues, [], code)
            self.assertEqual(len(records), 1, code)
            rec = records[0]
            self.assertEqual(rec.skip_reason, "training", code)
            self.assertFalse(rec.filled, code)

    def test_training_cells_count_toward_aggregates(self):
        records, _ = _parse_cell("EDU")
        self.assertTrue(records[0].included_in_aggregates)

    def test_manager_row_training_excluded_from_aggregates(self):
        records, _ = _parse_cell("EDU", manager=True)
        self.assertEqual(records[0].skip_reason, "training")
        self.assertFalse(records[0].included_in_aggregates)

    def test_training_event_type(self):
        records, _ = _parse_cell("CCT")
        self.assertEqual(_person_shift_event_type(records[0]), "training")


def _training_rec(**kwargs):
    defaults = {
        "date": date(2026, 1, 4),
        "base": "",
        "service_type": "",
        "day_night": "D",
        "role": "RN",
        "filled": False,
        "overtime": False,
        "leave_type": None,
        "source_tab": "RN & Medic (RN)",
        "source_cell": "C4",
        "raw_value": "EDU",
        "person_display": "Smith, Jane",
        "is_manager_row": False,
        "skip_reason": "training",
        "included_in_aggregates": True,
    }
    defaults.update(kwargs)
    return ShiftRecord(**defaults)


class TrainingAggregationTests(unittest.TestCase):
    def test_training_total_counted(self):
        records = [
            _training_rec(),
            _training_rec(raw_value="CCT"),
            _training_rec(raw_value="NEO SIM"),
        ]
        agg = aggregate_week_from_records("2026-01-04", records, ({}, {}, {}, {}))
        self.assertEqual(agg.training_total, 3)

    def test_training_excluded_from_leave_and_staffing(self):
        records = [_training_rec()]
        agg = aggregate_week_from_records("2026-01-04", records, ({}, {}, {}, {}))
        self.assertEqual(agg.filled_day, 0)
        self.assertEqual(agg.filled_night, 0)
        self.assertEqual(agg.leave_at, 0)

    def test_training_excluded_when_not_in_aggregates(self):
        records = [_training_rec(included_in_aggregates=False)]
        agg = aggregate_week_from_records("2026-01-04", records, ({}, {}, {}, {}))
        self.assertEqual(agg.training_total, 0)


class AdminAddedTrainingCodeTests(unittest.TestCase):
    """A Settings > Training codes entry must be recognized without a code
    change -- the whole point of making it admin-editable."""

    def test_unknown_code_flagged_without_extra_training_codes(self):
        records, issues = _parse_cell("ACLS")
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].issue_type, "unknown_unit")
        self.assertEqual(records, [])

    def test_admin_added_code_classified_as_training(self):
        records, issues = _parse_cell("ACLS", extra_training_codes=frozenset({"ACLS"}))
        self.assertEqual(issues, [])
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].skip_reason, "training")
        self.assertTrue(records[0].included_in_aggregates)


if __name__ == "__main__":
    unittest.main()
