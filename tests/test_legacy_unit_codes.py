"""Tests for historical Manchester unit aliases and retired FW shifts."""

import unittest
from datetime import date

from openpyxl import Workbook
from staffing_tool.schedule_import import (
    LEGACY_UNIT_ALIASES,
    RETIRED_UNIT_CODES,
    ShiftRecord,
    _classify_unit,
    _parse_grid,
    _split_unit_suffix,
    aggregate_week_from_records,
    weekly_person_shift_mappings,
)


class LegacyUnitAliasTests(unittest.TestCase):
    def test_legacy_codes_map_to_manchester_d11h(self):
        for legacy in ("D9P", "D9B", "D11B"):
            info = _classify_unit(legacy)
            self.assertIsNotNone(info, legacy)
            base, service_type, day_night = info  # type: ignore[misc]
            self.assertEqual(base, "Manchester")
            self.assertEqual(service_type, "RW")
            self.assertEqual(day_night, "D")

    def test_ot_suffix_on_legacy_codes(self):
        for raw, expected_core, expect_ot in (
            ("D9PC", "D9P", True),
            ("D9BC", "D9B", True),
            ("D11BC", "D11B", True),
            ("D9P C", "D9P", True),
        ):
            core, is_ot, _is_dual = _split_unit_suffix(raw)
            self.assertEqual(core, expected_core, raw)
            self.assertEqual(is_ot, expect_ot, raw)
            info = _classify_unit(core)
            self.assertIsNotNone(info, raw)
            self.assertEqual(info[0], "Manchester")  # type: ignore[index]

    def test_fw_is_retired(self):
        self.assertIn("FW", RETIRED_UNIT_CODES)
        self.assertIsNone(_classify_unit("FW"))
        core, is_ot, _ = _split_unit_suffix("FWC")
        self.assertEqual(core, "FW")
        self.assertTrue(is_ot)
        self.assertIsNone(_classify_unit(core))


class LegacyUnitGridParseTests(unittest.TestCase):
    def _parse_cell(self, cell_value: str, role: str = "RN") -> tuple[list, list]:
        wb = Workbook()
        ws = wb.active
        ws.title = "RN & Medic"
        ws["C1"] = date(2024, 1, 7)
        ws["A4"] = "Smith"
        ws["B4"] = "Jane"
        ws["C4"] = cell_value
        records, issues = _parse_grid(
            ws=ws,
            header_row_idx=1,
            first_row_idx=4,
            last_row_idx=4,
            role=role,
            sheet_label="RN & Medic (RN)",
            week_start_date=date(2024, 1, 7),
            week_end_date=date(2024, 1, 13),
        )
        return records, issues

    def test_legacy_manchester_cells_staffed_with_canonical_unit(self):
        for raw in ("D9P", "D9B", "D11B", "D9PC"):
            records, issues = self._parse_cell(raw)
            self.assertEqual(issues, [], raw)
            staffed = [r for r in records if r.filled]
            self.assertEqual(len(staffed), 1, raw)
            rec = staffed[0]
            self.assertEqual(rec.base, "Manchester")
            self.assertEqual(rec.unit_code, "D11H")
            self.assertEqual(rec.raw_value, raw)
            if raw.endswith("C"):
                self.assertTrue(rec.overtime)

    def test_fw_skipped_not_staffed(self):
        records, issues = self._parse_cell("FW")
        self.assertEqual(issues, [])
        staffed = [r for r in records if r.filled]
        skipped = [r for r in records if r.skip_reason == "retired_unit"]
        self.assertEqual(staffed, [])
        self.assertEqual(len(skipped), 1)
        self.assertEqual(skipped[0].raw_value, "FW")
        self.assertFalse(skipped[0].included_in_aggregates)

    def test_fw_ot_variant_skipped(self):
        records, issues = self._parse_cell("FWC")
        self.assertEqual(issues, [])
        self.assertEqual(len([r for r in records if r.filled]), 0)
        skipped = [r for r in records if r.skip_reason == "retired_unit"]
        self.assertEqual(len(skipped), 1)
        self.assertEqual(skipped[0].raw_value, "FWC")

    def test_legacy_manchester_in_aggregates(self):
        records, _ = self._parse_cell("D9P")
        rn = records[0]
        medic = ShiftRecord(
            date=rn.date,
            base="Manchester",
            service_type="RW",
            day_night="D",
            role="MEDIC",
            filled=True,
            overtime=False,
            leave_type=None,
            source_tab=rn.source_tab,
            source_cell="D4",
            raw_value="D9P",
            unit_code="D11H",
            person_display="Jones, Bob",
        )
        agg = aggregate_week_from_records("2024-01-07", [rn, medic], None)
        self.assertEqual(agg.filled_day, 1)
        self.assertEqual(agg.base_rw_staffed.get("Manchester", 0), 1)

    def test_fw_excluded_from_aggregates(self):
        records, _ = self._parse_cell("FW")
        agg = aggregate_week_from_records("2024-01-07", records, None)
        self.assertEqual(agg.filled_day, 0)
        rows = weekly_person_shift_mappings("2024-01-07", records)
        self.assertEqual(len([r for r in rows if r["event_type"] == "staffed"]), 0)
        retired = [r for r in rows if r["skip_reason"] == "retired_unit"]
        self.assertGreaterEqual(len(retired), 1)
        self.assertTrue(all(r["included_in_aggregates"] == 0 for r in retired))


class LegacyAliasConfigTests(unittest.TestCase):
    def test_all_legacy_aliases_target_d11h(self):
        for _legacy, canonical in LEGACY_UNIT_ALIASES.items():
            self.assertEqual(canonical, "D11H")


if __name__ == "__main__":
    unittest.main()
