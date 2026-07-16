"""End-to-end tests for the Excel builders: weekly board pack + monthly report.

These drive export_board_pack / export_monthly_report against a seeded
temporary database and assert the workbook structure plus the non-negotiable
constraints from docs/report-generator-spec.md (no formulas, no Excel
conditional formatting, no Green/Yellow/Red words in cells).
"""

import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from staffing_tool.db import get_engine, init_db, session_scope
from staffing_tool.models import WeeklyPersonShift, WeeklyStaffing
from staffing_tool.monthly_report import export_monthly_report
from staffing_tool.report import export_board_pack

RAG_WORDS = {"Green", "Yellow", "Red"}


def _spec_violations(path: str) -> list[tuple]:
    """Scan every sheet for spec violations; empty list means compliant."""
    wb = load_workbook(path)
    bad: list[tuple] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if cell.value.startswith("="):
                        bad.append(("formula", ws.title, cell.coordinate))
                    if cell.value.strip() in RAG_WORDS:
                        bad.append(("rag_word", ws.title, cell.coordinate))
        if len(list(ws.conditional_formatting)):
            bad.append(("conditional_formatting", ws.title))
    return bad


class BoardPackExcelTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.db_path = os.path.join(self.tmp.name, "test.db")
        self.out_dir = os.path.join(self.tmp.name, "output")
        init_db(self.db_path)
        with session_scope(self.db_path) as session:
            for i, ws in enumerate(
                ["2026-04-19", "2026-04-26", "2026-05-03", "2026-05-10"]
            ):
                session.add(
                    WeeklyStaffing(
                        week_start=ws,
                        filled_day=48 + i,
                        filled_night=20,
                        ot_rn=2,
                        ot_medic=1,
                        ot_emt=0,
                        leave_at=3,
                        leave_sick=1,
                    )
                )
            session.commit()
            for i in range(12):
                session.add(
                    WeeklyPersonShift(
                        week_start="2026-05-10",
                        person_display=f"Person {i}",
                        shift_date="2026-05-11",
                        role="RN" if i < 7 else "MEDIC",
                        event_type="staffed" if i % 4 else "ot",
                        included_in_aggregates=1,
                    )
                )
            session.commit()

    def tearDown(self):
        get_engine(self.db_path).dispose()
        self.tmp.cleanup()

    def test_weekly_board_pack_structure(self):
        path = export_board_pack(self.db_path, "2026-05-10", output_dir=self.out_dir)
        self.assertTrue(os.path.isfile(path))
        self.assertIn("Weekly_staffing_summary_2026-05-10_to_2026-05-16", path)
        wb = load_workbook(path)
        self.assertEqual(
            wb.sheetnames,
            ["Board_Summary", "Weekly_Detail", "Trend_12_Weeks", "Data_Dump"],
        )
        summary_labels = [wb["Board_Summary"].cell(r, 1).value for r in range(1, 30)]
        for label in ("Staffing Rate", "System RW Coverage %", "System GR Coverage %"):
            self.assertIn(label, summary_labels)

    def test_weekly_board_pack_spec_constraints(self):
        path = export_board_pack(self.db_path, "2026-05-10", output_dir=self.out_dir)
        self.assertEqual(_spec_violations(path), [])

    def test_weekly_board_pack_unknown_week_raises(self):
        with self.assertRaises(ValueError):
            export_board_pack(self.db_path, "2001-01-07", output_dir=self.out_dir)

    def test_monthly_report_structure_and_spec(self):
        path = export_monthly_report(
            self.db_path, "2026-05-01", "2026-05-31", output_dir=self.out_dir
        )
        self.assertTrue(os.path.isfile(path))
        wb = load_workbook(path)
        self.assertIn("Monthly_Summary", wb.sheetnames)
        self.assertIn("By_Week", wb.sheetnames)
        self.assertEqual(_spec_violations(path), [])


if __name__ == "__main__":
    unittest.main()
