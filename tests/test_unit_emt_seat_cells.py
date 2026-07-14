""" "<UNIT> EMT" cells on RN/Medic rows are cross-references, not shifts.

A manager or medic covering a ground unit's EMT seat is written as e.g.
"MG EMT" on their RN & Medic row, while the seat assignment itself lives on
the EMT sheet (and OPS View). Parsing the cross-reference as a staffed
shift would double-count the person, and before this handling it surfaced
as a bogus "unknown unit code" review issue.
"""

import unittest
from datetime import date

from openpyxl import Workbook
from staffing_tool.schedule_import import _parse_grid


def _parse_cell(cell_value: str, *, role: str = "MEDIC"):
    wb = Workbook()
    ws = wb.active
    ws.title = "RN & Medic"
    ws["C1"] = date(2024, 1, 7)
    ws["A4"] = "Holst"
    ws["C4"] = cell_value
    records, issues = _parse_grid(
        ws=ws,
        header_row_idx=1,
        first_row_idx=4,
        last_row_idx=4,
        role=role,
        sheet_label=f"RN & Medic ({role.title()})",
        week_start_date=date(2024, 1, 7),
        week_end_date=date(2024, 1, 13),
    )
    return records, issues


class UnitEmtSeatCellTests(unittest.TestCase):
    def test_unit_emt_on_medic_row_skipped_not_unknown(self):
        records, issues = _parse_cell("MG EMT")
        self.assertEqual(issues, [])
        self.assertEqual(len(records), 1)
        rec = records[0]
        self.assertEqual(rec.skip_reason, "admin")
        self.assertFalse(rec.filled)
        self.assertFalse(rec.included_in_aggregates)

    def test_unresolvable_prefix_still_flagged_unknown(self):
        records, issues = _parse_cell("ZZZ EMT")
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].issue_type, "unknown_unit")


if __name__ == "__main__":
    unittest.main()
