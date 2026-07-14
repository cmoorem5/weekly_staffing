"""Integration test: importing a workbook with a training cell persists
WeeklyStaffing.training_shifts (schedule_apply.py wiring)."""

import os
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from staffing_tool.db import get_engine, init_db, session_scope
from staffing_tool.models import TrainingCode, WeeklyStaffing
from staffing_tool.schedule_apply import apply_schedule_workbook

WEEK_START = "2024-01-07"


def _build_workbook(path: Path, *, extra_code: str | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RN & Medic"
    ws["C1"] = date(2024, 1, 7)
    ws["D1"] = date(2024, 1, 8)
    ws["E1"] = date(2024, 1, 9)

    # RN block (rows 4-50): one staffed RW crew slot + one training cell.
    ws["A4"] = "Smith"
    ws["B4"] = "Jane"
    ws["C4"] = "D7B"
    ws["A5"] = "Doe"
    ws["B5"] = "Roe"
    ws["D5"] = "EDU"
    if extra_code:
        ws["A6"] = "Fox"
        ws["B6"] = "Amy"
        ws["E6"] = extra_code

    # Medic block (rows 52-100): pairs with the RN staffed slot above.
    ws["A52"] = "Jones"
    ws["B52"] = "Bob"
    ws["C52"] = "D7B"

    wb.save(str(path))


class ScheduleApplyTrainingTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.db_path = os.path.join(self.tmp.name, "test.db")
        init_db(self.db_path)
        self.upload_path = Path(self.tmp.name) / "schedule.xlsx"
        _build_workbook(self.upload_path)

    def tearDown(self):
        get_engine(self.db_path).dispose()
        self.tmp.cleanup()

    def test_training_shifts_persisted_on_import(self):
        with session_scope(self.db_path) as session:
            result, err = apply_schedule_workbook(
                session,
                week_start=WEEK_START,
                upload_path=str(self.upload_path),
                source_filename="schedule.xlsx",
            )
        self.assertIsNone(err)
        assert result is not None

        with session_scope(self.db_path) as session:
            row = (
                session.query(WeeklyStaffing)
                .filter(WeeklyStaffing.week_start == WEEK_START)
                .first()
            )
            self.assertEqual(row.training_shifts, 1)
            self.assertEqual(row.filled_day, 1)

    def test_admin_added_training_code_applied_automatically(self):
        # An "ACLS" cell isn't in the built-in training list -- it must be
        # picked up automatically from the DB, with no code change and no
        # extra_training_codes passed explicitly by the caller.
        upload_with_extra = Path(self.tmp.name) / "schedule_extra.xlsx"
        _build_workbook(upload_with_extra, extra_code="ACLS")

        with session_scope(self.db_path) as session:
            session.add(TrainingCode(code="ACLS", created_at="2026-01-01T00:00:00Z"))

        with session_scope(self.db_path) as session:
            result, err = apply_schedule_workbook(
                session,
                week_start=WEEK_START,
                upload_path=str(upload_with_extra),
                source_filename="schedule_extra.xlsx",
            )
        self.assertIsNone(err)
        assert result is not None

        with session_scope(self.db_path) as session:
            row = (
                session.query(WeeklyStaffing)
                .filter(WeeklyStaffing.week_start == WEEK_START)
                .first()
            )
            # EDU (built-in) + ACLS (admin-added) = 2.
            self.assertEqual(row.training_shifts, 2)


if __name__ == "__main__":
    unittest.main()
