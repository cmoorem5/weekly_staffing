"""OPS View vehicle blocks must survive non-unit section labels.

The OPS View sheet lists vehicle blocks (unit code in column A, role rows
below) but also non-vehicle sections whose labels sit in the same column:
FLOAT, OPEN, DAY/NIGHT Extra, footer text. A non-unit label used to reset
the in-progress block's role rows while leaving ``current_unit`` set, so
the unit immediately before the section (PG before FLOAT in the July 2026
workbook) had its RN/Medic/EMT rows thrown away and replaced by the
section's rows -- erasing that base's staffed unit-days from Base
Coverage.
"""

import unittest
from datetime import datetime

from openpyxl import Workbook
from staffing_tool.schedule_import import (
    _ops_vehicle_blocks,
    _parse_ops_view_worksheet,
)

WEEK = "2026-07-12"


def _build_ops_ws():
    """Minimal OPS View: PG block, then FLOAT section, then N7B block."""
    wb = Workbook()
    ws = wb.active
    ws.title = "OPS View"
    # Date header row (columns C.. = days of the week).
    start = datetime.strptime(WEEK, "%Y-%m-%d")
    for i in range(7):
        ws.cell(row=1, column=3 + i, value=start.replace(day=start.day + i))

    def block(row, unit, roles):
        ws.cell(row=row, column=1, value=unit)
        for offset, (role, names) in enumerate(roles):
            ws.cell(row=row + offset, column=2, value=role)
            for i, name in enumerate(names):
                if name:
                    ws.cell(row=row + offset, column=3 + i, value=name)
        return row + len(roles)

    # PG staffed (RN+Medic+EMT) on Sun and Wed only.
    r = block(
        4,
        "PG",
        [
            ("RN", ["McGrath", None, None, "Eastman", None, None, None]),
            ("Medic", ["Sacco", None, None, "Sturtevant", None, None, None]),
            ("EMT", ["Finn", "Quinn", "Quinn", "Finn", "Finn", "Quinn", "Quinn"]),
        ],
    )
    # FLOAT section: role rows that must NOT be attributed to PG. Its Medic
    # is staffed on Tue, a day PG has no RN -- leakage would add a bogus day.
    r = block(
        r,
        "FLOAT",
        [
            ("RN", [None] * 7),
            ("Medic", [None, None, "Fielding", None, None, None, None]),
        ],
    )
    # A later real unit so the PG block gets flushed by a unit boundary too.
    block(
        r,
        "N7B",
        [
            ("RN", ["Johnson"] * 7),
            ("Medic", ["McKinnon"] * 7),
            ("PIC", ["Hurst"] * 7),
        ],
    )
    return ws


class OpsVehicleBlocksTests(unittest.TestCase):
    def test_float_section_does_not_replace_previous_unit_rows(self):
        ws = _build_ops_ws()
        blocks = {
            unit: role_rows for unit, _start, role_rows in _ops_vehicle_blocks(ws)
        }
        self.assertIn("PG", blocks)
        self.assertEqual(blocks["PG"].get("RN"), [4])
        self.assertEqual(blocks["PG"].get("Medic"), [5])
        self.assertEqual(blocks["PG"].get("EMT"), [6])
        self.assertNotIn("FLOAT", blocks)
        self.assertIn("N7B", blocks)

    def test_base_coverage_counts_unit_followed_by_float(self):
        ws = _build_ops_ws()
        rw_day, rw_night, gr_day, gr_night = _parse_ops_view_worksheet(ws, WEEK)
        # PG = Plymouth GR day; staffed Sun + Wed. FLOAT's Tue Medic must not
        # leak in, and N7B still counts normally.
        self.assertEqual(gr_day.get("Plymouth"), 2)
        self.assertEqual(rw_night.get("Bedford"), 7)
        self.assertEqual(gr_night, {})


if __name__ == "__main__":
    unittest.main()
