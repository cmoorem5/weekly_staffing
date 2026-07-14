"""OPEN/EXTRA summary rows must be skipped by label, not by a fixed row
number.

The RN & Medic (and EMT) grids end with an "OPEN" row (unfilled units per
day) and an "EXTRA" row (float staff names per day) below the real roster.
Those rows used to be located by a hardcoded row-number set
(RN_MEDIC_SKIP_SCHEDULE_ROWS = {45, 46, 91, 92, ...}). Adding staff pushes
every row below them down the sheet, so the hardcoded numbers silently go
stale and the OPEN/EXTRA rows' contents (unit-code lists, staff names) leak
into schedule parsing as if they were real shift codes -- surfacing as
bogus "unknown unit code" review entries.
"""

import unittest
from datetime import date

from openpyxl import Workbook
from staffing_tool.schedule_import import _parse_grid


def _build_ws(*, roster_rows: int, open_row: int, extra_row: int):
    """RN & Medic sheet: `roster_rows` real people, then OPEN, then EXTRA."""
    wb = Workbook()
    ws = wb.active
    ws.title = "RN & Medic"
    ws["C1"] = date(2024, 1, 7)

    first_row = 4
    for i in range(roster_rows):
        row = first_row + i
        ws.cell(row=row, column=1, value=f"Staff{i}")
        ws.cell(row=row, column=2, value="A")
        ws.cell(row=row, column=3, value="D7B")

    ws.cell(row=open_row, column=1, value="OPEN")
    ws.cell(row=open_row, column=3, value="MG, LG, PG")

    ws.cell(row=extra_row, column=1, value="EXTRA")
    ws.cell(row=extra_row, column=3, value="Phillips R.")

    return ws, first_row


class OpenExtraRowSkipTests(unittest.TestCase):
    def _run(self, *, roster_rows: int, open_row: int, extra_row: int):
        ws, first_row = _build_ws(
            roster_rows=roster_rows, open_row=open_row, extra_row=extra_row
        )
        records, issues = _parse_grid(
            ws=ws,
            header_row_idx=1,
            first_row_idx=first_row,
            last_row_idx=extra_row,
            role="RN",
            sheet_label="RN & Medic (RN)",
            week_start_date=date(2024, 1, 7),
            week_end_date=date(2024, 1, 13),
        )
        self.assertEqual(issues, [])
        return records

    def test_open_and_extra_rows_skipped_at_original_positions(self):
        # Mirrors the historical hardcoded rows (45, 46) relative offset.
        records = self._run(roster_rows=6, open_row=10, extra_row=11)
        open_recs = [r for r in records if r.excel_row == 10]
        extra_recs = [r for r in records if r.excel_row == 11]
        self.assertTrue(open_recs and all(r.skip_reason for r in open_recs))
        self.assertTrue(extra_recs and all(r.skip_reason for r in extra_recs))
        # Never surfaced as unfilled shifts / staffed unit codes.
        self.assertFalse(any(r.filled for r in open_recs + extra_recs))

    def test_open_and_extra_rows_skipped_after_roster_growth(self):
        # Two more staff rows inserted above OPEN/EXTRA -- everything shifts
        # down by 2, same as adding personnel to a real workbook. The
        # hardcoded-row-number approach would silently stop matching here.
        records = self._run(roster_rows=8, open_row=12, extra_row=13)
        open_recs = [r for r in records if r.excel_row == 12]
        extra_recs = [r for r in records if r.excel_row == 13]
        self.assertTrue(open_recs and all(r.skip_reason for r in open_recs))
        self.assertTrue(extra_recs and all(r.skip_reason for r in extra_recs))
        self.assertFalse(any(r.filled for r in open_recs + extra_recs))

    def test_real_roster_rows_above_open_are_not_skipped(self):
        records = self._run(roster_rows=6, open_row=10, extra_row=11)
        roster_recs = [r for r in records if r.excel_row < 10]
        self.assertTrue(roster_recs)
        self.assertTrue(any(r.filled for r in roster_recs))


if __name__ == "__main__":
    unittest.main()
