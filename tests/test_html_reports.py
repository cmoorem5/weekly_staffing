"""Smoke tests for the monthly and quarterly HTML report exports."""

import os
import tempfile
import unittest
from pathlib import Path

from staffing_tool.db import get_engine, init_db, session_scope
from staffing_tool.metrics import compute_role_fill
from staffing_tool.models import WeeklyPersonShift, WeeklyStaffing
from staffing_tool.monthly_html_report import (
    export_monthly_report_html,
    load_monthly_board_data,
)
from staffing_tool.quarterly_pdf_report import export_quarterly_staffing_html
from staffing_tool.weekly_pdf_report import export_weekly_staffing_html


class HtmlReportExportTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.db_path = os.path.join(self.tmp.name, "test.db")
        self.out_dir = os.path.join(self.tmp.name, "output")
        init_db(self.db_path)
        with session_scope(self.db_path) as session:
            # Two months so the December board report has a prior period.
            for ws, filled in [
                ("2025-11-02", 48),
                ("2025-11-09", 47),
                ("2025-12-07", 50),
                ("2025-12-14", 52),
            ]:
                session.add(
                    WeeklyStaffing(
                        week_start=ws,
                        filled_day=filled,
                        filled_night=20,
                        ot_rn=2,
                        ot_medic=1,
                        ot_emt=0,
                        leave_at=3,
                        leave_sick=1,
                    )
                )
            session.commit()
            # Person shifts for role fill (worked = staffed + ot).
            for i in range(20):
                session.add(
                    WeeklyPersonShift(
                        week_start="2025-12-07",
                        person_display=f"Person {i}",
                        shift_date="2025-12-08",
                        role="RN" if i < 12 else "MEDIC",
                        event_type="staffed" if i % 4 else "ot",
                        included_in_aggregates=1,
                    )
                )
            session.commit()

    def tearDown(self):
        get_engine(self.db_path).dispose()
        self.tmp.cleanup()

    def test_monthly_board_data_includes_prior_period(self):
        data = load_monthly_board_data(self.db_path, "2025-12-01", "2025-12-31")
        self.assertEqual(data.weeks_count, 2)
        self.assertIsNotNone(data.prior_rollups)
        self.assertEqual(len(data.weekly_detail), 2)

    def test_monthly_html_export(self):
        path = export_monthly_report_html(
            self.db_path, "2025-12-01", "2025-12-31", self.out_dir
        )
        self.assertTrue(os.path.isfile(path))
        html = Path(path).read_text(encoding="utf-8")
        self.assertIn("MONTHLY STAFFING REPORT", html)
        self.assertIn("KEY PERFORMANCE INDICATORS", html)
        self.assertIn("data:image/png;base64,", html)  # embedded charts
        self.assertIn("pts</span>", html)  # prior-period delta rendered

    def test_monthly_html_rejects_empty_range(self):
        with self.assertRaises(ValueError):
            export_monthly_report_html(
                self.db_path, "2020-01-01", "2020-01-31", self.out_dir
            )

    def test_compute_role_fill_counts_staffed_and_ot(self):
        with session_scope(self.db_path) as session:
            fill = {rf.role: rf for rf in compute_role_fill(session, ["2025-12-07"])}
        self.assertEqual(fill["RN"].worked, 12)
        self.assertEqual(fill["RN"].capacity, 84)
        self.assertEqual(fill["MEDIC"].worked, 8)
        self.assertEqual(fill["EMT"].worked, 0)

    def test_compute_role_fill_excludes_extra_bedford_units(self):
        with session_scope(self.db_path) as session:
            session.add(
                WeeklyPersonShift(
                    week_start="2025-12-07",
                    person_display="Required Line",
                    shift_date="2025-12-08",
                    role="EMT",
                    event_type="staffed",
                    unit_code="NG",
                    included_in_aggregates=1,
                )
            )
            for unit_code in ("GR2", "NG2"):
                session.add(
                    WeeklyPersonShift(
                        week_start="2025-12-07",
                        person_display="Extra Ambulance",
                        shift_date="2025-12-08",
                        role="EMT",
                        event_type="staffed",
                        unit_code=unit_code,
                        included_in_aggregates=1,
                    )
                )
            session.commit()
            fill = {rf.role: rf for rf in compute_role_fill(session, ["2025-12-07"])}
        self.assertEqual(fill["EMT"].worked, 1)

    def test_compute_role_fill_counts_emt_pair_cell_once(self):
        # EMT partner rows list two people for one grid cell;
        # weekly_person_shift_mappings writes one row per person. The seat
        # was worked once, so role fill must count the cell, not the rows.
        with session_scope(self.db_path) as session:
            for person in ("Chatigny, Aaron", "Belliveau, Kelly"):
                session.add(
                    WeeklyPersonShift(
                        week_start="2025-12-07",
                        person_display=person,
                        shift_date="2025-12-08",
                        role="EMT",
                        event_type="staffed",
                        unit_code="GR",
                        source_tab="EMT",
                        source_cell="C3",
                        included_in_aggregates=1,
                    )
                )
            session.commit()
            fill = {rf.role: rf for rf in compute_role_fill(session, ["2025-12-07"])}
        self.assertEqual(fill["EMT"].worked, 1)

    def test_compute_role_fill_pair_cell_once_end_to_end(self):
        from datetime import date

        from openpyxl import Workbook
        from staffing_tool.schedule_import import (
            _parse_grid,
            weekly_person_shift_mappings,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "EMT"
        ws["C1"] = date(2025, 12, 7)
        ws["D1"] = date(2025, 12, 8)
        ws["A3"] = "1-Chatigny, Aaron"
        ws["B3"] = "Belliveau, Kelly"
        ws["C3"] = "GR"
        ws["D3"] = "NG"
        records, issues = _parse_grid(
            ws=ws,
            header_row_idx=1,
            first_row_idx=3,
            last_row_idx=3,
            role="EMT",
            sheet_label="EMT",
            week_start_date=date(2025, 12, 7),
            week_end_date=date(2025, 12, 13),
        )
        self.assertEqual(issues, [])
        rows = weekly_person_shift_mappings("2025-12-07", records)
        # Two people per staffed cell → four person rows for two seat-days.
        self.assertEqual(len(rows), 4)
        with session_scope(self.db_path) as session:
            session.bulk_insert_mappings(WeeklyPersonShift, rows)
            session.commit()
            fill = {rf.role: rf for rf in compute_role_fill(session, ["2025-12-07"])}
        self.assertEqual(fill["EMT"].worked, 2)

    def test_monthly_html_has_day_night_and_role_fill(self):
        path = export_monthly_report_html(
            self.db_path, "2025-12-01", "2025-12-31", self.out_dir
        )
        html = Path(path).read_text(encoding="utf-8")
        self.assertIn("Day Fill", html)
        self.assertIn("Night Fill", html)
        self.assertIn("FILL RATE BY ROLE", html)

    def test_weekly_html_has_day_night_and_role_fill(self):
        path = export_weekly_staffing_html(self.db_path, "2025-12-07", self.out_dir)
        html = Path(path).read_text(encoding="utf-8")
        self.assertIn("DAY / NIGHT", html)
        self.assertIn("Day (56 required)", html)
        self.assertIn("RN (Flight Nurse)", html)

    def test_quarterly_html_export(self):
        path = export_quarterly_staffing_html(self.db_path, 2026, 2, self.out_dir)
        self.assertTrue(os.path.isfile(path))
        self.assertTrue(path.endswith(".html"))
        html = Path(path).read_text(encoding="utf-8")
        self.assertIn("QUARTERLY STAFFING REPORT", html)
        self.assertIn("FY2026 Q2", html)
        self.assertIn("PERIOD VOLUMES BY ROLE", html)
        self.assertIn("data:image/png;base64,", html)


if __name__ == "__main__":
    unittest.main()
